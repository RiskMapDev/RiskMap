"""Тесты слоя 8.3 «Бюджетные риски».

Разделены на две части. Первая проверяет методику на синтетических значениях
и работает без книги. Вторая (`@pytest.mark.golden`) сверяет независимый
пересчёт с контрольными значениями книги: расчётный лист 8.3 содержит формулы,
и сойтись обязано всё, до последнего знака.

Что здесь намеренно зафиксировано как факт, а не как проблема:

* три индикатора (R03, R10, R12) дают ноль на всех 240 строках — 14 % веса
  модели не работает;
* критическое переопределение — это пол в 75 баллов, поэтому единственная
  критическая строка имеет балл ровно 75,0, и сравнивать приходится на
  равенство;
* полнота данных в этом слое на уровень риска не влияет.
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest

from app.core.config import get_settings
from app.importers.budget_8_3 import (
    EXPECTED_MONTHLY_ROWS,
    EXPECTED_PERIODS,
    EXPECTED_RAW_ROWS,
    EXPECTED_SHEETS,
    EXPECTED_TERRITORIES,
    SOURCE_SPELLINGS,
    BudgetMonthlyRow,
    check_parameters_match_model,
    evaluate_rows,
    iter_raw_facts,
    load_monthly_rows,
    load_parameters,
    normalize_region_name,
    parse_period,
    resolve_workbook,
    territory_aliases,
)
from app.risk.core import IndicatorDirection, RiskLevel
from app.risk.layers.budget import (
    BANDS,
    BUDGET_8_3,
    DEGENERATE_INDICATORS,
    INDICATORS,
    OVERRIDE_FLOOR,
    BudgetRawIndicators,
    BudgetRowInputs,
    NormalizationBand,
    data_completeness,
    evaluate_row,
    override_triggered,
    rank_within_month,
)

# --- Контрольные значения книги (docs/audit/02-…, п. 6) ----------------------

BOOK_SCORE_MIN = 4.006757335657181
BOOK_SCORE_MAX = 75.0
BOOK_SCORE_SUM = 4516.173271
BOOK_SCORE_AVG = 18.817389
BOOK_LEVEL_COUNTS = {"Низкий": 187, "Средний": 52, "Высокий": 0, "Критический": 1}
BOOK_OVERRIDE_ROWS = 1
BOOK_INCOMPLETE_ROWS = 32
BOOK_REFERENCE_SCORES = {
    "REG-001-01": 23.506041560605112,
    "REG-001-02": 21.46732812343851,
    "REG-019-01": 75.0,
    "REG-002-12": 4.006757335657181,
    "REG-020-01": 42.968469238646584,
}


# --- Фикстуры ----------------------------------------------------------------


@pytest.fixture(scope="module")
def workbook_path() -> Path:
    """Путь к книге 8.3 или пропуск, если каталог источников недоступен."""
    source_dir = get_settings().source_data_dir
    if not source_dir.is_dir():
        pytest.skip(f"Каталог источников недоступен: {source_dir}")
    try:
        return resolve_workbook(source_dir)
    except FileNotFoundError as error:
        pytest.skip(str(error))


@pytest.fixture(scope="module")
def monthly_rows(workbook_path: Path) -> list[BudgetMonthlyRow]:
    return load_monthly_rows(workbook_path)


# --- Нормировка --------------------------------------------------------------


class TestNormirovka:
    """Линейная интерполяция между порогом «без риска» и критическим."""

    def test_porog_bez_riska_daet_nol(self) -> None:
        band = BANDS["R06"]  # HIGH: 0,02 → 0 баллов, 0,12 → 100
        assert band.score(0.02) == pytest.approx(0.0)

    def test_kriticheskiy_porog_daet_sto(self) -> None:
        assert BANDS["R06"].score(0.12) == pytest.approx(100.0)

    def test_seredina_daet_polovinu(self) -> None:
        assert BANDS["R06"].score(0.07) == pytest.approx(50.0)

    def test_za_predelami_otsekaetsya(self) -> None:
        band = BANDS["R06"]
        assert band.score(-5.0) == pytest.approx(0.0)
        assert band.score(99.0) == pytest.approx(100.0)

    def test_napravlenie_low_rastet_pri_padenii(self) -> None:
        """R01 «недобор доходов»: 0,98 → 0 баллов, 0,80 → 100."""
        band = BANDS["R01"]
        assert band.direction is IndicatorDirection.LOWER_IS_RISKIER
        assert band.score(0.98) == pytest.approx(0.0)
        assert band.score(0.80) == pytest.approx(100.0)
        assert band.score(0.89) == pytest.approx(50.0)
        # Исполнение выше плана риска не создаёт.
        assert band.score(1.30) == pytest.approx(0.0)

    def test_sovpavshie_porogi_ne_delyat_na_nol(self) -> None:
        """Пороги совпали — интерполировать не по чему, но падать нельзя."""
        low = NormalizationBand(1.0, 1.0, IndicatorDirection.LOWER_IS_RISKIER)
        assert low.score(0.5) == pytest.approx(100.0)
        assert low.score(1.5) == pytest.approx(0.0)

        high = NormalizationBand(1.0, 1.0, IndicatorDirection.HIGHER_IS_RISKIER)
        assert high.score(1.5) == pytest.approx(100.0)
        assert high.score(0.5) == pytest.approx(0.0)


# --- Модель ------------------------------------------------------------------


class TestModel:
    def test_summa_vesov_ravna_sto(self) -> None:
        """Правило управления моделью с листа «Параметры»."""
        assert sum(spec.weight for spec in INDICATORS) == pytest.approx(100.0)

    def test_pyatnadtsat_indikatorov(self) -> None:
        assert len(INDICATORS) == 15

    def test_u_kazhdogo_indikatora_est_porogi(self) -> None:
        assert {spec.code for spec in INDICATORS} == set(BANDS)

    def test_porogi_urovney_0_25_50_75(self) -> None:
        assert BUDGET_8_3.level_for(0.0) is RiskLevel.LOW
        assert BUDGET_8_3.level_for(24.999) is RiskLevel.LOW
        assert BUDGET_8_3.level_for(25.0) is RiskLevel.MEDIUM
        assert BUDGET_8_3.level_for(50.0) is RiskLevel.HIGH
        assert BUDGET_8_3.level_for(75.0) is RiskLevel.CRITICAL

    def test_serogo_urovnya_v_metodike_net(self) -> None:
        """В отличие от слоя 8.4, неполнота здесь не даёт отдельного статуса."""
        assert BUDGET_8_3.min_completeness is None


# --- Переопределение ---------------------------------------------------------


def _row(
    *,
    r01: float = 1.0,
    r02: float = 1.0,
    balance: float = 1_000.0,
    flags: float = 0.0,
    territory_id: str = "REG-001",
    month: int = 1,
    **overrides: float,
) -> BudgetRowInputs:
    """Строка со «здоровыми» значениями, кроме явно переданных."""
    defaults: dict[str, float] = {
        "r04_intensivnost_utochneniy": 0.0,
        "r05_oshibka_profilya": 0.0,
        "r06_otklonenie_saldo": 0.0,
        "r07_kassovyy_bufer": 5.0,
        "r08_izbytochnye_ostatki": 0.0,
        "r09_davlenie_ostatka": 0.0,
        "r10_otstavanie_obyazatelstv": 0.0,
        "r11_neoplachennye_obyazatelstva": 0.0,
        "r12_shirina_nedoispolneniya": 0.0,
        "r13_hhi": 0.0,
        "r14_finansovye_operatsii": 0.0,
    }
    defaults.update(overrides)
    raw = BudgetRawIndicators(
        r01_dohody_ispolnenie=r01,
        r02_zatraty_ispolnenie=r02,
        r15_flagi_kachestva=flags,
        **defaults,  # type: ignore[arg-type]
    )
    return BudgetRowInputs(
        territory_id=territory_id,
        territory_name="Абайская область",
        month=month,
        period=f"{month:02d}.2025",
        raw=raw,
        closing_balance=balance,
    )


class TestKriticheskoePereopredelenie:
    """`MAX(взвешенная сумма; 75)` — это пол балла, а не замена уровня."""

    @pytest.mark.parametrize(
        ("opisanie", "stroka"),
        [
            ("исполнение доходов ниже 70 %", _row(r01=0.69)),
            ("исполнение расходов ниже 60 %", _row(r02=0.59)),
            ("отрицательный конечный остаток", _row(balance=-1.0)),
            ("три флага качества данных", _row(flags=3.0)),
        ],
    )
    def test_kazhdoe_uslovie_podnimaet_ball_do_75(
        self, opisanie: str, stroka: BudgetRowInputs
    ) -> None:
        assert override_triggered(stroka), opisanie
        result = evaluate_row(stroka)
        assert result.score >= OVERRIDE_FLOOR
        assert result.level is RiskLevel.CRITICAL

    def test_granichnye_znacheniya_ne_srabatyvayut(self) -> None:
        """Ровно 0,7 и ровно 0,6 — ещё не переопределение: сравнение строгое."""
        assert not override_triggered(_row(r01=0.7))
        assert not override_triggered(_row(r02=0.6))
        assert not override_triggered(_row(balance=0.0))
        assert not override_triggered(_row(flags=2.0))

    def test_pol_daet_rovno_75_a_ne_100(self) -> None:
        """Ключевое отличие пола от замены уровня.

        Замена уровня дала бы «критический» при любом балле. Пол поднимает
        балл ровно до 75 — и уровень получается критическим уже по порогу.
        Единственная критическая строка книги имеет балл ровно 75,0.
        """
        result = evaluate_row(_row(r01=0.5))
        assert result.score == pytest.approx(OVERRIDE_FLOOR)
        assert result.level is RiskLevel.CRITICAL

    def test_pol_ne_snizhaet_bolshiy_ball(self) -> None:
        """Если взвешенная сумма выше 75, пол её не трогает.

        Набираем 82 балла из 100: R01 13 + R02 15 + R04 8 + R05 6 + R06 10 +
        R07 5 + R09 9 + R11 6 + R13 2 + R14 4 + R15 4.
        """
        stroka = _row(
            r01=0.5,
            r02=0.0,
            flags=3.0,
            r04_intensivnost_utochneniy=1.0,
            r05_oshibka_profilya=1.0,
            r06_otklonenie_saldo=1.0,
            r07_kassovyy_bufer=0.0,
            r09_davlenie_ostatka=1.0,
            r11_neoplachennye_obyazatelstva=1.0,
            r13_hhi=1.0,
            r14_finansovye_operatsii=1.0,
        )
        result = evaluate_row(stroka)
        assert result.override_triggered
        assert result.score == pytest.approx(82.0)
        assert result.score > OVERRIDE_FLOOR

    def test_bez_povoda_pol_ne_primenyaetsya(self) -> None:
        result = evaluate_row(_row())
        assert not result.override_triggered
        assert result.score < OVERRIDE_FLOOR


class TestPolnota:
    @pytest.mark.parametrize(
        ("flagi", "ozhidaemo"),
        [(0.0, 1.0), (1.0, 2 / 3), (2.0, 1 / 3), (3.0, 0.0), (5.0, 0.0)],
    )
    def test_formula_polnoty(self, flagi: float, ozhidaemo: float) -> None:
        assert data_completeness(flagi) == pytest.approx(ozhidaemo)

    def test_polnota_ne_menyaet_uroven(self) -> None:
        """В слое 8.3 «серого» уровня нет — неполнота учтена баллом R15."""
        result = evaluate_row(_row(flags=1.0))
        assert result.data_completeness == pytest.approx(2 / 3)
        assert result.level is not RiskLevel.UNKNOWN
        assert result.risk.completeness == pytest.approx(1.0)


class TestRang:
    def test_rang_schitaetsya_vnutri_mesyatsa(self) -> None:
        """Ранг 1 у самой рискованной территории месяца."""
        riskier = evaluate_row(_row(territory_id="REG-002", r06_otklonenie_saldo=0.12))
        safer = evaluate_row(_row(territory_id="REG-001", r06_otklonenie_saldo=0.02))

        ranks = rank_within_month([safer, riskier])
        assert ranks[riskier.key] == 1
        assert ranks[safer.key] == 2

    def test_raznye_mesyatsy_rangiruyutsya_nezavisimo(self) -> None:
        january = evaluate_row(_row(month=1, r06_otklonenie_saldo=0.02))
        february = evaluate_row(_row(month=2, r06_otklonenie_saldo=0.12))

        ranks = rank_within_month([january, february])
        assert ranks[january.key] == 1
        assert ranks[february.key] == 1

    def test_tay_brek_po_kodu_territorii(self) -> None:
        """При равных баллах выше идёт территория с меньшим кодом."""
        first = evaluate_row(_row(territory_id="REG-001"))
        second = evaluate_row(_row(territory_id="REG-002"))

        ranks = rank_within_month([second, first])
        assert ranks["REG-001-01"] == 1
        assert ranks["REG-002-01"] == 2


class TestRazborIstochnika:
    def test_period_stroka_a_ne_data(self) -> None:
        """В книге все 74 831 значение периода — строки формата «MM.YYYY»."""
        assert parse_period("01.2025") == (1, 2025)
        assert parse_period("12.2025") == (12, 2025)

    @pytest.mark.parametrize(
        ("iz_knigi", "norma"),
        [
            ("Западно-Казахстанкая область", "Западно-Казахстанская область"),
            ("Мангыстауская область", "Мангистауская область"),
            ("Северо-Казахстанкая область", "Северо-Казахстанская область"),
            ("Туркистанская область", "Туркестанская область"),
        ],
    )
    def test_chetyre_napisaniya_normalizuyutsya(self, iz_knigi: str, norma: str) -> None:
        assert normalize_region_name(iz_knigi) == norma

    def test_sovpadayushchie_nazvaniya_ne_menyayutsya(self) -> None:
        assert normalize_region_name("Абайская область") == "Абайская область"
        assert normalize_region_name("г. Астана") == "г. Астана"

    def test_rasshozhdeniy_rovno_chetyre(self) -> None:
        assert len(SOURCE_SPELLINGS) == 4


# --- Сверка с книгой ---------------------------------------------------------


@pytest.mark.golden
class TestSverkaSKnigoy:
    def test_parametry_knigi_sovpadayut_s_modelyu(self, workbook_path: Path) -> None:
        """Лист «Параметры» объявлен редактируемым — расхождение обязано быть видно."""
        assert check_parameters_match_model(workbook_path) == []

    def test_summa_vesov_v_knige_sto(self, workbook_path: Path) -> None:
        parameters = load_parameters(workbook_path)
        assert sum(p.weight for p in parameters.values()) == pytest.approx(100.0)

    def test_240_strok_20_territoriy_12_mesyatsev(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        assert len(monthly_rows) == EXPECTED_MONTHLY_ROWS
        assert len({row.inputs.territory_id for row in monthly_rows}) == EXPECTED_TERRITORIES
        assert len({row.inputs.period for row in monthly_rows}) == EXPECTED_PERIODS

    def test_vse_stroki_urovnya_region(self, monthly_rows: list[BudgetMonthlyRow]) -> None:
        """Слой 8.3 общереспубликанский: районов в нём нет."""
        assert {row.geo_level for row in monthly_rows} == {"REGION"}
        assert {row.parent_territory_id for row in monthly_rows} == {"KZ"}

    def test_ball_sovpadaet_so_vsemi_240_strokami(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        """Главная сверка: независимый пересчёт против колонки Risk Score."""
        results = evaluate_rows(monthly_rows)
        for row, result in zip(monthly_rows, results, strict=True):
            assert result.score == pytest.approx(row.book_score, abs=1e-9), row.book_key

    def test_agregaty_sovpadayut(self, monthly_rows: list[BudgetMonthlyRow]) -> None:
        scores = [result.score for result in evaluate_rows(monthly_rows)]
        assert min(scores) == pytest.approx(BOOK_SCORE_MIN)
        assert max(scores) == pytest.approx(BOOK_SCORE_MAX)
        assert sum(scores) == pytest.approx(BOOK_SCORE_SUM, abs=1e-4)
        assert sum(scores) / len(scores) == pytest.approx(BOOK_SCORE_AVG, abs=1e-6)

    def test_raspredelenie_urovney(self, monthly_rows: list[BudgetMonthlyRow]) -> None:
        counts = Counter(result.level.label_ru for result in evaluate_rows(monthly_rows))
        assert counts["Низкий"] == BOOK_LEVEL_COUNTS["Низкий"]
        assert counts["Средний"] == BOOK_LEVEL_COUNTS["Средний"]
        assert counts["Высокий"] == BOOK_LEVEL_COUNTS["Высокий"]
        assert counts["Критический"] == BOOK_LEVEL_COUNTS["Критический"]

    def test_uroven_sovpadaet_postrochno(self, monthly_rows: list[BudgetMonthlyRow]) -> None:
        for row, result in zip(monthly_rows, evaluate_rows(monthly_rows), strict=True):
            assert result.level.label_ru == row.book_level, row.book_key

    def test_pereopredelenie_srabotalo_rovno_odin_raz(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        results = evaluate_rows(monthly_rows)
        triggered = [r for r in results if r.override_triggered]
        assert len(triggered) == BOOK_OVERRIDE_ROWS
        assert triggered[0].key == "REG-019-01"
        # Пол, а не замена уровня: балл равен ровно 75,0.
        assert triggered[0].score == 75.0

    def test_pereopredelenie_sovpadaet_s_knigoy_postrochno(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        for row, result in zip(monthly_rows, evaluate_rows(monthly_rows), strict=True):
            assert result.override_triggered == row.book_override, row.book_key

    def test_polnota_menshe_edinitsy_u_32_strok(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        """32 среза без раздела «IV. Сальдо по операциям с финансовыми активами»."""
        results = evaluate_rows(monthly_rows)
        incomplete = [r for r in results if r.data_completeness < 1.0]
        assert len(incomplete) == BOOK_INCOMPLETE_ROWS
        assert all(r.data_completeness == pytest.approx(2 / 3) for r in incomplete)

    def test_flag_otsutstvuyushchikh_korney_daet_te_zhe_32_stroki(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        assert sum(row.missing_roots_flag for row in monthly_rows) == BOOK_INCOMPLETE_ROWS

    def test_polnota_sovpadaet_s_knigoy(self, monthly_rows: list[BudgetMonthlyRow]) -> None:
        for row, result in zip(monthly_rows, evaluate_rows(monthly_rows), strict=True):
            assert result.data_completeness == pytest.approx(row.book_completeness), row.book_key

    @pytest.mark.parametrize(("klyuch", "ball"), sorted(BOOK_REFERENCE_SCORES.items()))
    def test_etalonnye_stroki(
        self, monthly_rows: list[BudgetMonthlyRow], klyuch: str, ball: float
    ) -> None:
        by_key = {r.key: r for r in evaluate_rows(monthly_rows)}
        assert by_key[klyuch].score == pytest.approx(ball, abs=1e-9)

    def test_astana_yanvar_edinstvennaya_kriticheskaya(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        results = evaluate_rows(monthly_rows)
        critical = [r for r in results if r.level is RiskLevel.CRITICAL]
        assert [r.key for r in critical] == ["REG-019-01"]
        assert critical[0].territory_name == "г. Астана"

    def test_rang_sovpadaet_s_knigoy(self, monthly_rows: list[BudgetMonthlyRow]) -> None:
        results = evaluate_rows(monthly_rows)
        ranks = rank_within_month(results)
        for row, result in zip(monthly_rows, results, strict=True):
            assert ranks[result.key] == row.book_rank, row.book_key

    def test_tri_indikatora_vyrozhdeny_na_vsekh_strokakh(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        """R03, R10, R12 не срабатывают ни разу — 14 % веса модели мертвы.

        Это факт данных 2025 года, а не дефект модели: пороги «перерасход
        периода», «отставание обязательств» и «ширина недоисполнения» в этой
        выборке не достигаются. Индикаторы намеренно оставлены в методике —
        их вырожденность должна быть видна, а не скрыта удалением.
        """
        never_fired: set[str] = set(BANDS)
        for result in evaluate_rows(monthly_rows):
            for factor in result.risk.factors:
                if factor.value:
                    never_fired.discard(factor.code)
        assert never_fired == DEGENERATE_INDICATORS

    def test_dominiruyushchiy_indikator_r11(self, monthly_rows: list[BudgetMonthlyRow]) -> None:
        """R11 «неоплаченные обязательства» — средний балл около 81 из 100."""
        results = evaluate_rows(monthly_rows)
        values = [
            factor.value
            for result in results
            for factor in result.risk.factors
            if factor.code == "R11" and factor.value is not None
        ]
        assert sum(values) / len(values) * 100 == pytest.approx(81.192, abs=0.01)

    def test_kato_v_knige_otsutstvuet(self, workbook_path: Path) -> None:
        """Ни в сырых данных, ни в расчёте нет ни одного поля КАТО.

        Из-за этого геопривязка возможна только по текстовому названию
        области, и справочник алиасов — не удобство, а единственный способ
        связать слой с геометрией.
        """
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            headers: list[str] = []
            for sheet_name in ("Расчет_месяц", "RAW_DATA_Бюджет_все_регионы_КЗ_"):
                sheet = workbook[sheet_name]
                first = next(sheet.iter_rows(max_row=1, values_only=True))
                headers.extend(str(cell) for cell in first if cell is not None)
        finally:
            workbook.close()
        assert not [h for h in headers if "като" in h.casefold()]

    def test_alias_zavoditsya_dlya_kazhdogo_napisaniya(
        self, monthly_rows: list[BudgetMonthlyRow]
    ) -> None:
        """Опечатки заводятся алиасами, а не «исправляются» в данных."""
        aliases = territory_aliases(monthly_rows)
        source_spellings = {alias for alias, _, kind in aliases if kind == "source_spelling"}
        assert source_spellings == set(SOURCE_SPELLINGS)

        # Исходное написание при этом сохраняется в самой строке.
        assert {row.source_region_name for row in monthly_rows if row.spelling_differs} == set(
            SOURCE_SPELLINGS
        )


@pytest.mark.golden
@pytest.mark.slow
class TestSyryeDannye:
    def test_odinnadtsat_listov(self, workbook_path: Path) -> None:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            assert len(workbook.sheetnames) == EXPECTED_SHEETS
        finally:
            workbook.close()

    def test_74831_strok_syrykh_dannykh(self, workbook_path: Path) -> None:
        facts = list(iter_raw_facts(workbook_path))
        assert len(facts) == EXPECTED_RAW_ROWS

    def test_vse_periody_2025_goda(self, workbook_path: Path) -> None:
        periods = {fact.period for fact in iter_raw_facts(workbook_path)}
        assert len(periods) == EXPECTED_PERIODS
        assert all(fact_period.endswith("2025") for fact_period in periods)

    def test_korni_ierarkhii_bez_roditelya(self, workbook_path: Path) -> None:
        """Пропуски `parent_id` в точности равны числу строк с level = 0."""
        roots = 0
        without_parent = 0
        for fact in iter_raw_facts(workbook_path):
            roots += fact.level == 0
            without_parent += fact.parent_id is None
        assert roots == without_parent == 2_608
