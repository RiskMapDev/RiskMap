"""Отчётность и экспорт.

Проверяется не «файл скачался», а то, ради чего отчёт вообще существует как
документ: что он назван, датирован, подписан, что в нём перечислены применённые
фильтры и источники с датами, и — главное — что он признаётся в собственной
неполноте числом, когда неполнота есть.

Отдельная тема — «нет данных». В исходных книгах заказчика неизмеренный
показатель попадал в таблицу нулём и молча снижал балл. Тесты ниже следят,
чтобы в отчётах эта ошибка не воспроизвелась ни в одном из трёх форматов: ни
нулём, ни пустой ячейкой.

Проверки идут по **содержимому файла**, а не по промежуточной структуре.
Структуру можно собрать правильно и потерять при отрисовке, и такую потерю
видно только если распаковать docx и прочитать xlsx.

Про изоляцию. База теста — общая с уже загруженными данными слоёв, поэтому
выборка «всё, что есть» здесь ничего не доказывает: числа в ней зависят от
того, какие книги успели импортировать. Каждый набор объектов помечается
уникальной строкой в наименовании, а выборка сужается поиском по этой метке.
Так утверждения тестов остаются верными независимо от содержимого базы.
"""

from __future__ import annotations

import io
import re
import uuid
import zipfile
from collections.abc import Iterator
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any
from urllib.parse import unquote

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.queryspec import ObjectType, QuerySpec
from app.api.report_routes import _ascii_fallback, _content_disposition
from app.api.report_routes import router as report_router
from app.core import security
from app.core.config import get_settings
from app.db.models.access import (
    AuditAction,
    AuditLogEntry,
    RoleCode,
    SensitiveDataAccess,
)
from app.db.models.access import User as UserModel
from app.db.models.source import SourceDataset, SourceFile
from app.db.models.subsidy import SubsidyRecipient
from app.db.models.territory import Territory
from app.risk.core import RiskLevel
from app.services import report_render, reports
from app.services.report_render import PdfUnavailableError, ReportFormat
from app.services.reports import Cell, ReportTemplate
from tests.conftest import UserFactory

REPORTS_URL = "/api/v1/reports"


# --- Приспособления ----------------------------------------------------------


@pytest.fixture
def report_app(db_session: Session) -> Iterator[FastAPI]:
    """Приложение с подключённым роутером отчётов.

    Роутер подключается здесь, а не берётся готовым из `create_app`: проверять
    права надо настоящим запросом, а зависимость прав существует только на
    зарегистрированном маршруте — безупречная проверка, не навешенная на
    обработчик, снаружи выглядит как открытая дверь.
    """
    from app.db.session import get_db
    from app.main import create_app

    application = create_app()
    application.include_router(report_router, prefix=get_settings().api_prefix)
    application.dependency_overrides[get_db] = lambda: db_session
    try:
        yield application
    finally:
        application.dependency_overrides.clear()


@pytest.fixture
def report_client(report_app: FastAPI) -> Iterator[TestClient]:
    with TestClient(report_app) as client:
        yield client
    security.clear_revoked_tokens()


def _token(user: UserModel) -> str:
    token, _ = security.create_access_token(
        user_id=user.id, login=user.login, role=str(user.role.code)
    )
    return token


def _headers(user: UserModel) -> dict[str, str]:
    return {"Authorization": f"Bearer {_token(user)}"}


@pytest.fixture
def marker() -> str:
    """Метка прогона: по ней выборка сужается до объектов этого теста."""
    return f"проба-{uuid.uuid4().hex[:10]}"


def _spec(marker: str, **overrides: Any) -> QuerySpec:
    """Выборка, ограниченная объектами текущего теста."""
    payload: dict[str, Any] = {
        "search": marker,
        "object_types": [ObjectType.SUBSIDY_RECIPIENT],
    }
    payload.update(overrides)
    return QuerySpec(**payload)


class _Recipient:
    """Пара «модель + ожидаемые значения» — чтобы тесты не хардкодили ИИН."""

    def __init__(self, model: SubsidyRecipient, xin: str) -> None:
        self.model = model
        self.xin = xin

    @property
    def masked(self) -> str:
        return f"{self.xin[:4]}{'*' * 6}{self.xin[-2:]}"


def _make_recipient(
    session: Session,
    *,
    name: str,
    level: RiskLevel,
    score: float | None,
    territory: Territory | None,
    amount: Decimal | None = Decimal("1000000.00"),
    completeness: float = 1.0,
    s1: float | None = 0.5,
) -> _Recipient:
    """Получатель субсидий — самый дешёвый в сборке объект выборки.

    Слой 8.5 выбран не случайно: у него есть и балл, и уровень, и полнота, и
    персональный идентификатор, то есть всё, что проверяют тесты ниже.
    """
    # ИИН генерируется, а не задаётся константой: база теста общая, и
    # захардкоженный номер рано или поздно столкнётся с ограничением
    # уникальности либо совпадёт с реально загруженной строкой.
    xin = f"84{uuid.uuid4().int % 10**10:010d}"
    recipient = SubsidyRecipient(
        xin=xin,
        name=name,
        natural_key=f"{xin}:1",
        territory_id=territory.id if territory is not None else None,
        territory_resolution="resolved" if territory is not None else "empty",
        total_amount=amount if amount is not None else Decimal("0.00"),
        payments_count=3,
        programs_count=1,
        s1_concentration=s1,
        s2_repetition=None,
        model_code="subsidies-8.5",
        model_version="1.0",
        risk_score=score,
        risk_level=str(level),
        risk_completeness=completeness,
        director_name="Петров П.П.",
    )
    session.add(recipient)
    session.flush()
    return _Recipient(recipient, xin)


@pytest.fixture
def selection(
    db_session: Session, territories: dict[str, Territory], marker: str
) -> dict[str, _Recipient]:
    """Выборка без пробелов: у всех объектов уровень измерен."""
    karasay = territories["karasay"]
    items = {
        "alpha": _make_recipient(
            db_session,
            name=f"ТОО «Альфа» {marker}",
            level=RiskLevel.CRITICAL,
            score=91.0,
            territory=karasay,
        ),
        "beta": _make_recipient(
            db_session,
            name=f"ТОО «Бета» {marker}",
            level=RiskLevel.HIGH,
            score=72.5,
            territory=karasay,
        ),
        "gamma": _make_recipient(
            db_session,
            name=f"ТОО «Гамма» {marker}",
            level=RiskLevel.LOW,
            score=12.0,
            territory=territories["talgar"],
        ),
    }
    db_session.flush()
    return items


@pytest.fixture
def selection_with_gaps(
    db_session: Session,
    territories: dict[str, Territory],
    selection: dict[str, _Recipient],
    marker: str,
) -> dict[str, _Recipient]:
    """Та же выборка плюс объекты с серым уровнем, без балла и без территории."""
    selection["delta"] = _make_recipient(
        db_session,
        name=f"ТОО «Дельта» без оценки {marker}",
        level=RiskLevel.UNKNOWN,
        score=None,
        territory=territories["karasay"],
        amount=None,
        completeness=0.1,
        s1=None,
    )
    selection["epsilon"] = _make_recipient(
        db_session,
        name=f"ТОО «Эпсилон» без территории {marker}",
        level=RiskLevel.UNKNOWN,
        score=None,
        territory=None,
        completeness=0.0,
        s1=None,
    )
    db_session.flush()
    return selection


@pytest.fixture
def sources(db_session: Session) -> SourceFile:
    """Источник с датой актуальности, источник без неё и лист методики."""
    suffix = uuid.uuid4().hex[:10]
    source_file = SourceFile(
        file_name=f"Субсидии-{suffix}.xlsx",
        normalized_name=f"субсидии-{suffix}.xlsx",
        sha256=uuid.uuid4().hex + uuid.uuid4().hex,
        size_bytes=1024,
    )
    db_session.add(source_file)
    db_session.flush()

    db_session.add_all(
        [
            SourceDataset(
                source_file_id=source_file.id,
                layer_code="8.5",
                sheet_name="Расчёт по получателям",
                role="raw",
                row_count=4231,
                data_as_of=date(2025, 12, 31),
            ),
            SourceDataset(
                source_file_id=source_file.id,
                layer_code="8.7",
                sheet_name="Организации",
                role="raw",
                row_count=100,
                data_as_of=None,
            ),
            # Лист методики описывает, как считать, а не что показывать:
            # в перечень источников отчёта попасть не должен.
            SourceDataset(
                source_file_id=source_file.id,
                layer_code="8.5",
                sheet_name="Методика",
                role="model_config",
            ),
        ]
    )
    db_session.flush()
    return source_file


@pytest.fixture
def analyst(make_user: UserFactory, db_session: Session) -> UserModel:
    user = make_user(RoleCode.ANALYST)
    db_session.commit()
    return user


# --- Чтение готовых файлов ---------------------------------------------------


def docx_text(payload: bytes) -> str:
    """Весь видимый текст документа Word.

    Читается разметка файла, а не структура python-docx: проверять надо то, что
    окажется у пользователя, а не то, что было в памяти до сохранения.
    """
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        xml = archive.read("word/document.xml").decode("utf-8")
    return " ".join(re.findall(r"<w:t[^>]*>([^<]*)</w:t>", xml))


def xlsx_cells(payload: bytes) -> list[tuple[str, Any]]:
    """Все непустые ячейки книги как пары «лист → значение»."""
    workbook = load_workbook(io.BytesIO(payload))
    collected: list[tuple[str, Any]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    collected.append((sheet.title, cell.value))
    return collected


def xlsx_text(payload: bytes) -> str:
    return " ".join(str(value) for _, value in xlsx_cells(payload))


# --- Ячейка: «нет данных» вместо нуля ----------------------------------------


class TestYacheyka:
    """Правила печати значения. Здесь решается судьба «нет данных»."""

    def test_otsutstvuyushchee_znachenie_ne_nol(self) -> None:
        for cell in (
            Cell.of(None),
            Cell.of("   "),
            Cell.money(None),
            Cell.score(None),
            Cell.percent(None),
            Cell.whole(None),
            Cell.when(None),
        ):
            assert cell.text == reports.NO_DATA
            assert cell.number is None
            assert cell.is_missing

    def test_nastoyashchiy_nol_pechataetsya_nulyom(self) -> None:
        """Измеренный ноль — не отсутствие, и подменять его нельзя."""
        assert Cell.whole(0).text == "0"
        assert Cell.whole(0).number == 0.0
        assert Cell.whole(0).is_missing is False
        assert Cell.score(0.0).number == 0.0
        assert Cell.percent(0.0).text == "0%"

    def test_dolya_ot_nulya_ne_opredelena(self) -> None:
        """Доля в пустой выборке — «нет данных», а не «0%»."""
        assert Cell.share(0, 0).is_missing
        assert Cell.share(0, 0).text == reports.NO_DATA
        assert Cell.share(1, 4).number == pytest.approx(0.25)

    def test_predvaritelnyy_ball_pomechen_i_ne_chislo(self) -> None:
        """Предварительный балл не смешивается с окончательным.

        Пометка стоит в тексте, числового представления нет: иначе Excel
        просуммирует и усреднит несравнимое.
        """
        preliminary = Cell.score(77.0, preliminary=True)
        final = Cell.score(77.0)

        assert reports.PRELIMINARY_MARK in preliminary.text
        assert preliminary.number is None
        assert preliminary.is_missing is False
        assert reports.PRELIMINARY_MARK not in final.text
        assert final.number == 77.0

    def test_summa_ostayotsya_chislom_dlya_excel(self) -> None:
        cell = Cell.money(Decimal("1234567.50"))
        assert cell.number == pytest.approx(1234567.5)
        assert cell.number_format


# --- Предупреждение о полноте ------------------------------------------------


class TestPreduprezhdenieOPolnote:
    """Отчёт, умалчивающий о неполноте, опаснее отсутствия отчёта."""

    def test_bez_probelov_preduprezhdeniya_net(self) -> None:
        warning = reports.CompletenessWarning(
            total=10, unknown_level=0, preliminary_score=0, without_territory=0
        )
        assert warning.has_gaps is False
        assert warning.lines[0] == reports.WARNING_CLEAN

    def test_serye_obekty_nazvany_chislom(self) -> None:
        warning = reports.CompletenessWarning(
            total=10, unknown_level=3, preliminary_score=0, without_territory=0
        )
        text = " ".join(warning.lines)

        assert warning.has_gaps is True
        assert warning.lines[0] == reports.WARNING_MARKER
        # Именно число, а не «часть объектов» и не «некоторые».
        assert "3" in text
        assert "10" in text

    def test_predvaritelnyy_ball_nazvan_chislom(self) -> None:
        warning = reports.CompletenessWarning(
            total=10, unknown_level=0, preliminary_score=4, without_territory=0
        )
        text = " ".join(warning.lines)

        assert warning.has_gaps is True
        assert "4" in text
        assert reports.PRELIMINARY_MARK in text

    def test_usechenie_vyborki_nazvano(self) -> None:
        warning = reports.CompletenessWarning(
            total=9000,
            unknown_level=0,
            preliminary_score=0,
            without_territory=0,
            truncated_to=5000,
        )
        text = " ".join(warning.lines)

        assert warning.has_gaps is True
        assert "5000" in text
        assert "9000" in text


# --- Сборка данных отчёта ----------------------------------------------------


class TestSborkaOtchyota:
    """Данные отчёта до отрисовки."""

    def test_vosem_shablonov(self) -> None:
        """Ровно восемь шаблонов ТЗ и референса, ни больше ни меньше."""
        assert len(list(ReportTemplate)) == 8
        catalog = reports.template_catalog()
        assert len(catalog) == 8
        for item in catalog:
            assert item["title"]
            assert item["description"]

    def test_shapka_soderzhit_avtora_i_rol(
        self,
        db_session: Session,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        document = reports.build_report(
            db_session, ReportTemplate.TERRITORY, _spec(marker), user=analyst
        )

        assert document.generated_by_name == analyst.full_name
        assert document.generated_by_role == analyst.role.title
        assert document.generated_at.tzinfo is not None
        assert "UTC" in document.generated_at_text

    def test_filtry_chelovekochitaemy_a_ne_json(
        self, db_session: Session, territories: dict[str, Territory]
    ) -> None:
        spec = QuerySpec(
            year=2025,
            territory_codes=[territories["karasay"].code],
            object_types=[ObjectType.SUBSIDY_RECIPIENT],
            risk_levels=[RiskLevel.CRITICAL, RiskLevel.HIGH],
            search="Альфа",
        )

        filters = dict(reports.describe_filters(db_session, spec))

        assert filters["Период"] == "2025 год"
        # Код территории подменён названием: код читателю отчёта ничего не говорит.
        assert filters["Территории"] == "Карасайский район"
        assert filters["Типы объектов"] == "Получатель субсидий"
        assert "Критический" in filters["Уровни риска"]
        assert filters["Поисковый запрос"] == "Альфа"
        # Отсутствие серого уровня в фильтре обязано быть названо вслух.
        assert "НЕ включены" in filters["Уровни риска"]

    def test_pustye_filtry_nazvany_pustymi(self, db_session: Session) -> None:
        filters = dict(reports.describe_filters(db_session, QuerySpec()))
        assert "не применялись" in filters["Фильтры"]
        # Строка об уровнях присутствует и здесь: серые объекты включены.
        assert "все уровни" in filters["Уровни риска"]

    def test_neizvestnyy_kod_territorii_ne_teryaetsya(self, db_session: Session) -> None:
        filters = dict(
            reports.describe_filters(db_session, QuerySpec(territory_codes=["net-takogo"]))
        )
        assert "net-takogo" in filters["Территории"]
        assert "не найден" in filters["Территории"]

    def test_ogranichenie_roli_nazvano_v_filtrakh(self, db_session: Session) -> None:
        filters = dict(
            reports.describe_filters(
                db_session, QuerySpec(), scope_territory_name="Карасайский район"
            )
        )
        assert "Карасайский район" in filters["Территориальное ограничение роли"]

    def test_istochniki_s_datami_aktualnosti(
        self, db_session: Session, sources: SourceFile
    ) -> None:
        collected = reports.collect_sources(db_session, ["8.5", "8.7"])
        mine = [item for item in collected if item.file_name == sources.file_name]
        by_layer = {item.layer_code: item for item in mine}

        assert by_layer["8.5"].data_as_of == date(2025, 12, 31)
        assert by_layer["8.5"].as_of_text == "31.12.2025"
        # Дата отсутствует — так и написано, а не пусто.
        assert by_layer["8.7"].data_as_of is None
        assert reports.NO_DATA in by_layer["8.7"].as_of_text

    def test_list_metodiki_ne_istochnik_dannykh(
        self, db_session: Session, sources: SourceFile
    ) -> None:
        collected = reports.collect_sources(db_session, ["8.5"])
        mine = [item for item in collected if item.file_name == sources.file_name]
        sheets = {name for item in mine for name in item.sheet_names}

        assert "Расчёт по получателям" in sheets
        assert "Методика" not in sheets

    def test_preduprezhdenie_poyavlyaetsya_pri_serykh_obektakh(
        self,
        db_session: Session,
        analyst: UserModel,
        selection_with_gaps: dict[str, _Recipient],
        marker: str,
    ) -> None:
        document = reports.build_report(
            db_session, ReportTemplate.REGION_SUMMARY, _spec(marker), user=analyst
        )

        assert document.warning.total == 5
        assert document.warning.has_gaps is True
        assert document.warning.unknown_level == 2
        assert document.warning.without_territory == 1

    def test_preduprezhdeniya_net_kogda_probelov_net(
        self,
        db_session: Session,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        document = reports.build_report(
            db_session, ReportTemplate.REGION_SUMMARY, _spec(marker), user=analyst
        )

        assert document.warning.total == 3
        assert document.warning.unknown_level == 0
        assert document.warning.has_gaps is False
        assert document.warning.lines[0] == reports.WARNING_CLEAN

    def test_pustaya_vyborka_obyasnyaetsya(
        self, db_session: Session, analyst: UserModel, marker: str
    ) -> None:
        """Пустой отчёт без объяснения читается как «рисков нет»."""
        document = reports.build_report(
            db_session, ReportTemplate.TERRITORY, _spec(marker), user=analyst
        )
        text = " ".join(
            line for section in document.sections for line in section.paragraphs
        )

        assert document.warning.total == 0
        assert "не означает" in text

    def test_istochniki_perechisleny_i_u_pustoy_vyborki(
        self, db_session: Session, analyst: UserModel, sources: SourceFile, marker: str
    ) -> None:
        """Читатель должен видеть, по каким данным система искала и не нашла."""
        document = reports.build_report(
            db_session, ReportTemplate.TERRITORY, _spec(marker), user=analyst
        )
        assert document.sources


# --- Отрисовка: восемь шаблонов во всех форматах -----------------------------


@pytest.mark.parametrize("template", list(ReportTemplate))
@pytest.mark.parametrize("report_format", list(ReportFormat))
def test_shablon_sobiraetsya_vo_vsekh_formatakh(
    db_session: Session,
    analyst: UserModel,
    selection_with_gaps: dict[str, _Recipient],
    sources: SourceFile,
    marker: str,
    template: ReportTemplate,
    report_format: ReportFormat,
) -> None:
    """Каждый из восьми шаблонов собирается в каждом из трёх форматов."""
    document = reports.build_report(db_session, template, _spec(marker), user=analyst)
    payload = report_render.render(document, report_format)

    assert len(payload) > 1000
    if report_format is ReportFormat.PDF:
        assert payload.startswith(b"%PDF")
        # Шрифт встроен подмножеством — без него кириллица не отобразится.
        assert b"FontFile2" in payload
    else:
        # docx и xlsx — ZIP-контейнеры OOXML.
        assert payload.startswith(b"PK\x03\x04")


class TestSoderzhimoeFayla:
    """Что реально оказалось в файле, а не в структуре до отрисовки."""

    @pytest.fixture
    def document(
        self,
        db_session: Session,
        analyst: UserModel,
        selection_with_gaps: dict[str, _Recipient],
        sources: SourceFile,
        territories: dict[str, Territory],
        marker: str,
    ) -> reports.ReportDocument:
        return reports.build_report(
            db_session,
            ReportTemplate.HIGH_RISK,
            _spec(marker, year=2025, territory_codes=[territories["karasay"].code]),
            user=analyst,
        )

    def test_word_soderzhit_datu_avtora_filtry_i_istochniki(
        self, document: reports.ReportDocument, sources: SourceFile
    ) -> None:
        text = docx_text(report_render.render_docx(document))

        assert document.generated_at_text in text
        assert document.generated_by_name in text
        assert document.generated_by_role in text
        assert "Применённые фильтры" in text
        assert "2025 год" in text
        assert "Карасайский район" in text
        assert "Источники данных" in text
        assert sources.file_name in text
        assert "31.12.2025" in text

    def test_excel_soderzhit_datu_avtora_filtry_i_istochniki(
        self, document: reports.ReportDocument, sources: SourceFile
    ) -> None:
        text = xlsx_text(report_render.render_xlsx(document))

        assert document.generated_at_text in text
        assert document.generated_by_name in text
        assert document.generated_by_role in text
        assert "2025 год" in text
        assert sources.file_name in text
        assert "31.12.2025" in text

    def test_pdf_soderzhit_vstroennyy_kirillicheskiy_shrift(
        self, document: reports.ReportDocument
    ) -> None:
        payload = report_render.render_pdf(document)

        assert payload.startswith(b"%PDF")
        assert b"FontFile2" in payload
        assert b"DejaVuSans" in payload

    def test_preduprezhdenie_v_kazhdom_formate(
        self, document: reports.ReportDocument
    ) -> None:
        assert document.warning.has_gaps is True

        for text in (
            docx_text(report_render.render_docx(document)),
            xlsx_text(report_render.render_xlsx(document)),
        ):
            assert reports.WARNING_HEADING in text
            assert reports.WARNING_MARKER in text

    def test_bez_probelov_preduprezhdenie_ne_krichit(
        self,
        db_session: Session,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        document = reports.build_report(
            db_session, ReportTemplate.HIGH_RISK, _spec(marker), user=analyst
        )
        text = docx_text(report_render.render_docx(document))

        # Раздел присутствует всегда: его отсутствие читалось бы как
        # «проверку полноты не делали».
        assert reports.WARNING_HEADING in text
        assert reports.WARNING_CLEAN in text
        assert reports.WARNING_MARKER not in text

    def test_net_dannykh_ne_prevrashchaetsya_v_nol_v_word(
        self, document: reports.ReportDocument
    ) -> None:
        text = docx_text(report_render.render_docx(document))
        assert reports.NO_DATA in text

    def test_net_dannykh_ne_prevrashchaetsya_v_nol_v_excel(
        self, document: reports.ReportDocument
    ) -> None:
        """В книге отсутствующее значение — текст «нет данных», а не 0 и не пусто."""
        payload = report_render.render_xlsx(document)
        values = [value for _, value in xlsx_cells(payload)]

        assert reports.NO_DATA in values

        # Ни одна ячейка расшифровки факторов не должна оказаться числовым
        # нулём: у «Дельты» факторы не измерены, и ноль был бы утверждением,
        # которого источник не делал.
        workbook = load_workbook(io.BytesIO(payload))
        for sheet in workbook.worksheets:
            if "фактор" not in sheet.title.lower():
                continue
            zeroes = [
                cell.value
                for row in sheet.iter_rows(min_row=3)
                for cell in row
                if isinstance(cell.value, (int, float))
                and not isinstance(cell.value, bool)
                and cell.value == 0
            ]
            assert not zeroes

    def test_excel_khranit_chisla_chislami(self, document: reports.ReportDocument) -> None:
        """Книгу открывают, чтобы сортировать и складывать."""
        numbers = [
            value
            for _, value in xlsx_cells(report_render.render_xlsx(document))
            if isinstance(value, (int, float)) and not isinstance(value, bool)
        ]
        assert numbers

    def test_pustaya_tablitsa_ne_propuskaetsya_a_obyasnyaetsya(
        self, db_session: Session, analyst: UserModel, marker: str
    ) -> None:
        """Пропуск читается как «раздел забыли», а не как «строк нет»."""
        document = reports.build_report(
            db_session, ReportTemplate.RISK_CATEGORY, _spec(marker), user=analyst
        )
        text = docx_text(report_render.render_docx(document))

        assert "Строк, удовлетворяющих условиям, нет." in text

    def test_primechanie_o_nule_v_dokumente(
        self, document: reports.ReportDocument
    ) -> None:
        text = docx_text(report_render.render_docx(document))
        assert "Ноль в таблице" in text


# --- Маскирование ------------------------------------------------------------


class TestMaskirovanie:
    """ИИН в отчёте показывается по роли, полное раскрытие журналируется."""

    def _document(
        self, session: Session, user: UserModel, marker: str
    ) -> reports.ReportDocument:
        return reports.build_report(
            session, ReportTemplate.HIGH_RISK, _spec(marker), user=user
        )

    def _identifiers(self, document: reports.ReportDocument) -> list[str]:
        for section in document.sections:
            for table in section.tables:
                titles = [column.title for column in table.columns]
                if "ИИН / БИН" not in titles:
                    continue
                index = titles.index("ИИН / БИН")
                return [row[index].text for row in table.rows]
        return []

    def test_analitik_vidit_masku(
        self,
        db_session: Session,
        make_user: UserFactory,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        analyst = make_user(RoleCode.ANALYST)
        db_session.flush()
        # Значение приходит из базы строкой, а не членом перечисления:
        # колонка объявлена как String, и сравнивать надо по значению.
        assert str(analyst.role.sensitive_data_access) == str(SensitiveDataAccess.MASKED)

        values = self._identifiers(self._document(db_session, analyst, marker))

        assert values
        assert all("*" in value for value in values)
        assert selection["alpha"].xin not in values
        assert selection["alpha"].masked in values

    def test_prosmotr_ne_vidit_vovse(
        self,
        db_session: Session,
        make_user: UserFactory,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        viewer = make_user(RoleCode.VIEWER)
        db_session.flush()

        values = self._identifiers(self._document(db_session, viewer, marker))

        assert values
        assert all(value == reports.CLOSED_BY_ROLE for value in values)
        # «Скрыто ролью» — не «нет данных»: значение есть, просто не для вас.
        assert reports.NO_DATA not in values

    def test_administrator_vidit_polnostyu_i_eto_v_zhurnale(
        self,
        db_session: Session,
        make_user: UserFactory,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        admin = make_user(RoleCode.ADMIN)
        db_session.flush()

        values = self._identifiers(self._document(db_session, admin, marker))

        assert selection["alpha"].xin in values

        records = (
            db_session.execute(
                select(AuditLogEntry)
                .where(AuditLogEntry.user_id == admin.id)
                .where(AuditLogEntry.action == AuditAction.SENSITIVE_VIEW)
            )
            .scalars()
            .all()
        )
        assert records
        assert any((record.details or {}).get("count", 0) >= 1 for record in records)

    def test_maska_ne_utekaet_v_gotovyy_fayl(
        self,
        db_session: Session,
        make_user: UserFactory,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        """Проверка по содержимому файла, а не по структуре до отрисовки."""
        analyst = make_user(RoleCode.ANALYST)
        db_session.flush()
        document = self._document(db_session, analyst, marker)

        text = docx_text(report_render.render_docx(document))

        assert selection["alpha"].xin not in text
        assert selection["alpha"].masked in text


# --- Журналирование ----------------------------------------------------------


class TestZhurnal:
    """Каждое формирование и каждая выгрузка оставляют след."""

    def test_formirovanie_pishetsya_v_zhurnal(
        self,
        db_session: Session,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        reports.build_report(
            db_session, ReportTemplate.RATINGS, _spec(marker, year=2025), user=analyst
        )

        record = db_session.execute(
            select(AuditLogEntry)
            .where(AuditLogEntry.user_id == analyst.id)
            .where(AuditLogEntry.action == AuditAction.REPORT_GENERATED)
        ).scalar_one()

        assert record.entity_type == "report"
        assert record.entity_id == str(ReportTemplate.RATINGS)
        assert record.details is not None
        assert record.details["template"] == str(ReportTemplate.RATINGS)
        assert record.details["filters"]["year"] == "2025"
        assert record.details["objects_total"] == 3

    def test_kazhdoe_formirovanie_otdelnaya_zapis(
        self,
        db_session: Session,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        reports.build_report(db_session, ReportTemplate.TERRITORY, _spec(marker), user=analyst)
        reports.build_report(db_session, ReportTemplate.RATINGS, _spec(marker), user=analyst)

        records = (
            db_session.execute(
                select(AuditLogEntry)
                .where(AuditLogEntry.user_id == analyst.id)
                .where(AuditLogEntry.action == AuditAction.REPORT_GENERATED)
            )
            .scalars()
            .all()
        )
        assert len(records) == 2

    def test_v_zhurnal_ne_popadayut_personalnye_dannye(
        self,
        db_session: Session,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        """Журнал фиксирует факт обращения, а не содержимое."""
        secret = selection["alpha"].xin
        reports.build_report(
            db_session, ReportTemplate.HIGH_RISK, _spec(marker), user=analyst
        )
        records = (
            db_session.execute(
                select(AuditLogEntry).where(AuditLogEntry.user_id == analyst.id)
            )
            .scalars()
            .all()
        )

        for record in records:
            assert secret not in str(record.details)

    def test_vygruzka_pishetsya_otdelnym_sobytiem(
        self,
        db_session: Session,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
        report_client: TestClient,
    ) -> None:
        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.TERRITORY.value}?format=docx",
            json=_spec(marker).model_dump(mode="json"),
            headers=_headers(analyst),
        )
        assert response.status_code == 200

        actions = {
            record.action
            for record in db_session.execute(
                select(AuditLogEntry).where(AuditLogEntry.user_id == analyst.id)
            ).scalars()
        }
        assert AuditAction.REPORT_GENERATED in actions
        assert AuditAction.EXPORT in actions

        export = db_session.execute(
            select(AuditLogEntry)
            .where(AuditLogEntry.user_id == analyst.id)
            .where(AuditLogEntry.action == AuditAction.EXPORT)
        ).scalar_one()
        assert export.details is not None
        assert export.details["format"] == "docx"
        assert export.details["size_bytes"] > 0
        assert export.details["file_name"].endswith(".docx")


# --- Эндпоинты ---------------------------------------------------------------


class TestEndpointy:
    """HTTP-край: права, форматы, имя файла, территориальное ограничение."""

    def test_katalog_shablonov(self, report_client: TestClient) -> None:
        response = report_client.get(f"{REPORTS_URL}/templates")
        assert response.status_code == 200
        assert len(response.json()) == 8

    def test_katalog_formatov_soobshchaet_o_dostupnosti_pdf(
        self, report_client: TestClient
    ) -> None:
        response = report_client.get(f"{REPORTS_URL}/formats")
        assert response.status_code == 200
        codes = {item["code"]: item for item in response.json()}
        assert codes["docx"]["available"] is True
        assert codes["xlsx"]["available"] is True
        assert "available" in codes["pdf"]

    @pytest.mark.parametrize("report_format", ["docx", "xlsx", "pdf"])
    def test_otchyot_otdayotsya_faylom(
        self,
        report_client: TestClient,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
        report_format: str,
    ) -> None:
        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.REGION_SUMMARY.value}?format={report_format}",
            json=_spec(marker).model_dump(mode="json"),
            headers=_headers(analyst),
        )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            ReportFormat(report_format).media_type
        )
        assert response.headers["cache-control"] == "no-store"
        assert len(response.content) > 1000

    def test_format_po_umolchaniyu_word(
        self,
        report_client: TestClient,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.TERRITORY.value}",
            json=_spec(marker).model_dump(mode="json"),
            headers=_headers(analyst),
        )
        assert response.status_code == 200
        assert response.content.startswith(b"PK\x03\x04")

    def test_imya_fayla_kirillicey_po_rfc_5987(
        self,
        report_client: TestClient,
        analyst: UserModel,
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.HIGH_RISK.value}?format=docx",
            json=_spec(marker).model_dump(mode="json"),
            headers=_headers(analyst),
        )
        disposition = response.headers["content-disposition"]

        assert disposition.startswith("attachment;")
        # Запасное ASCII-имя для клиентов, не знающих RFC 5987.
        ascii_part = re.search(r'filename="([^"]*)"', disposition)
        assert ascii_part is not None
        assert ascii_part.group(1).isascii()
        # Настоящее имя — в filename*, в UTF-8 и с процентным кодированием.
        encoded = re.search(r"filename\*=UTF-8''(\S+)", disposition)
        assert encoded is not None
        decoded = unquote(encoded.group(1))
        assert "Перечень высокорисковых объектов" in decoded
        assert decoded.endswith(".docx")

    def test_bez_prava_otkaz(
        self, report_client: TestClient, make_user: UserFactory, db_session: Session
    ) -> None:
        viewer = make_user(RoleCode.VIEWER)
        db_session.commit()

        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.TERRITORY.value}",
            json=QuerySpec().model_dump(mode="json"),
            headers=_headers(viewer),
        )

        assert response.status_code == 403

    def test_bez_tokena_otkaz(self, report_client: TestClient) -> None:
        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.TERRITORY.value}",
            json=QuerySpec().model_dump(mode="json"),
        )
        assert response.status_code == 401

    def test_neizvestnyy_shablon_otvergaetsya(
        self, report_client: TestClient, analyst: UserModel
    ) -> None:
        response = report_client.post(
            f"{REPORTS_URL}/net-takogo-shablona",
            json=QuerySpec().model_dump(mode="json"),
            headers=_headers(analyst),
        )
        assert response.status_code == 422

    def test_territorialnoe_ogranichenie_soblyudaetsya_i_nazvano(
        self,
        report_client: TestClient,
        make_user: UserFactory,
        db_session: Session,
        territories: dict[str, Territory],
        selection: dict[str, _Recipient],
        marker: str,
    ) -> None:
        """Ограничение роли — тоже фильтр, и умолчать о нём нельзя."""
        local = make_user(RoleCode.ANALYST, territory_id=territories["karasay"].id)
        db_session.commit()

        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.TERRITORY.value}?format=docx",
            json=_spec(marker).model_dump(mode="json"),
            headers=_headers(local),
        )
        text = docx_text(response.content)

        assert "Территориальное ограничение роли" in text
        assert "Карасайский район" in text
        assert "ТОО «Альфа»" in text
        # Талгарский район в зону доступа не входит.
        assert "ТОО «Гамма»" not in text

    def test_pdf_nedostupen_501_a_ne_bityy_fayl(
        self,
        report_client: TestClient,
        analyst: UserModel,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Честный отказ лучше документа, который открывается и не читается."""

        def net_shrifta(_document: object) -> bytes:
            raise PdfUnavailableError("Не найден файл шрифта DejaVuSans.ttf.")

        monkeypatch.setitem(report_render._RENDERERS, ReportFormat.PDF, net_shrifta)

        response = report_client.post(
            f"{REPORTS_URL}/{ReportTemplate.TERRITORY.value}?format=pdf",
            json=QuerySpec(search="zavedomo-net").model_dump(mode="json"),
            headers=_headers(analyst),
        )

        assert response.status_code == 501
        detail = response.json()["detail"]
        assert "DejaVuSans.ttf" in detail
        assert "Word" in detail


# --- Шрифт для PDF -----------------------------------------------------------


class TestShriftPdf:
    """Шрифт лежит в репозитории, а не тянется из сети при первом запуске."""

    def test_fayly_shrifta_na_meste(self) -> None:
        directory = report_render.font_directory()
        assert (directory / "DejaVuSans.ttf").is_file()
        assert (directory / "DejaVuSans-Bold.ttf").is_file()

    def test_litsenziya_i_proiskhozhdenie_zafiksirovany(self) -> None:
        directory = report_render.font_directory()
        assert (directory / "LICENSE_DEJAVU.txt").is_file()
        provenance = (directory / "PROVENANCE.md").read_text(encoding="utf-8")
        assert "DejaVu" in provenance
        assert "SHA-256" in provenance
        assert "закрыт" in provenance

    def test_shrift_pokryvaet_kirillitsu_kazakhskie_bukvy_i_tenge(self) -> None:
        """Проверка по таблице cmap: без этих знаков отчёт нечитаем."""
        from reportlab.pdfbase.ttfonts import TTFont

        font = TTFont("Probe", str(report_render.font_directory() / "DejaVuSans.ttf"))
        cmap = font.face.charToGlyph
        for symbol in "АБЯабяЁёҚқҒғӘәІі№₸—«»":
            assert ord(symbol) in cmap, f"в шрифте нет знака {symbol!r}"

    def test_otsutstvie_shrifta_daet_ponyatnuyu_oshibku(
        self, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
    ) -> None:
        monkeypatch.setattr(report_render, "_fonts_registered", False)
        monkeypatch.setattr(report_render, "font_directory", lambda: tmp_path)

        with pytest.raises(PdfUnavailableError) as excinfo:
            report_render.register_pdf_fonts()

        assert "DejaVuSans.ttf" in str(excinfo.value)
        assert "PROVENANCE" in str(excinfo.value)


# --- Имя файла ---------------------------------------------------------------


class TestImyaFayla:
    def test_translitatsiya_dayot_uznavaemoe_imya(self) -> None:
        fallback = _ascii_fallback("Отчёт по территории.docx")
        assert fallback.isascii()
        assert "Otchet" in fallback

    def test_zagolovok_soderzhit_oba_imeni(self) -> None:
        header = _content_disposition("Сводный отчёт.xlsx")
        assert 'filename="' in header
        assert "filename*=UTF-8''" in header
        # Процентное кодирование обязано быть ASCII целиком.
        assert header.split("filename*=UTF-8''")[1].isascii()

    def test_kavychki_ne_lomayut_zagolovok(self) -> None:
        """Кавычка в имени файла разорвала бы заголовок пополам."""
        header = _content_disposition('Отчёт "особый".docx')
        assert header.count('"') == 2
