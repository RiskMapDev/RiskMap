"""Тесты слоя 8.5 — субсидии и господдержка.

Главное, что здесь доказывается:

* балл считает наш код, а не книга (в книге его попросту нет — формулы без кэша);
* ведущие нули идентификаторов переживают импорт;
* пустая ячейка индикатора не превращается в ноль;
* и при всём этом наши числа сходятся с контрольными значениями аудита.

Golden-тесты сравнивают посчитанное **нашим** кодом с числами из книги. Там,
где расхождение неустранимо, оно зафиксировано явным ожиданием с объяснением
причины, а не ослабленным допуском, — см. `docs/assumptions-and-gaps.md`.
"""

from __future__ import annotations

from collections import Counter
from decimal import Decimal
from typing import Any, ClassVar

import pytest

from app.importers.subsidies_8_5 import (
    BOOK_CONTROL_VALUES,
    HEADER_ROW,
    SHEET_DISTRICTS,
    SHEET_RECIPIENTS,
    TERRITORY_ALIASES,
    ImportResult,
    build_orm_rows,
    default_resolver,
    factors_payload,
    find_source,
    normalize_bid_number,
    normalize_xin,
    parse_datetime,
    parse_flag,
    program_code,
    run_import,
)
from app.risk.core import RiskLevel
from app.risk.layers.subsidies import (
    INDICATOR_CODES,
    MODEL_CODE,
    REFERENCE_THRESHOLDS,
    REFERENCE_WEIGHTS,
    book_indicator_values,
    build_spec,
    build_thresholds,
    indicator_values,
    risk_exposure,
    score,
    score_as_book,
)
from app.services.territory_resolver import ResolutionStatus

# --- Контрольные значения аудита ---------------------------------------------

ЭТАЛОННЫЕ_ПОЛУЧАТЕЛИ: dict[str, float] = {
    "780702300265": 72.095,
    "190640012953": 64.885,
    "150240027741": 53.115,
    "060240001134": 31.175,
    "551016301150": 16.530,
}

#: Суммарная риск-экспозиция в семантике книги (пустая ячейка = 0).
КНИЖНАЯ_ЭКСПОЗИЦИЯ = 20_393_585_538.27

#: Она же по методике проекта: у 66 получателей балл нормируется на доступный
#: вес, поэтому итог выше книжного. Расхождение объяснимо и задокументировано.
ПРОЕКТНАЯ_ЭКСПОЗИЦИЯ = 20_395_614_586.64

#: Сумма колонки «Риск-экспозиция» листа «Риск_районы». Не совпадает с суммой
#: по получателям: свод считает экспозицию от среднего риска, округлённого до
#: одного знака. Дефект книги, зафиксирован явно.
ЭКСПОЗИЦИЯ_СВОДА = 20_390_445_713


# --- Фикстуры -----------------------------------------------------------------


@pytest.fixture(scope="session")
def результат_импорта() -> ImportResult:
    """Разбор книги целиком — один раз на весь прогон.

    Чтение 21 521 строки достаточно дорого, чтобы не повторять его в каждом
    тесте, и достаточно важно, чтобы не подменять фикстурой-заглушкой.
    """
    try:
        find_source()
    except (FileNotFoundError, SystemExit) as exc:  # pragma: no cover — среда без исходников
        pytest.skip(f"Книга слоя 8.5 недоступна: {exc}")
    return run_import()


@pytest.fixture
def спецификация() -> Any:
    """Модель с эталонными весами — для тестов, не читающих книгу."""
    return build_spec(REFERENCE_WEIGHTS, REFERENCE_THRESHOLDS)


# --- Идентификаторы -----------------------------------------------------------


class TestИдентификаторы:
    """Ведущий ноль в БИН/ИИН — пятая часть связей с другими слоями."""

    @pytest.mark.parametrize(
        ("исходное", "ожидаемое"),
        [
            ("780702300265", "780702300265"),
            ("080340015131", "080340015131"),
            (80340015131, "080340015131"),  # число потеряло ведущий ноль
            (80340015131.0, "080340015131"),  # оно же через float
            ("  080340015131  ", "080340015131"),
        ],
    )
    def test_bin_vsegda_12_znakov(self, исходное: object, ожидаемое: str) -> None:
        assert normalize_xin(исходное) == ожидаемое
        assert len(normalize_xin(исходное)) == 12

    def test_nomer_zayavki_ne_dopolnyaetsya_do_fiksirovannoy_dliny(self) -> None:
        """Номер заявки бывает 12 и 14 знаков — дополнять его нельзя."""
        assert normalize_bid_number("023001015578") == "023001015578"
        assert normalize_bid_number("00230010155781") == "00230010155781"
        assert normalize_bid_number(23001015578) == "23001015578"

    def test_kod_programmy_ustoychiv(self) -> None:
        """Повторный импорт обязан давать тот же код — иначе будут дубли."""
        имя = "Заявка на получение субсидий на удешевление стоимости затрат на корма"
        assert program_code(имя) == program_code(имя + " ")
        assert program_code(имя) != program_code(имя + " (иная)")


class TestРазборЗначений:
    def test_data_bez_chasovogo_poyasa(self) -> None:
        значение = parse_datetime("2022-12-18T17:39:13")
        assert значение is not None
        assert значение.year == 2022
        assert значение.tzinfo is None

    def test_nerazbornaya_data_ne_ronyaet_import(self) -> None:
        assert parse_datetime("не дата") is None
        assert parse_datetime(None) is None

    def test_flag_da_i_pustota(self) -> None:
        assert parse_flag("да") is True
        assert parse_flag("Да") is True
        assert parse_flag(None) is False
        assert parse_flag("") is False


# --- Модель риска -------------------------------------------------------------


class TestМодельРиска:
    def test_summa_vesov_edinitsa(self) -> None:
        assert sum(REFERENCE_WEIGHTS.values()) == pytest.approx(1.0)

    def test_vesa_ne_summiruyushchiesya_v_edinitsu_otvergayutsya(self) -> None:
        плохие = {**REFERENCE_WEIGHTS, "s1": 0.5}
        with pytest.raises(ValueError, match="сумма весов"):
            build_spec(плохие, REFERENCE_THRESHOLDS)

    def test_nepolnyy_nabor_vesov_otvergaetsya(self) -> None:
        with pytest.raises(ValueError, match="не прочитаны веса"):
            build_spec({"s1": 1.0}, REFERENCE_THRESHOLDS)

    def test_otsutstvuyushchiy_porog_otvergaetsya(self) -> None:
        with pytest.raises(ValueError, match="не заданы пороги"):
            build_thresholds({RiskLevel.MEDIUM: 35.0})

    def test_serogo_urovnya_v_metodike_net(self, спецификация: Any) -> None:
        """В 8.5 порога полноты нет — это отличие от слоёв 8.6 и 8.7."""
        assert спецификация.min_completeness is None

    def test_porogi_35_55_75(self, спецификация: Any) -> None:
        assert спецификация.level_for(34.999) is RiskLevel.LOW
        assert спецификация.level_for(35.0) is RiskLevel.MEDIUM
        assert спецификация.level_for(54.999) is RiskLevel.MEDIUM
        assert спецификация.level_for(55.0) is RiskLevel.HIGH
        assert спецификация.level_for(74.999) is RiskLevel.HIGH
        assert спецификация.level_for(75.0) is RiskLevel.CRITICAL

    def test_formula_sovpadaet_s_listom_metodiki(self, спецификация: Any) -> None:
        """R = 100 × Σ(wₖ·sₖ) при полностью измеренных индикаторах."""
        значения = {"s1": 1.0, "s2": 0.473, "s3": 0.0, "s4": 1.0, "s5": 1.0}
        ожидаемое = 100 * (0.3 * 1.0 + 0.15 * 0.473 + 0.2 * 0.0 + 0.2 * 1.0 + 0.15 * 1.0)

        результат = score(спецификация, значения)
        assert результат.score == pytest.approx(ожидаемое, abs=0.001)
        assert результат.completeness == pytest.approx(1.0)

    def test_risk_ekspozitsiya(self) -> None:
        assert risk_exposure(65_829_614, 72.095) == pytest.approx(47_459_860.21, abs=0.01)

    def test_ekspozitsiya_bez_balla_ne_nol(self) -> None:
        """Ноль означал бы «денег под риском нет», а не «оценить не смогли»."""
        assert risk_exposure(1_000_000, None) is None
        assert risk_exposure(None, 50.0) is None


class TestПустаяЯчейкаНеНоль:
    """Центральное требование ядра, применённое к слою 8.5."""

    ЗНАЧЕНИЯ: ClassVar[dict[str, float | None]] = {
        "s1": None,  # район неизвестен — концентрация не рассчитана
        "s2": 0.4,
        "s3": 0.0,
        "s4": 0.0,
        "s5": 0.0,
    }

    def test_proektnaya_semantika_ne_izmeryaet(self, спецификация: Any) -> None:
        значения = indicator_values(self.ЗНАЧЕНИЯ)
        assert значения["s1"].is_measured is False
        assert "район" in значения["s1"].note

        результат = score(спецификация, self.ЗНАЧЕНИЯ)
        # Знаменатель — 0.7, а не 1.0: неизмеренный вес выброшен из обеих частей.
        assert результат.completeness == pytest.approx(0.7)
        assert результат.score == pytest.approx(100 * 0.15 * 0.4 / 0.7, abs=0.001)

    def test_knizhnaya_semantika_schitaet_nulyom(self, спецификация: Any) -> None:
        значения = book_indicator_values(self.ЗНАЧЕНИЯ)
        assert значения["s1"].is_measured is True
        assert значения["s1"].value == 0.0
        assert "ноль" in значения["s1"].note

        результат = score_as_book(спецификация, self.ЗНАЧЕНИЯ)
        assert результат.completeness == pytest.approx(1.0)
        assert результат.score == pytest.approx(100 * 0.15 * 0.4, abs=0.001)

    def test_dve_semantiki_dayut_raznyy_ball(self, спецификация: Any) -> None:
        """Ровно та цена, которую книга платит за неразличение пустоты и нуля."""
        наш = score(спецификация, self.ЗНАЧЕНИЯ)
        книжный = score_as_book(спецификация, self.ЗНАЧЕНИЯ)
        assert наш.score is not None and книжный.score is not None
        assert наш.score > книжный.score

    def test_polnye_dannye_dayut_odinakovyy_ball(self, спецификация: Any) -> None:
        """Расхождение возникает только там, где ячейка пуста."""
        значения: dict[str, float | None] = dict.fromkeys(INDICATOR_CODES, 0.5)
        assert score(спецификация, значения).score == pytest.approx(
            score_as_book(спецификация, значения).score
        )


class TestСправочникТерриторий:
    def test_vse_napisaniya_knigi_sopostavlyayutsya(self) -> None:
        сопоставитель = default_resolver()
        for написание in TERRITORY_ALIASES:
            результат = сопоставитель.resolve(написание)
            assert результат.status is ResolutionStatus.RESOLVED, написание

    def test_neopoznannoe_nazvanie_ne_ugadyvaetsya(self) -> None:
        результат = default_resolver().resolve("Несуществующий район")
        assert результат.status is ResolutionStatus.NOT_FOUND
        assert результат.territory_code is None

    def test_pustoe_nazvanie_otlichaetsya_ot_nenaydennogo(self) -> None:
        """66 получателей без района — это отсутствие данных, а не опечатка."""
        assert default_resolver().resolve(None).status is ResolutionStatus.EMPTY

    def test_napisaniya_ne_perekryvayut_drug_druga(self) -> None:
        assert default_resolver().ambiguous_names == ()


# --- Golden: сверка с книгой ---------------------------------------------------


@pytest.mark.golden
@pytest.mark.slow
class TestМетодикаИзКниги:
    def test_vesa_prochitany_iz_yacheek_b9_b13(self, результат_импорта: ImportResult) -> None:
        assert dict(результат_импорта.methodology.weights) == pytest.approx(
            dict(REFERENCE_WEIGHTS)
        )

    def test_summa_vesov_edinitsa(self, результат_импорта: ImportResult) -> None:
        """`Методика!B14` — формула без кэша, сумму обязаны считать мы."""
        assert результат_импорта.methodology.weight_sum == pytest.approx(1.0, abs=1e-9)

    def test_porogi_prochitany_iz_yacheek_b16_b18(self, результат_импорта: ImportResult) -> None:
        assert результат_импорта.methodology.thresholds == dict(REFERENCE_THRESHOLDS)

    def test_model_bez_serogo_urovnya(self, результат_импорта: ImportResult) -> None:
        assert результат_импорта.spec.min_completeness is None
        assert результат_импорта.spec.code == MODEL_CODE


@pytest.mark.golden
@pytest.mark.slow
class TestФормулыБезКэша:
    """Ловушка №1 книги: у всех 10 240 формул нет кэшированного значения."""

    def test_kolonki_r_uroven_ekspozitsiya_chitayutsya_kak_none(self) -> None:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        try:
            путь = find_source()
        except (FileNotFoundError, SystemExit) as exc:  # pragma: no cover
            pytest.skip(f"Книга слоя 8.5 недоступна: {exc}")

        книга = load_workbook(путь, data_only=True, read_only=True)
        try:
            строки = list(книга[SHEET_RECIPIENTS].iter_rows(values_only=True))
        finally:
            книга.close()

        данные = [строка for строка in строки[HEADER_ROW:] if строка[0] is not None]
        assert len(данные) == 3413
        # Колонки T, U, V — индексы 19, 20, 21.
        assert all(строка[19] is None for строка in данные)
        assert all(строка[20] is None for строка in данные)
        assert all(строка[21] is None for строка in данные)

    def test_ball_poyavlyaetsya_tolko_posle_nashego_rascheta(
        self, результат_импорта: ImportResult
    ) -> None:
        баллы = [строка.result.score for строка in результат_импорта.recipients]
        assert all(балл is not None for балл in баллы)
        assert max(б for б in баллы if б is not None) > 0


@pytest.mark.golden
@pytest.mark.slow
class TestОбъёмыИСуммы:
    def test_chislo_strok(self, результат_импорта: ImportResult) -> None:
        assert len(результат_импорта.recipients) == 3413
        assert len(результат_импорта.payments) == 21521

    def test_chislo_rayonov(self, результат_импорта: ImportResult) -> None:
        коды = {
            строка.territory_code
            for строка in результат_импорта.recipients
            if строка.territory_code
        }
        assert len(коды) == 24

    def test_summa_subsidiy(self, результат_импорта: ImportResult) -> None:
        по_получателям = sum(
            строка.total_amount for строка in результат_импорта.recipients
        )
        по_vyplatam = sum(строка.amount_total for строка in результат_импорта.payments)
        assert по_получателям == Decimal("67535553445")
        assert по_vyplatam == по_получателям

    def test_chislo_programm_i_vidov(self, результат_импорта: ImportResult) -> None:
        assert len(результат_импорта.programs) == 46
        виды = {p.animal_type for p in результат_импорта.programs}
        assert len(виды) == 8

    def test_flagi_listа_dannye(self, результат_импорта: ImportResult) -> None:
        выплаты = результат_импорта.payments
        assert sum(1 for p in выплаты if p.flag_paid_before_decision) == 1209
        assert sum(1 for p in выплаты if p.flag_abnormal_lag) == 1052
        assert sum(1 for p in выплаты if p.flag_amount_outlier) == 882

    def test_vedushchie_nuli_perezhili_import(self, результат_импорта: ImportResult) -> None:
        сколько = sum(
            1 for строка in результат_импорта.recipients if строка.xin.startswith("0")
        )
        assert сколько == 70
        assert all(len(строка.xin) == 12 for строка in результат_импорта.recipients)


@pytest.mark.golden
@pytest.mark.slow
class TestРаспределениеРиска:
    def test_maksimalnyy_ball(self, результат_импорта: ImportResult) -> None:
        баллы = [
            строка.result.score
            for строка in результат_импорта.recipients
            if строка.result.score is not None
        ]
        assert max(баллы) == pytest.approx(72.095, abs=0.001)

    def test_raspredelenie_po_urovnyam(self, результат_импорта: ImportResult) -> None:
        """0 критических / 2 высоких / 67 средних / 3344 низких.

        Совпадает с книгой, несмотря на другую трактовку пустой s1: все 66
        получателей без района остаются низкорисковыми и при нормировке на
        доступный вес.
        """
        уровни = результат_импорта.level_counts
        assert уровни[RiskLevel.CRITICAL] == 0
        assert уровни[RiskLevel.HIGH] == 2
        assert уровни[RiskLevel.MEDIUM] == 67
        assert уровни[RiskLevel.LOW] == 3344
        assert уровни[RiskLevel.UNKNOWN] == 0

    @pytest.mark.parametrize(("бин", "ожидаемый_r"), sorted(ЭТАЛОННЫЕ_ПОЛУЧАТЕЛИ.items()))
    def test_etalonnye_poluchateli(
        self, результат_импорта: ImportResult, бин: str, ожидаемый_r: float
    ) -> None:
        строка = next(с for с in результат_импорта.recipients if с.xin == бин)
        assert строка.result.score == pytest.approx(ожидаемый_r, abs=0.001)
        assert строка.book_result.score == pytest.approx(ожидаемый_r, abs=0.001)

    def test_oba_vysokoriskovykh_naydeny(self, результат_импорта: ImportResult) -> None:
        высокие = [
            с for с in результат_импорта.recipients if с.result.level is RiskLevel.HIGH
        ]
        assert {с.xin for с in высокие} == {"780702300265", "190640012953"}

    def test_ekspozitsiya_etalonnykh_strok(self, результат_импорта: ImportResult) -> None:
        по_бин = {с.xin: с for с in результат_импорта.recipients}
        assert по_бин["780702300265"].exposure == pytest.approx(47_459_860.21, abs=0.01)
        assert по_бин["190640012953"].exposure == pytest.approx(360_415_435.81, abs=0.01)


@pytest.mark.golden
@pytest.mark.slow
class TestРискЭкспозиция:
    def test_knizhnaya_ekspozitsiya_skhoditsya(self, результат_импорта: ImportResult) -> None:
        """Сверка с контрольным числом аудита — 20 393 585 538.27 ₸."""
        assert результат_импорта.book_total_exposure == pytest.approx(
            КНИЖНАЯ_ЭКСПОЗИЦИЯ, abs=0.5
        )

    def test_proektnaya_ekspozitsiya_vyshe_knizhnoy(
        self, результат_импорта: ImportResult
    ) -> None:
        """Расхождение с книгой — намеренное и ровно на 66 строках.

        Книга засчитывает пустую s1 как ноль и тем занижает балл получателям
        без района. Мы нормируем на доступный вес, поэтому итог выше на
        2 029 048.37 ₸ (0.01 %). Тест не ослаблен: обе величины проверены
        точно, расхождение объяснено в docs/assumptions-and-gaps.md.
        """
        assert результат_импорта.total_exposure == pytest.approx(
            ПРОЕКТНАЯ_ЭКСПОЗИЦИЯ, abs=0.5
        )
        разница = результат_импорта.total_exposure - результат_импорта.book_total_exposure
        assert разница == pytest.approx(2_029_048.37, abs=1.0)

    def test_raskhozhdenie_tolko_tam_gde_pusta_s1(
        self, результат_импорта: ImportResult
    ) -> None:
        """Две семантики расходятся только на строках с неизмеренной s1.

        Их 66, но балл отличается у 24: у остальных 42 все прочие индикаторы
        нулевые, и нормировка нуля на любой знаменатель даёт ноль. Число 24 —
        не подгонка, а следствие данных, поэтому проверяется точно.
        """
        разошлись = [
            с
            for с in результат_импорта.recipients
            if с.result.score is not None
            and с.book_result.score is not None
            and abs(с.result.score - с.book_result.score) > 1e-9
        ]
        assert all(с.indicators["s1"] is None for с in разошлись)
        assert len(разошлись) == 24

        совпали = [
            с
            for с in результат_импорта.recipients
            if с.indicators["s1"] is None and с not in разошлись
        ]
        assert len(совпали) == 42
        assert all(с.book_result.score == pytest.approx(0.0) for с in совпали)


@pytest.mark.golden
@pytest.mark.slow
class TestНеизмеренноеИТерритории:
    def test_66_poluchateley_bez_rayona(self, результат_импорта: ImportResult) -> None:
        без_района = [
            с for с in результат_импорта.recipients if с.territory_code is None
        ]
        assert len(без_района) == 66
        assert all(с.territory_status is ResolutionStatus.EMPTY for с in без_района)

    def test_96_vyplat_bez_rayona(self, результат_импорта: ImportResult) -> None:
        assert sum(1 for p in результат_импорта.payments if p.territory_code is None) == 96

    def test_zhurnal_kachestva_fiksiruet_neizmerennyy_indikator(
        self, результат_импорта: ImportResult
    ) -> None:
        замечания = [
            з for з in результат_импорта.issues if з.code == "indicator_not_measured"
        ]
        assert len(замечания) == 66
        assert {з.context["indicator"] for з in замечания} == {"s1"}

    def test_zhurnal_kachestva_fiksiruet_territoriyu(
        self, результат_импорта: ImportResult
    ) -> None:
        замечания = [
            з for з in результат_импорта.issues if з.code == "territory_not_resolved"
        ]
        # 66 получателей + 96 выплат: у каждой строки своя запись в журнале.
        assert len(замечания) == 162

    def test_ni_odno_nazvanie_ne_ostalos_neopoznannym(
        self, результат_импорта: ImportResult
    ) -> None:
        """Все 24 написания книги распознаны; несопоставленное — только пустое."""
        отчёт = результат_импорта.territory_report
        assert отчёт.not_found == ()
        assert отчёт.ambiguous == ()
        assert отчёт.empty == 66

    def test_polnota_66_strok_ravna_0_7(self, результат_импорта: ImportResult) -> None:
        без_района = [с for с in результат_импорта.recipients if с.territory_code is None]
        assert all(
            с.result.completeness == pytest.approx(0.7) for с in без_района
        )

    def test_ostalnye_stroki_polny(self, результат_импорта: ImportResult) -> None:
        полные = [с for с in результат_импорта.recipients if с.territory_code is not None]
        assert len(полные) == 3347
        assert all(с.result.completeness == pytest.approx(1.0) for с in полные)


@pytest.mark.golden
@pytest.mark.slow
class TestСверкаИмпорта:
    def test_svodka_skhoditsya_s_kontrolnymi_znacheniyami(
        self, результат_импорта: ImportResult
    ) -> None:
        расхождения = результат_импорта.reconciliation.compare(BOOK_CONTROL_VALUES)
        assert расхождения == (), f"расхождения с книгой: {расхождения}"

    def test_svodka_lovit_podmenu_kontrolnogo_znacheniya(
        self, результат_импорта: ImportResult
    ) -> None:
        """Сверка обязана сообщать о расхождении, а не молчать."""
        расхождения = результат_импорта.reconciliation.compare({"recipients": 9999})
        assert len(расхождения) == 1
        assert расхождения[0]["metric"] == "recipients"
        assert расхождения[0]["delta"] == pytest.approx(3413 - 9999)


@pytest.mark.golden
@pytest.mark.slow
class TestДефектыКниги:
    """Расхождения, которые не наши, — но замалчивать их нельзя."""

    def test_svod_po_rayonam_zanizhaet_ekspozitsiyu(self) -> None:
        """Лист «Риск_районы» не сходится с расчётом по получателям.

        Свод считает экспозицию как «сумма × средний взвешенный риск / 100»,
        округляя средний риск до одного знака, а по получателям она считается
        построчно. Отсюда 3 139 825.27 ₸ разницы на 20.4 млрд — 0.015 %.
        Подгонять расчёт под свод нельзя: построчная величина точнее.
        """
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        try:
            путь = find_source()
        except (FileNotFoundError, SystemExit) as exc:  # pragma: no cover
            pytest.skip(f"Книга слоя 8.5 недоступна: {exc}")

        книга = load_workbook(путь, data_only=True, read_only=True)
        try:
            строки = [
                строка
                for строка in list(книга[SHEET_DISTRICTS].iter_rows(values_only=True))[
                    HEADER_ROW:
                ]
                if строка[0] is not None
            ]
        finally:
            книга.close()

        assert len(строки) == 24
        assert sum(строка[4] for строка in строки) == 3347  # получателей в своде
        assert sum(строка[3] for строка in строки) == 67_439_109_971  # сумма субсидий
        assert sum(строка[8] for строка in строки) == 2  # высокорисковых
        assert sum(строка[9] for строка in строки) == ЭКСПОЗИЦИЯ_СВОДА
        assert pytest.approx(
            3_139_825.27, abs=0.5
        ) == КНИЖНАЯ_ЭКСПОЗИЦИЯ - ЭКСПОЗИЦИЯ_СВОДА


# --- Отображение в ORM ---------------------------------------------------------


@pytest.mark.golden
@pytest.mark.slow
class TestОтображениеВORM:
    def test_stroki_prevrashchayutsya_v_obekty(self, результат_импорта: ImportResult) -> None:
        урезанный = ImportResult(
            source_path=результат_импорта.source_path,
            methodology=результат_импорта.methodology,
            spec=результат_импорта.spec,
            programs=результат_импорта.programs[:3],
            recipients=результат_импорта.recipients[:5],
            payments=(),
            issues=(),
            territory_report=результат_импорта.territory_report,
            reconciliation=результат_импорта.reconciliation,
        )
        объекты = build_orm_rows(урезанный)

        assert len(объекты["recipients"]) == 5
        первый = объекты["recipients"][0]
        assert первый.xin == "780702300265"
        assert первый.risk_score == pytest.approx(72.095, abs=0.001)
        assert первый.risk_level == "high"
        assert первый.book_risk_score == pytest.approx(72.095, abs=0.001)
        assert первый.territory_id is None  # справочник территорий не передан

    def test_rasshifrovka_soderzhit_vse_pyat_indikatorov(
        self, результат_импорта: ImportResult
    ) -> None:
        строка = next(
            с for с in результат_импорта.recipients if с.indicators["s1"] is None
        )
        расшифровка = factors_payload(строка.result)
        коды = [ф["code"] for ф in расшифровка["factors"]]
        assert коды == list(INDICATOR_CODES)

        неизмеренный = next(ф for ф in расшифровка["factors"] if ф["code"] == "s1")
        assert неизмеренный["measured"] is False
        assert неизмеренный["effect"] == "не измерено"
        assert расшифровка["completeness"] == pytest.approx(0.7)


@pytest.mark.golden
@pytest.mark.slow
def test_uroven_kazhdoy_stroki_sootvetstvuet_porogam(
    результат_импорта: ImportResult,
) -> None:
    """Никакой уровень не назначен в обход порогов 35/55/75."""
    спец = результат_импорта.spec
    несоответствия = Counter[str]()
    for строка in результат_импорта.recipients:
        балл = строка.result.score
        assert балл is not None
        if спец.level_for(балл) is not строка.result.level:
            несоответствия[строка.xin] += 1
    assert несоответствия == Counter()
