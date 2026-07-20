"""Тесты слоя 8.6 — инфраструктурные и инвестиционные проекты.

Слой 8.6 — единственный, где в одной книге лежат **две несвязанные популяции**:
проекты ГЧП (тип A, W_total = 110) и заключения строительной экспертизы
(тип B, W_total = 90). Половина тестов этого файла существует ровно затем,
чтобы попытка свести их в одну сущность немедленно роняла прогон.

Что здесь доказывается:

* две модели остаются двумя — разные коды, разные индикаторы, разный полный
  вес, и ни один ключ не связывает популяции (а «очевидный» ключ по номеру даёт
  198 ложных совпадений — это проверяется явно);
* ведущие нули регистрационного номера восстанавливаются: без `zfill(6)` джойн
  витрины с сырьём даёт 0 из 4842, с ним — 4842 из 4842;
* единица учёта типа B — заключение, а не объект: 111 строк с повторной
  экспертизой стоят за 52 различными объектами;
* у проектов ГЧП нет районной привязки, и на районную карту они не выводятся;
* пустая ячейка нигде не превращается в ноль;
* серый уровень остаётся официальным, даже когда балл посчитан.

Golden-тесты сравнивают посчитанное **нашим** кодом с числами книги и аудита
(`docs/audit/03-sloi-8-5-8-6-8-7.md`, разделы 2.9 и 2.10). Расхождений по
контрольным значениям нет: все 6165 строк воспроизводятся построчно.
"""

from __future__ import annotations

import datetime as dt
import re
from collections import Counter
from pathlib import Path
from typing import NamedTuple

import pytest
from sqlalchemy import CheckConstraint

from app.core.config import get_settings
from app.db.models.infrastructure import (
    ConstructionExpertiseObject,
    PppProject,
    ProjectEntity,
    ProjectEntityKind,
    ProjectParticipant,
    TerritoryPrecision,
)
from app.db.models.territory import TerritoryLevel
from app.importers.infrastructure_8_6 import (
    EXPECTED_CONTEST_COUNT,
    EXPECTED_CONTRACT_COUNT,
    EXPECTED_EXPERTISE_COUNT,
    EXPECTED_PPP_COUNT,
    EXPERTISE_CODES,
    PPP_CODES,
    REGISTRATION_NUMBER_LENGTH,
    SHEET_EXPERTISE_RAW,
    ExpertiseConclusionRow,
    ExpertisePopulation,
    PppPopulation,
    PppProjectRow,
    as_stored,
    clean_text,
    is_terminated,
    link_contracts_to_contests,
    normalize_registration_number,
    parse_number,
    parse_source_date,
    read_contests,
    read_contracts,
    read_expertise_book_rows,
    read_expertise_conclusions,
    read_ppp_book_rows,
    read_ppp_projects,
    read_source_dir,
    squash,
)
from app.risk.core import IndicatorValue, RiskLevel, RiskResult, aggregate_levels
from app.risk.layers.infrastructure import (
    A2_SCALE,
    A6_SCALE,
    B2_SCALE,
    DATA_ERROR_REASON,
    EXPERTISE_MODEL,
    MIN_COMPLETENESS_8_6,
    PPP_MODEL,
    THRESHOLDS_8_6,
    evaluate_expertise_conclusion,
    evaluate_ppp_project,
    expertise_significance_k,
    graded,
    ppp_significance_k,
    preliminary_level,
)
from app.services.layers import get_layer

# --- Контрольные значения аудита ---------------------------------------------

#: Тип A: № проекта → Risk Score книги (раздел 2.9, «5 эталонных строк»).
ЭТАЛОННЫЕ_ПРОЕКТЫ: dict[int, float] = {268: 85.7, 271: 85.7, 1180: 81.2, 241: 38.2, 213: 0.0}

#: Тип B: «Рег. №» витрины (со срезанными нулями) → Risk Score книги.
ЭТАЛОННЫЕ_ЗАКЛЮЧЕНИЯ: dict[str, float] = {
    "5617": 67.1,
    "7910": 65.0,
    "5545": 58.5,
    "4952": 44.4,
    "4976": 13.3,
}

#: Распределение типа A. Серых 64 = 59 «недостаточно данных» + 5 «ошибка в данных».
РАСПРЕДЕЛЕНИЕ_A: dict[RiskLevel, int] = {
    RiskLevel.LOW: 1066,
    RiskLevel.MEDIUM: 141,
    RiskLevel.HIGH: 44,
    RiskLevel.CRITICAL: 8,
    RiskLevel.UNKNOWN: 64,
}

#: Распределение типа B. Серых нет вовсе: у заключений полнота никогда не ниже 70/90.
РАСПРЕДЕЛЕНИЕ_B: dict[RiskLevel, int] = {
    RiskLevel.LOW: 4221,
    RiskLevel.MEDIUM: 567,
    RiskLevel.HIGH: 54,
    RiskLevel.CRITICAL: 0,
    RiskLevel.UNKNOWN: 0,
}

СЕРЫХ_БЕЗ_ДАННЫХ = 59
СЕРЫХ_ПО_ОШИБКЕ_ДАННЫХ = 5

#: Уровни книги в терминах ядра. Два серых уровня книги — один уровень модели:
#: причина серости хранится отдельно, в `override_applied` и `is_preliminary`.
УРОВНИ_КНИГИ: dict[str, RiskLevel] = {
    "низкий": RiskLevel.LOW,
    "средний": RiskLevel.MEDIUM,
    "высокий": RiskLevel.HIGH,
    "критический": RiskLevel.CRITICAL,
    "серый (недостаточно данных)": RiskLevel.UNKNOWN,
    "серый (ошибка в данных)": RiskLevel.UNKNOWN,
}


# --- Вспомогательное ----------------------------------------------------------


def значения_расходятся(наше: float | None, книжное: float | None) -> bool:
    """Расходится ли наше значение индикатора с книжным.

    `None` и `0.0` считаются разными значениями — именно это различие книга
    местами теряет, и сверка обязана его видеть.
    """
    if (наше is None) != (книжное is None):
        return True
    if наше is None or книжное is None:
        return False
    return abs(наше - книжное) > 1e-9


def уровень_вне_порогов(оценка: RiskResult) -> bool:
    """Уровень назначен не по порогам и не по объявленному исключению."""
    if оценка.score is None:
        return оценка.level is not RiskLevel.UNKNOWN
    if оценка.level is RiskLevel.UNKNOWN:
        return not (оценка.is_preliminary or bool(оценка.override_applied))
    return PPP_MODEL.level_for(оценка.score) is not оценка.level


# --- Фабрики строк ------------------------------------------------------------


def проект(**поля: object) -> PppProjectRow:
    """Синтетический проект ГЧП — для тестов, не читающих книгу.

    Значения по умолчанию нейтральны: ни один индикатор не срабатывает, поэтому
    тест меняет ровно то поле, которое проверяет, и не зависит от остальных.
    """
    значения: dict[str, object] = {
        "row_number": 8,
        "registry_number": 1,
        "region_raw": "Алматинская область",
        "project_level": "местный",
        "title": "Строительство школы",
        "object_kind": "",
        "status_raw": "Реализуемые: эксплуатация",
        "capacity": "",
        "sector": "",
        "initiative_kind": "",
        "contract_date": None,
        "construction_start": None,
        "construction_end": None,
        "operation_start": None,
        "operation_end": None,
        "contract_kind": "",
        "government_partner_raw": "Акимат Алматинской области",
        "private_partner_raw": 'ТОО "Партнёр"',
        "cost_initial": None,
        "investments": None,
        "government_participation_form": "",
        "source_url": "",
    }
    значения.update(поля)
    return PppProjectRow(**значения)  # type: ignore[arg-type]


def заключение(**поля: object) -> ExpertiseConclusionRow:
    """Синтетическое заключение экспертизы."""
    значения: dict[str, object] = {
        "row_number": 2,
        "registration_number": "005617",
        "author_supervision_status": "Согласован",
        "work_kind": "",
        "design_stage": "",
        "industry": "",
        "object_kind": "",
        "title": "Строительство школы",
        "customer_raw": "ГУ Отдел строительства",
        "designer_raw": 'ТОО "Проектировщик"',
        "location_raw": "Республика Казахстан, Алматинская область;",
        "conclusion_number": "01-0144/26",
        "issue_date_raw": "12.06.2026",
        "capacity": "",
        "capacity_unit": "",
        "has_cost_estimate": True,
        "technological_complexity": "",
        "responsibility_level": "2 уровень",
        "hazard_class": "3 класс опасности",
        "category": "",
        "efficiency_class": "",
        "funding_source": "",
        "expertise_place": "",
        "full_set_cost": "",
        "external_id": 1_000_001,
    }
    значения.update(поля)
    return ExpertiseConclusionRow(**значения)  # type: ignore[arg-type]


# --- Фикстуры -----------------------------------------------------------------


class РасчётГЧП(NamedTuple):
    """Разобранная и посчитанная популяция типа A."""

    строки: tuple[PppProjectRow, ...]
    выборка: PppPopulation
    оценки: dict[int, RiskResult]


class РасчётЭкспертизы(NamedTuple):
    """Разобранная и посчитанная популяция типа B."""

    строки: tuple[ExpertiseConclusionRow, ...]
    выборка: ExpertisePopulation
    оценки: dict[str, RiskResult]


@pytest.fixture(scope="session")
def путь_к_книге() -> Path:
    """Путь к книге 8.6. Имя файла в NFD — искать можно только листингом."""
    try:
        return read_source_dir(get_settings().source_data_dir)
    except (FileNotFoundError, SystemExit, OSError) as exc:  # pragma: no cover — среда без книг
        pytest.skip(f"Книга слоя 8.6 недоступна: {exc}")


@pytest.fixture(scope="session")
def расчёт_гчп(путь_к_книге: Path) -> РасчётГЧП:
    """1323 проекта, посчитанные один раз на весь прогон."""
    строки = tuple(read_ppp_projects(путь_к_книге))
    выборка = PppPopulation(строки)
    оценки = {
        строка.registry_number: evaluate_ppp_project(
            выборка.indicator_values(строка),
            significance_k=выборка.significance_k(строка),
            has_data_error=строка.has_date_error,
        )
        for строка in строки
    }
    return РасчётГЧП(строки, выборка, оценки)


@pytest.fixture(scope="session")
def расчёт_экспертизы(путь_к_книге: Path) -> РасчётЭкспертизы:
    """4842 заключения, посчитанные один раз на весь прогон."""
    строки = tuple(read_expertise_conclusions(путь_к_книге))
    выборка = ExpertisePopulation(строки)
    оценки = {
        строка.registration_number: evaluate_expertise_conclusion(
            выборка.indicator_values(строка),
            significance_k=выборка.significance_k(строка),
        )
        for строка in строки
    }
    return РасчётЭкспертизы(строки, выборка, оценки)


# --- Две популяции остаются двумя ---------------------------------------------


class TestДвеПопуляцииНеОдна:
    """Главное требование слоя: тип A и тип B нельзя связывать общим ключом.

    Аудит проверил все мыслимые кандидаты в ключ (раздел 2.8) и не нашёл ни
    одного. Эти тесты обязаны падать, если кто-то попробует свести популяции
    в одну сущность — потому что после такого объединения половина полей
    каждой популяции превращается в «не измерено», и полнота обрушивается
    у всех 6165 объектов.
    """

    def test_eto_dve_modeli_a_ne_odna(self) -> None:
        assert PPP_MODEL.code != EXPERTISE_MODEL.code
        assert PPP_MODEL.code == "8.6-ppp"
        assert EXPERTISE_MODEL.code == "8.6-expertise"

    def test_nabory_indikatorov_ne_peresekayutsya(self) -> None:
        коды_a = {индикатор.code for индикатор in PPP_MODEL.indicators}
        коды_b = {индикатор.code for индикатор in EXPERTISE_MODEL.indicators}

        assert коды_a == set(PPP_CODES)
        assert коды_b == set(EXPERTISE_CODES)
        assert коды_a & коды_b == set()

    def test_polnyy_ves_metodiki_raznyy(self) -> None:
        """110 против 90. Общий знаменатель означал бы общую методику."""
        assert PPP_MODEL.total_weight == pytest.approx(110.0)
        assert EXPERTISE_MODEL.total_weight == pytest.approx(90.0)

    def test_vesa_tipa_a_sovpadayut_s_reestrom_indikatorov(self) -> None:
        ожидаемые = {
            "A1": 25.0,
            "A2": 20.0,
            "A3": 15.0,
            "A4": 15.0,
            "A5": 15.0,
            "A6": 10.0,
            "A7": 10.0,
        }
        фактические = {и.code: и.weight for и in PPP_MODEL.indicators}
        assert фактические == ожидаемые

    def test_vesa_tipa_b_sovpadayut_s_reestrom_indikatorov(self) -> None:
        ожидаемые = {"B1": 20.0, "B2": 20.0, "B3": 15.0, "B4": 15.0, "B5": 10.0, "B6": 10.0}
        фактические = {и.code: и.weight for и in EXPERTISE_MODEL.indicators}
        assert фактические == ожидаемые

    def test_indikatory_odnoy_populyatsii_ne_prinimayutsya_drugoy(self) -> None:
        """Ядро обязано отказать, а не посчитать «как получится».

        Именно этот отказ ловит попытку прогнать заключение через модель ГЧП
        (и наоборот) — самый вероятный способ незаметно склеить популяции.
        """
        with pytest.raises(KeyError, match="неизвестных индикаторов"):
            evaluate_ppp_project({"B1": IndicatorValue(code="B1", value=1.0)})

        with pytest.raises(KeyError, match="неизвестных индикаторов"):
            evaluate_expertise_conclusion({"A1": IndicatorValue(code="A1", value=1.0)})

    def test_tipov_obekta_rovno_dva(self) -> None:
        """Третье значение дискриминатора — это и есть попытка слить популяции."""
        assert set(ProjectEntityKind) == {
            ProjectEntityKind.PPP_PROJECT,
            ProjectEntityKind.EXPERTISE_CONCLUSION,
        }

    def test_supertip_ne_soderzhit_predmetnykh_poley(self) -> None:
        """Как только предметное поле переедет в супертип, начнётся склеивание."""
        общие = set(ProjectEntity.__table__.columns.keys())
        предметные = {
            "registry_number",
            "registration_number",
            "private_partner_key",
            "government_partner_key",
            "customer_key",
            "designer_key",
            "object_identity_key",
        }
        assert общие & предметные == set()

    def test_predmetnye_polya_zhivut_v_podtipakh(self) -> None:
        assert "registry_number" in PppProject.__table__.columns
        assert "registration_number" in ConstructionExpertiseObject.__table__.columns
        assert "registration_number" not in PppProject.__table__.columns
        assert "registry_number" not in ConstructionExpertiseObject.__table__.columns

    def test_uchastnik_ne_ssylaetsya_na_organizatsii_sloya_8_7(self) -> None:
        """БИН участника заполнен у единиц записей — жёсткая ссылка невозможна.

        «Мягкая» подстановка по наименованию дала бы 2 совпадения из 809 × 769,
        то есть шум, выданный за связь.
        """
        цели = {
            ключ.target_fullname.split(".")[0]
            for ключ in ProjectParticipant.__table__.foreign_keys
        }
        # Единственная предметная ссылка — на объект слоя 8.6; остальное —
        # служебная ссылка на задание импорта из ProvenanceMixin.
        assert "project_entities" in цели
        assert "organizations" not in цели
        assert "persons" not in цели


# --- Нормализация значений ----------------------------------------------------


class TestРегистрационныйНомер:
    """Ловушка №2 книги: ведущие нули срезаны в витрине, но есть в сырье."""

    @pytest.mark.parametrize(
        ("исходное", "ожидаемое"),
        [
            ("5617", "005617"),
            ("005617", "005617"),
            ("1", "000001"),
            (5617, "005617"),
            ("  5617  ", "005617"),
        ],
    )
    def test_nomer_dopolnyaetsya_do_shesti_znakov(self, исходное: object, ожидаемое: str) -> None:
        assert normalize_registration_number(исходное) == ожидаемое
        assert len(normalize_registration_number(исходное)) == REGISTRATION_NUMBER_LENGTH

    def test_slishkom_dlinnyy_nomer_otvergaetsya(self) -> None:
        """Семь знаков — это не «плохо отформатированный» номер, а другая сущность."""
        with pytest.raises(ValueError, match="длиннее"):
            normalize_registration_number("1234567")

    def test_nechislovoy_nomer_otvergaetsya(self) -> None:
        with pytest.raises(ValueError, match="не является числом"):
            normalize_registration_number("01-0144/26")


class TestРазборЗначений:
    def test_nol_v_summe_eto_ne_zapolneno(self) -> None:
        """Ноль в стоимости и инвестициях означает «поле пустое».

        Отличить «инвестиций ноль» от «не заполнено» по данным нельзя, и книга
        трактует ноль как отсутствие. Если бы `parse_number` вернул 0.0,
        индикатор A6 у 348 проектов посчитался бы как «роста нет», хотя про
        рост вообще ничего не известно.
        """
        assert parse_number(0) is None
        assert parse_number(0.0) is None
        assert parse_number(1_500_000) == pytest.approx(1_500_000.0)
        assert parse_number(None) is None
        assert parse_number("1 500 000") is None

    def test_flag_ne_schitaetsya_chislom(self) -> None:
        """`True` — это флаг, а не единица тенге."""
        assert parse_number(True) is None

    @pytest.mark.parametrize(
        ("исходное", "ожидаемое"),
        [
            (dt.date(2024, 5, 1), dt.date(2024, 5, 1)),
            (dt.datetime(2024, 5, 1, 12, 30), dt.date(2024, 5, 1)),
            (2024, dt.date(2024, 12, 31)),
        ],
    )
    def test_data_razbiraetsya(self, исходное: object, ожидаемое: dt.date) -> None:
        assert parse_source_date(исходное) == ожидаемое

    def test_god_bez_mesyatsa_eto_konets_goda(self) -> None:
        """Плановый срок «до конца 2024» просрочен не с первого января."""
        assert parse_source_date(2024) == dt.date(2024, 12, 31)

    def test_stroka_kompozit_ne_ugadyvaetsya(self) -> None:
        """Выбрать одну дату из композита можно только гаданием.

        Недоступный индикатор дешевле, чем угаданная дата: угадав, мы получим
        просрочку там, где её, возможно, нет.
        """
        композит = "04.11.2019 (осн.договор); 10.10.2024 (ДС №004)"
        assert parse_source_date(композит) is None
        assert parse_source_date(None) is None

    def test_registr_statusa_ne_teryaet_dve_stroki(self) -> None:
        """«Расторгнут» 150 раз и «расторгнут» 2 раза — всего 152, а не 150."""
        assert is_terminated("Расторгнут") is True
        assert is_terminated("расторгнут") is True
        assert is_terminated("  Расторгнут  ") is True
        assert is_terminated("Реализуемые: эксплуатация") is False
        assert is_terminated(None) is False

    def test_nan_eto_pustaya_yacheyka(self) -> None:
        assert clean_text("nan") == ""
        assert clean_text("NaN") == ""
        assert clean_text(None) == ""
        assert clean_text("  ТОО «А»  ") == "ТОО «А»"


class TestКлючиГруппировки:
    """Книга нормализует контрагентов непоследовательно, и это приходится повторять."""

    def test_chastnyy_partner_svorachivaetsya_do_bukv_i_tsifr(self) -> None:
        assert squash('ТОО "Zhetysu Energy"') == squash("ТОО Zhetysu Energy")
        assert squash('ТОО «А-Б», ') == squash("ТОО АБ")

    def test_gospartner_khranitsya_kak_v_istochnike(self) -> None:
        """Обрезка хвостовых пробелов меняет индикатор A4 у проекта № 908.

        Госпартнёр с хвостовым пробелом образует отдельную группу, в которой
        меньше трёх проектов, и A4 становится недоступным. Выглядит небрежно,
        но это буквальное поведение книги, и «причёсывание» ключа меняет число.
        """
        assert as_stored("Акимат ") == "Акимат "
        assert as_stored("Акимат ") != as_stored("Акимат")
        assert as_stored(None) == ""


# --- Модель риска -------------------------------------------------------------


class TestМодельРиска86:
    def test_porogi_25_50_75(self) -> None:
        """Пороги 8.6 отличаются от порогов 8.5 (35/55/75), и это не разнобой.

        Несогласованность порогов между слоями — зафиксированный факт аудита
        (раздел 4). Приведение их к общему виду здесь было бы подменой методики.
        """
        assert [граница for граница, _ in THRESHOLDS_8_6] == [0.0, 25.0, 50.0, 75.0]

        for модель in (PPP_MODEL, EXPERTISE_MODEL):
            assert модель.level_for(24.999) is RiskLevel.LOW
            assert модель.level_for(25.0) is RiskLevel.MEDIUM
            assert модель.level_for(49.999) is RiskLevel.MEDIUM
            assert модель.level_for(50.0) is RiskLevel.HIGH
            assert модель.level_for(74.999) is RiskLevel.HIGH
            assert модель.level_for(75.0) is RiskLevel.CRITICAL

    def test_seryy_uroven_pri_polnote_nizhe_poloviny(self) -> None:
        assert MIN_COMPLETENESS_8_6 == 0.5
        assert PPP_MODEL.min_completeness == 0.5
        assert EXPERTISE_MODEL.min_completeness == 0.5

    def test_vse_istochniki_podklyucheny(self) -> None:
        """В отличие от 8.7, здесь неподключённых индикаторов нет."""
        for модель in (PPP_MODEL, EXPERTISE_MODEL):
            assert модель.available_total_weight == pytest.approx(модель.total_weight)

    def test_versiya_modeli_zapisana(self) -> None:
        """Оценка ссылается на версию, иначе правка веса перепишет историю."""
        результат = evaluate_ppp_project({"A1": IndicatorValue(code="A1", value=1.0)})
        assert результат.model_code == PPP_MODEL.code
        assert результат.model_version == PPP_MODEL.version


class TestГрадуированныеШкалы:
    def test_neizmerennaya_velichina_ostaetsya_neizmerennoy(self) -> None:
        """`graded(None)` обязан вернуть `None`, а не «риска нет»."""
        assert graded(None, A2_SCALE) is None
        assert graded(None, B2_SCALE) is None

    @pytest.mark.parametrize(
        ("доля", "ожидаемое"),
        [(0.5, 1.0), (0.49, 0.6), (0.3, 0.6), (0.29, 0.0), (0.0, 0.0)],
    )
    def test_shkala_a2_po_porogham_50_i_30(self, доля: float, ожидаемое: float) -> None:
        assert graded(доля, A2_SCALE) == pytest.approx(ожидаемое)

    @pytest.mark.parametrize(
        ("отношение", "ожидаемое"),
        [(1.5, 1.0), (1.49, 0.5), (1.2, 0.5), (1.19, 0.0)],
    )
    def test_shkala_a6_po_porogham_1_5_i_1_2(self, отношение: float, ожидаемое: float) -> None:
        assert graded(отношение, A6_SCALE) == pytest.approx(ожидаемое)

    @pytest.mark.parametrize(
        ("заключений", "ожидаемое"),
        [(1.0, 0.0), (2.0, 0.6), (3.0, 1.0), (7.0, 1.0)],
    )
    def test_shkala_b2_schitaet_zaklyucheniya_po_obektu(
        self, заключений: float, ожидаемое: float
    ) -> None:
        assert graded(заключений, B2_SCALE) == pytest.approx(ожидаемое)


class TestКоэффициентЗначимости:
    def test_diapazon_1_00_1_30(self) -> None:
        assert ppp_significance_k(top_quartile_cost=False, republican_level=False) == 1.0
        assert ppp_significance_k(top_quartile_cost=True, republican_level=False) == 1.15
        assert ppp_significance_k(top_quartile_cost=False, republican_level=True) == 1.15
        assert ppp_significance_k(top_quartile_cost=True, republican_level=True) == 1.30

    def test_tip_b_schitaet_svoi_priznaki(self) -> None:
        assert expertise_significance_k(hazard_class_1_2=False, responsibility_level_1=False) == 1.0
        assert expertise_significance_k(hazard_class_1_2=True, responsibility_level_1=True) == 1.30

    def test_ball_ne_prevyshaet_sta(self) -> None:
        """min(100; S_norm × K) — иначе K вынес бы балл за шкалу."""
        значения = {код: IndicatorValue(code=код, value=1.0) for код in PPP_CODES}
        результат = evaluate_ppp_project(значения, significance_k=1.30)
        assert результат.normalized_score == pytest.approx(100.0)
        assert результат.score == pytest.approx(100.0)


@pytest.mark.golden
class TestПроверочныеПримерыМетодики:
    """Примеры с листа «Методика Risk Score» — сверка без чтения витрины.

    Книгу здесь читать не нужно: значения индикаторов взяты из самого листа
    методики, и тест проверяет, что наша формула даёт ровно тот балл, который
    методика объявляет правильным.
    """

    def test_proekt_268_iz_yacheyki_b18(self) -> None:
        значения = {
            "A1": IndicatorValue(code="A1", value=1.0),
            "A2": IndicatorValue(code="A2", value=0.6),
            "A3": IndicatorValue(code="A3", value=1.0),
            "A4": IndicatorValue(code="A4", value=1.0),
            "A5": IndicatorValue(code="A5", value=1.0),
            "A6": IndicatorValue(code="A6", value=0.0),
            "A7": IndicatorValue(code="A7", value=0.0),
        }
        результат = evaluate_ppp_project(значения, significance_k=1.15)

        assert результат.raw_score == pytest.approx(82.0)
        assert результат.available_weight == pytest.approx(110.0)
        assert результат.normalized_score == pytest.approx(74.545, abs=0.001)
        assert результат.score == pytest.approx(85.7, abs=0.05)
        assert результат.level is RiskLevel.CRITICAL

    def test_zaklyuchenie_7910_iz_yacheyki_b19(self) -> None:
        """B6 не измерен — и знаменатель 80, а не 90. Это и есть суть шага 3."""
        значения = {
            "B1": IndicatorValue(code="B1", value=1.0),
            "B2": IndicatorValue(code="B2", value=0.0),
            "B3": IndicatorValue(code="B3", value=0.0),
            "B4": IndicatorValue(code="B4", value=1.0),
            "B5": IndicatorValue(code="B5", value=0.5),
            "B6": IndicatorValue(code="B6", value=None),
        }
        результат = evaluate_expertise_conclusion(значения, significance_k=1.30)

        assert результат.raw_score == pytest.approx(40.0)
        assert результат.available_weight == pytest.approx(80.0)
        assert результат.normalized_score == pytest.approx(50.0)
        assert результат.score == pytest.approx(65.0)
        assert результат.completeness == pytest.approx(80.0 / 90.0)
        assert результат.level is RiskLevel.HIGH


# --- Отсутствие данных не превращается в ноль ---------------------------------


class TestПустотаНеНоль:
    """Центральное требование ядра, применённое к обеим популяциям слоя 8.6."""

    def test_neizmerennyy_indikator_vybroshen_iz_obeikh_chastey_drobi(self) -> None:
        """Не измерено — значит ни в числителе, ни в знаменателе.

        Если бы неизмеренный A5 попал в сумму нулём, балл упал бы с 100 до 86.4
        и проект выглядел бы благополучнее, чем он есть.
        """
        измеренные = {код: IndicatorValue(code=код, value=1.0) for код in PPP_CODES}
        измеренные["A5"] = IndicatorValue(code="A5", value=None, note="срок не разобран")
        результат = evaluate_ppp_project(измеренные)

        assert результат.available_weight == pytest.approx(95.0)
        assert результат.score == pytest.approx(100.0)
        assert результат.completeness == pytest.approx(95.0 / 110.0)

    def test_esli_by_pustotu_schitali_nulyom_ball_byl_by_nizhe(self) -> None:
        """Цена ошибки, названная числом."""
        с_пустотой = {код: IndicatorValue(code=код, value=1.0) for код in PPP_CODES}
        с_пустотой["A5"] = IndicatorValue(code="A5", value=None)
        с_нулём = {**с_пустотой, "A5": IndicatorValue(code="A5", value=0.0)}

        наш = evaluate_ppp_project(с_пустотой)
        ошибочный = evaluate_ppp_project(с_нулём)

        assert наш.score is not None and ошибочный.score is not None
        assert наш.score > ошибочный.score
        assert ошибочный.score == pytest.approx(100.0 * 95.0 / 110.0, abs=0.001)

    def test_nol_investitsiy_delaet_a6_neizmerennym(self) -> None:
        """Ноль в объёме инвестиций — «поле не заполнено», а не «роста нет»."""
        выборка = PppPopulation([проект(cost_initial=1_000.0, investments=None)])
        значение = выборка.a6(выборка.rows[0])

        assert значение.is_measured is False
        assert "не заполнены" in значение.note

    def test_nerazbornyy_srok_delaet_a5_neizmerennym(self) -> None:
        выборка = PppPopulation([проект(construction_end=None)])
        значение = выборка.a5(выборка.rows[0])

        assert значение.is_measured is False
        assert "не разобрано" in значение.note

    def test_neizvestnoe_nalichie_smety_delaet_b4_neizmerennym(self) -> None:
        """«Нет сведений» не равно «сметы нет»: второе — это риск, первое — незнание."""
        выборка = ExpertisePopulation([заключение(has_cost_estimate=None)])
        значение = выборка.b4(выборка.rows[0])

        assert значение.is_measured is False
        assert значение.value is None

    def test_smeta_est_eto_izmerennyy_nol(self) -> None:
        """Обратная сторона: подтверждённое отсутствие риска — измеренный ноль."""
        выборка = ExpertisePopulation([заключение(has_cost_estimate=True)])
        значение = выборка.b4(выборка.rows[0])

        assert значение.is_measured is True
        assert значение.value == pytest.approx(0.0)

    def test_esli_ne_izmereno_nichego_balla_net_a_ne_nol(self) -> None:
        """`None` и `0.0` — разные ответы: «не знаем» против «риска нет»."""
        пусто = {код: IndicatorValue(code=код, value=None) for код in EXPERTISE_CODES}
        результат = evaluate_expertise_conclusion(пусто)

        assert результат.score is None
        assert результат.raw_score is None
        assert результат.completeness == pytest.approx(0.0)
        assert результат.level is RiskLevel.UNKNOWN

    def test_prichina_neizmerennosti_dokhodit_do_rasshifrovki(self) -> None:
        """Пользователь должен видеть, почему индикатор не измерен."""
        значения = {код: IndicatorValue(code=код, value=0.0) for код in PPP_CODES}
        значения["A4"] = IndicatorValue(
            code="A4", value=None, note="у государственного партнёра меньше трёх проектов"
        )
        результат = evaluate_ppp_project(значения)

        неизмеренный = next(ф for ф in результат.unmeasured_factors if ф.code == "A4")
        assert неизмеренный.effect == "не измерено"
        assert "меньше трёх проектов" in неизмеренный.note
        assert неизмеренный.contribution is None


# --- Серый уровень ------------------------------------------------------------


class TestСерыйУровеньОфициален:
    """Балл рядом с серым уровнем информативен, но уровень остаётся серым."""

    @staticmethod
    def _мало_данных() -> RiskResult:
        """Проект №213 книги: измерены только A1 и A7, W_avail = 35 из 110."""
        значения: dict[str, IndicatorValue] = {
            код: IndicatorValue(code=код, value=None) for код in PPP_CODES
        }
        значения["A1"] = IndicatorValue(code="A1", value=0.0)
        значения["A7"] = IndicatorValue(code="A7", value=0.0)
        return evaluate_ppp_project(значения)

    def test_polnota_nizhe_poloviny_daet_seryy_uroven(self) -> None:
        результат = self._мало_данных()

        assert результат.available_weight == pytest.approx(35.0)
        assert результат.completeness == pytest.approx(35.0 / 110.0)
        assert результат.level is RiskLevel.UNKNOWN
        assert результат.is_preliminary is True

    def test_ball_pri_etom_sokhranyaetsya(self) -> None:
        """Скрыть балл значило бы потерять информацию; выдать его за уровень —
        соврать. Поэтому хранится и балл, и признак предварительности."""
        результат = self._мало_данных()
        assert результат.score is not None

    def test_predvaritelnyy_uroven_ne_stanovitsya_ofitsialnym(self) -> None:
        значения: dict[str, IndicatorValue] = {
            код: IndicatorValue(code=код, value=None) for код in PPP_CODES
        }
        значения["A1"] = IndicatorValue(code="A1", value=1.0)
        результат = evaluate_ppp_project(значения)

        # Балл 100 из 100, но измерен ровно один индикатор из семи.
        assert результат.score == pytest.approx(100.0)
        assert preliminary_level(PPP_MODEL, результат) is RiskLevel.CRITICAL
        assert результат.level is RiskLevel.UNKNOWN

    def test_v_agregate_po_urovnyu_takoy_obekt_eto_net_dannykh(self) -> None:
        """Решение заказчика: в фильтрах и сводках объект относится к «нет данных».

        Если бы агрегат брал предварительный уровень, дашборд показал бы
        критический объект там, где на деле нечем его оценить.
        """
        значения: dict[str, IndicatorValue] = {
            код: IndicatorValue(code=код, value=None) for код in PPP_CODES
        }
        значения["A1"] = IndicatorValue(code="A1", value=1.0)
        результат = evaluate_ppp_project(значения)

        распределение = aggregate_levels([результат])
        assert распределение[RiskLevel.UNKNOWN] == 1
        assert распределение[RiskLevel.CRITICAL] == 0

    def test_seryy_ne_uezzhaet_v_konets_spiska_kak_blagopoluchnyy(self) -> None:
        """Ранг −1: объект без данных не притворяется низкорисковым."""
        assert RiskLevel.UNKNOWN.order < RiskLevel.LOW.order
        assert RiskLevel.UNKNOWN.label_ru == "Нет данных"

    def test_oshibka_v_dannykh_daet_seryy_nezavisimo_ot_balla(self) -> None:
        """Окончание строительства раньше начала: балл посчитан, но опираться
        на него нельзя — сроки, из которых он выведен, противоречивы."""
        значения = {код: IndicatorValue(code=код, value=1.0) for код in PPP_CODES}
        результат = evaluate_ppp_project(значения, has_data_error=True)

        assert результат.score == pytest.approx(100.0)
        assert результат.completeness == pytest.approx(1.0)
        assert результат.level is RiskLevel.UNKNOWN
        assert результат.override_applied == DATA_ERROR_REASON

    def test_oshibka_v_dannykh_silnee_polnoty(self) -> None:
        """Полнота 100 % не «спасает» строку с логически невозможными сроками."""
        значения = {код: IndicatorValue(code=код, value=0.0) for код in PPP_CODES}
        обычный = evaluate_ppp_project(значения)
        с_ошибкой = evaluate_ppp_project(значения, has_data_error=True)

        assert обычный.level is RiskLevel.LOW
        assert с_ошибкой.level is RiskLevel.UNKNOWN

    def test_data_error_v_stroke_opredelyaetsya_po_srokam(self) -> None:
        плохой = проект(
            construction_start=dt.date(2024, 1, 1), construction_end=dt.date(2023, 1, 1)
        )
        хороший = проект(
            construction_start=dt.date(2023, 1, 1), construction_end=dt.date(2024, 1, 1)
        )
        неизвестный = проект(construction_start=dt.date(2023, 1, 1), construction_end=None)

        assert плохой.has_date_error is True
        assert хороший.has_date_error is False
        assert неизвестный.has_date_error is False


# --- Территория ---------------------------------------------------------------


class TestТерриторияТипаA:
    """У проектов ГЧП есть область и нет района — и выдумывать его нельзя."""

    def test_est_otdelnyy_uroven_tochnosti_oblast(self) -> None:
        """«Район не указан» и «источник не содержит района» — разные состояния.

        Различать их обязательно: от этого зависит, можно ли объект вообще
        показывать на районной карте.
        """
        assert TerritoryPrecision.REGION.value == "region"
        assert set(TerritoryPrecision) == {
            TerritoryPrecision.DISTRICT,
            TerritoryPrecision.REGION,
            TerritoryPrecision.NONE,
        }

    def test_gchp_ne_vyvoditsya_na_rayonnuyu_kartu(self) -> None:
        """65 проектов Алматинской области физически невозможно разложить по
        районам, и показать их на районной карте значит соврать."""
        слой = get_layer("infrastructure_ppp")
        assert слой.available_at(TerritoryLevel.REGION)
        assert not слой.available_at(TerritoryLevel.DISTRICT)
        assert слой.unavailability_reason(TerritoryLevel.DISTRICT)

    def test_ekspertiza_naoborot_zhivet_na_rayonnom_urovne(self) -> None:
        """Ещё одно различие популяций: у типа B район в источнике есть."""
        слой = get_layer("infrastructure_expertise")
        assert слой.available_at(TerritoryLevel.DISTRICT)

    def test_u_gchp_tochnost_privyazki_po_umolchaniyu_oblast(self) -> None:
        """Проект ГЧП, созданный без явного указания, получает «область».

        Раньше он получал «none», то есть выглядел как объект вообще без
        территории. Разница существенна: «известна только область» и
        «территория неизвестна» по-разному отвечают на вопрос, можно ли
        показывать объект на районной карте.
        """
        проект = PppProject(title="Тестовый проект ГЧП")

        assert проект.territory_precision is TerritoryPrecision.REGION

    def test_yavno_zadannaya_tochnost_ne_perezapisyvaetsya(self) -> None:
        проект = PppProject(
            title="Тестовый проект ГЧП",
            territory_precision=TerritoryPrecision.NONE,
        )

        assert проект.territory_precision is TerritoryPrecision.NONE

    def test_pravilo_derzhitsya_ogranicheniem_bazy(self) -> None:
        """Значения по умолчанию мало: массовая вставка идёт мимо конструктора.

        Поэтому «тип A ⇒ точность region» закреплено ограничением на таблице
        супертипа, где эта колонка физически и лежит.
        """
        ограничения = {
            c.name
            for c in ProjectEntity.__table__.constraints
            if isinstance(c, CheckConstraint)
        }

        assert "ck_ppp_project_territory_is_region" in ограничения


# --- Golden: объёмы и распределения -------------------------------------------


@pytest.mark.golden
@pytest.mark.slow
class TestОбъёмыКниги:
    def test_chislo_proektov_i_zaklyucheniy(
        self, расчёт_гчп: РасчётГЧП, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        assert len(расчёт_гчп.строки) == EXPECTED_PPP_COUNT == 1323
        assert len(расчёт_экспертизы.строки) == EXPECTED_EXPERTISE_COUNT == 4842

    def test_sluzhebnaya_stroka_numeratsii_otbroshena(self, расчёт_гчп: РасчётГЧП) -> None:
        """В сыром реестре 1324 строки: одна из них — нумерация колонок 1..22."""
        assert len(расчёт_гчп.строки) == 1324 - 1
        assert {строка.registry_number for строка in расчёт_гчп.строки} == set(range(1, 1324))

    def test_imya_lista_ekspertizy_s_dvumya_probelami(self, путь_к_книге: Path) -> None:
        """Обращение по «очевидному» имени с одним пробелом даёт KeyError."""
        from openpyxl import load_workbook

        книга = load_workbook(путь_к_книге, read_only=True, data_only=True)
        try:
            имена = set(книга.sheetnames)
        finally:
            книга.close()

        assert SHEET_EXPERTISE_RAW in имена
        assert "Данные Экспертиза инфр проект" not in имена
        assert SHEET_EXPERTISE_RAW.count("  ") == 1

    def test_konkursy_i_dogovory(self, путь_к_книге: Path) -> None:
        конкурсы = read_contests(путь_к_книге)
        договоры = read_contracts(путь_к_книге)

        assert len(конкурсы) == EXPECTED_CONTEST_COUNT == 514
        assert len(договоры) == EXPECTED_CONTRACT_COUNT == 12


@pytest.mark.golden
@pytest.mark.slow
class TestРаспределениеТипаA:
    def test_raspredelenie_po_urovnyam(self, расчёт_гчп: РасчётГЧП) -> None:
        распределение = aggregate_levels(list(расчёт_гчп.оценки.values()))
        assert распределение == РАСПРЕДЕЛЕНИЕ_A

    def test_seryy_delitsya_na_dve_prichiny(self, расчёт_гчп: РасчётГЧП) -> None:
        """59 «недостаточно данных» + 5 «ошибка в данных» = 64 серых.

        Причина серости хранится отдельно от уровня: пользователю нужно знать,
        нечем оценить объект или данные о нём противоречивы.
        """
        серые = [
            оценка
            for оценка in расчёт_гчп.оценки.values()
            if оценка.level is RiskLevel.UNKNOWN
        ]
        по_ошибке = [о for о in серые if о.override_applied == DATA_ERROR_REASON]
        по_полноте = [о for о in серые if о.is_preliminary]

        assert len(по_полноте) == СЕРЫХ_БЕЗ_ДАННЫХ == 59
        assert len(по_ошибке) == СЕРЫХ_ПО_ОШИБКЕ_ДАННЫХ == 5
        assert len(серые) == 64

    def test_maksimalnyy_ball(self, расчёт_гчп: РасчётГЧП) -> None:
        баллы = [о.score for о in расчёт_гчп.оценки.values() if о.score is not None]
        assert max(баллы) == pytest.approx(85.7, abs=0.05)

    @pytest.mark.parametrize(("номер", "ожидаемый_балл"), sorted(ЭТАЛОННЫЕ_ПРОЕКТЫ.items()))
    def test_etalonnye_proekty(
        self, расчёт_гчп: РасчётГЧП, номер: int, ожидаемый_балл: float
    ) -> None:
        оценка = расчёт_гчп.оценки[номер]
        assert оценка.score is not None
        assert round(оценка.score, 1) == pytest.approx(ожидаемый_балл, abs=0.05)

    def test_etalon_268_kriticheskiy_pri_polnote_sto(self, расчёт_гчп: РасчётГЧП) -> None:
        for номер in (268, 271):
            оценка = расчёт_гчп.оценки[номер]
            assert оценка.raw_score == pytest.approx(82.0)
            assert оценка.available_weight == pytest.approx(110.0)
            assert оценка.completeness == pytest.approx(1.0)
            assert оценка.level is RiskLevel.CRITICAL

    def test_etalon_213_seryy_a_ne_nizkiy(self, расчёт_гчп: РасчётГЧП) -> None:
        """Балл 0 при полноте 32 % — это «не знаем», а не «риска нет»."""
        оценка = расчёт_гчп.оценки[213]
        assert оценка.score == pytest.approx(0.0)
        assert оценка.available_weight == pytest.approx(35.0)
        assert оценка.level is RiskLevel.UNKNOWN
        assert оценка.is_preliminary is True
        assert preliminary_level(PPP_MODEL, оценка) is RiskLevel.LOW

    def test_rastorgnutykh_152(self, расчёт_гчп: РасчётГЧП) -> None:
        """Сравнение без свёртки регистра дало бы 150 — потеря двух строк."""
        расторгнуто = sum(1 for строка in расчёт_гчп.строки if строка.is_terminated)
        assert расторгнуто == 152
        assert расторгнуто / len(расчёт_гчп.строки) == pytest.approx(0.115, abs=0.001)

    def test_almatinskaya_oblast_65_proektov(self, расчёт_гчп: РасчётГЧП) -> None:
        """64 + 1: одна строка записана как «Республика Казахстан (Алматинская
        област» — наименование обрезано на 40 знаках."""
        точное = [с for с in расчёт_гчп.строки if с.region_raw == "Алматинская область"]
        обрезанное = [
            с
            for с in расчёт_гчп.строки
            if "Алматинская област" in с.region_raw and с.region_raw != "Алматинская область"
        ]

        assert len(точное) == 64
        assert len(обрезанное) == 1
        assert len(точное) + len(обрезанное) == 65

    def test_v_almatinskoy_oblasti_rastorgnuto_20_iz_65(self, расчёт_гчп: РасчётГЧП) -> None:
        область = [с for с in расчёт_гчп.строки if "Алматинская област" in с.region_raw]
        расторгнуто = [с for с in область if с.is_terminated]

        assert len(область) == 65
        assert len(расторгнуто) == 20
        assert len(расторгнуто) / len(область) == pytest.approx(0.308, abs=0.001)

    def test_regionov_tridtsat_napisaniy(self, расчёт_гчп: РасчётГЧП) -> None:
        """30 написаний на 20 областей — справочник алиасов обязателен."""
        assert len({с.region_raw for с in расчёт_гчп.строки}) == 30


@pytest.mark.golden
@pytest.mark.slow
class TestРаспределениеТипаB:
    def test_raspredelenie_po_urovnyam(self, расчёт_экспертизы: РасчётЭкспертизы) -> None:
        распределение = aggregate_levels(list(расчёт_экспертизы.оценки.values()))
        assert распределение == РАСПРЕДЕЛЕНИЕ_B

    def test_serykh_i_kriticheskikh_net(self, расчёт_экспертизы: РасчётЭкспертизы) -> None:
        """У заключений полнота никогда не падает ниже 70 из 90 — серых нет.

        Это не значит, что данные полны: это значит, что четыре из шести
        индикаторов типа B считаются по самой строке и потому доступны всегда.
        """
        оценки = list(расчёт_экспертизы.оценки.values())

        assert all(оценка.completeness >= 70.0 / 90.0 for оценка in оценки)
        assert [о for о in оценки if о.level is RiskLevel.UNKNOWN] == []
        assert [о for о in оценки if о.level is RiskLevel.CRITICAL] == []
        assert [о for о in оценки if о.is_preliminary] == []

    def test_maksimalnyy_ball(self, расчёт_экспертизы: РасчётЭкспертизы) -> None:
        баллы = [о.score for о in расчёт_экспертизы.оценки.values() if о.score is not None]
        assert max(баллы) == pytest.approx(67.1, abs=0.05)

    @pytest.mark.parametrize(("номер", "ожидаемый_балл"), sorted(ЭТАЛОННЫЕ_ЗАКЛЮЧЕНИЯ.items()))
    def test_etalonnye_zaklyucheniya(
        self, расчёт_экспертизы: РасчётЭкспертизы, номер: str, ожидаемый_балл: float
    ) -> None:
        """Ключ эталона записан как в витрине — со срезанными нулями.

        Найти строку по нему можно только после `zfill(6)`, и это ровно та
        операция, без которой джойн даёт ноль совпадений.
        """
        оценка = расчёт_экспертизы.оценки[номер.zfill(6)]
        assert оценка.score is not None
        assert round(оценка.score, 1) == pytest.approx(ожидаемый_балл, abs=0.05)

    def test_etalon_5617_schitaetsya_na_semidesyati(
        self, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        оценка = расчёт_экспертизы.оценки["005617"]
        assert оценка.raw_score == pytest.approx(47.0)
        assert оценка.available_weight == pytest.approx(70.0)
        assert оценка.completeness == pytest.approx(70.0 / 90.0)
        assert оценка.level is RiskLevel.HIGH

    def test_korrektirovok_psd_764(self, расчёт_экспертизы: РасчётЭкспертизы) -> None:
        """Признак ищется в полном наименовании: витрина обрезает его до 180
        знаков, и на обрезанном тексте он теряется у 172 строк."""
        корректировки = sum(1 for с in расчёт_экспертизы.строки if с.has_correction)
        assert корректировки == 764
        assert корректировки / len(расчёт_экспертизы.строки) == pytest.approx(0.158, abs=0.001)

    def test_bez_smetnoy_dokumentatsii_76(self, расчёт_экспертизы: РасчётЭкспертизы) -> None:
        без_сметы = sum(1 for с in расчёт_экспертизы.строки if с.has_cost_estimate is False)
        неизвестно = sum(1 for с in расчёт_экспертизы.строки if с.has_cost_estimate is None)

        assert без_сметы == 76
        # Поле заполнено у всех 4842 строк: «нет сведений» здесь не встречается,
        # и потому B4 измерен всюду. Проверяем явно, чтобы подмена «Да/Нет» на
        # пустоту в будущей выгрузке была замечена, а не размазана по баллам.
        assert неизвестно == 0

    def test_obektov_almatinskoy_oblasti_442(self, расчёт_экспертизы: РасчётЭкспертизы) -> None:
        область = [
            с for с in расчёт_экспертизы.строки if "Алматинская область" in с.location_raw
        ]
        assert len(область) == 442


@pytest.mark.golden
@pytest.mark.slow
class TestЕдиницаУчётаТипаB:
    """Ловушка №13 аудита: «111 объектов» — это 111 строк, а не объектов."""

    def test_strok_s_povtornoy_ekspertizoy_111(
        self, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        assert расчёт_экспертизы.выборка.rows_with_repeated_expertise() == 111

    def test_razlichnykh_obektov_za_nimi_52(self, расчёт_экспертизы: РасчётЭкспертизы) -> None:
        """Считать объекты по числу строк значит завысить результат вдвое."""
        assert расчёт_экспертизы.выборка.distinct_objects_with_repeated_expertise() == 52

    def test_stroki_i_obekty_eto_raznye_velichiny(
        self, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        """Тест ловит подмену агрегата: если кто-то посчитает объекты по строкам,
        два числа совпадут, и здесь это станет ошибкой."""
        выборка = расчёт_экспертизы.выборка
        assert выборка.rows_with_repeated_expertise() != (
            выборка.distinct_objects_with_repeated_expertise()
        )

    def test_edinitsa_ucheta_zaklyuchenie_a_ne_obekt(
        self, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        """4842 заключения стоят за 4783 различными объектами."""
        assert расчёт_экспертизы.выборка.distinct_objects() == 4783
        assert len(расчёт_экспертизы.строки) == 4842

    def test_tozhdestvo_obekta_eto_para_naimenovanie_zakazchik(self) -> None:
        первое = заключение(registration_number="000001", title='Школа "А"')
        второе = заключение(registration_number="000002", title="Школа А")
        третье = заключение(
            registration_number="000003", title="Школа А", customer_raw="Другой заказчик"
        )

        assert первое.object_identity_key == второе.object_identity_key
        assert первое.object_identity_key != третье.object_identity_key


# --- Golden: ведущие нули и джойн ---------------------------------------------


@pytest.mark.golden
@pytest.mark.slow
class TestВедущиеНулиРегНомера:
    """Ловушка №2 книги, проверенная на всех 4842 строках."""

    def test_v_syrom_reestre_nomera_shestiznachnye(
        self, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        assert all(len(с.registration_number) == 6 for с in расчёт_экспертизы.строки)
        assert len({с.registration_number for с in расчёт_экспертизы.строки}) == 4842

    def test_v_vitrine_nuli_srezany(self, путь_к_книге: Path) -> None:
        витрина = read_expertise_book_rows(путь_к_книге)
        assert len(витрина) == 4842
        assert any(len(строка.key) < 6 for строка in витрина)

    def test_bez_zfill_dzhoyn_daet_nol_iz_4842(
        self, путь_к_книге: Path, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        """Прямое пересечение витрины с сырьём не находит ни одной строки.

        Это не гипотеза, а измеренный факт: загрузчик, собранный «в лоб»,
        молча потеряет весь тип B.
        """
        сырьё = {с.registration_number for с in расчёт_экспертизы.строки}
        витрина = {строка.key for строка in read_expertise_book_rows(путь_к_книге)}

        assert len(витрина & сырьё) == 0

    def test_s_zfill_dzhoyn_daet_4842_iz_4842(
        self, путь_к_книге: Path, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        сырьё = {с.registration_number for с in расчёт_экспертизы.строки}
        витрина = {
            normalize_registration_number(строка.key)
            for строка in read_expertise_book_rows(путь_к_книге)
        }

        assert len(витрина & сырьё) == 4842
        assert витрина == сырьё


@pytest.mark.golden
@pytest.mark.slow
class TestОбщегоКлючаНет:
    """Раздел 2.8 аудита, проверенный кодом, а не пересказом.

    Каждый тест — отдельный кандидат в ключ, и каждый провалился. Если
    когда-нибудь появится настоящий ключ, эти тесты упадут, и это будет
    правильный сигнал: связку нужно вводить осознанно, а не обнаружить
    случайно.
    """

    def test_naimenovaniya_ne_peresekayutsya(
        self, расчёт_гчп: РасчётГЧП, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        проекты = {squash(с.title) for с in расчёт_гчп.строки if с.title}
        объекты = {squash(с.title) for с in расчёт_экспертизы.строки if с.title}

        assert len(проекты) == 1265
        assert len(объекты) == 4781
        assert проекты & объекты == set()

    def test_kontragenty_ne_peresekayutsya(
        self, расчёт_гчп: РасчётГЧП, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        """Аудит сообщал о 2 и 17 совпадениях под своей нормализацией.

        Под каноническими ключами проекта (`squash`) совпадений нет вовсе —
        то есть вывод «общего ключа нет» не ослаблен, а усилен. Расхождение
        с числами аудита объясняется тем, что аудит сопоставлял колонки витрин
        (наименования там обрезаны), а мы — сырые реестры.
        """
        частные = {
            squash(с.private_partner_raw) for с in расчёт_гчп.строки if с.private_partner_raw
        }
        государственные = {
            squash(с.government_partner_raw)
            for с in расчёт_гчп.строки
            if с.government_partner_raw.strip()
        }
        заказчики = {
            squash(с.customer_raw) for с in расчёт_экспертизы.строки if с.customer_raw.strip()
        }
        проектировщики = {
            squash(с.designer_raw) for с in расчёт_экспертизы.строки if с.designer_raw.strip()
        }

        assert частные & заказчики == set()
        assert частные & проектировщики == set()
        assert государственные & заказчики == set()

    def test_naivnyy_dzhoyn_po_nomeru_daet_198_lozhnykh_sovpadeniy(
        self, расчёт_гчп: РасчётГЧП, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        """Самый опасный «очевидный» ключ — номер строки.

        № проекта ГЧП пробегает 1..1323, регистрационный номер заключения —
        шестизначный. Дополните первый нулями, и 198 пар «совпадут»: проект №1
        станет заключением 000001. Ни одна из этих пар не описывает один и тот
        же объект. Тест держит это число на виду именно затем, чтобы такой
        джойн никто не завёл «потому что сходится».
        """
        номера_проектов = {f"{с.registry_number:06d}" for с in расчёт_гчп.строки}
        номера_заключений = {с.registration_number for с in расчёт_экспертизы.строки}

        assert len(номера_проектов & номера_заключений) == 198

    def test_bin_v_reestre_ekspertizy_otsutstvuet(
        self, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        """0 из 4842 — аффилированность по ТЗ 9.3 здесь недоказуема."""
        бин = re.compile(r"\b\d{12}\b")
        с_бином = [
            с
            for с in расчёт_экспертизы.строки
            if бин.search(с.customer_raw) or бин.search(с.designer_raw)
        ]
        assert с_бином == []

    def test_edinstvennaya_svyazka_lezhit_vnutri_kontura_gchp(
        self, путь_к_книге: Path
    ) -> None:
        """12 договоров из 12 находят свой конкурс — и наружу это не ведёт."""
        конкурсы = read_contests(путь_к_книге)
        договоры = read_contracts(путь_к_книге)
        связано, сироты = link_contracts_to_contests(договоры, конкурсы)

        assert len(связано) == 12
        assert сироты == []
        assert sum(1 for к in конкурсы if к.organizer_bin) == 514


# --- Golden: сверка с витринами книги -----------------------------------------


@pytest.mark.golden
@pytest.mark.slow
class TestСверкаСВитринойКниги:
    """Построчная сверка нашего расчёта с витринами риска — 6165 строк.

    Расхождений нет ни по одному индикатору, ни по коэффициенту значимости,
    ни по баллу. Это и есть доказательство, что модель воспроизводит методику,
    а не подгоняет итог.
    """

    def test_tip_a_ball_sovpadaet_postrochno(
        self, путь_к_книге: Path, расчёт_гчп: РасчётГЧП
    ) -> None:
        витрина = {строка.key: строка for строка in read_ppp_book_rows(путь_к_книге)}
        расхождения = [
            номер
            for номер, оценка in расчёт_гчп.оценки.items()
            if оценка.score is None
            or abs(round(оценка.score, 1) - round(витрина[str(номер)].score, 1)) > 0.05
        ]
        assert расхождения == []

    def test_tip_a_indikatory_i_k_sovpadayut_postrochno(
        self, путь_к_книге: Path, расчёт_гчп: РасчётГЧП
    ) -> None:
        витрина = {строка.key: строка for строка in read_ppp_book_rows(путь_к_книге)}
        расхождения: list[tuple[int, str]] = []
        for строка in расчёт_гчп.строки:
            эталон = витрина[str(строка.registry_number)]
            наши = расчёт_гчп.выборка.indicator_values(строка)
            if расчёт_гчп.выборка.significance_k(строка) != pytest.approx(эталон.significance_k):
                расхождения.append((строка.registry_number, "K"))
            расхождения.extend(
                (строка.registry_number, код)
                for код in PPP_CODES
                if значения_расходятся(наши[код].value, эталон.indicator_values[код])
            )
        assert расхождения == []

    def test_tip_a_uroven_sovpadaet_postrochno(
        self, путь_к_книге: Path, расчёт_гчп: РасчётГЧП
    ) -> None:
        витрина = {строка.key: строка for строка in read_ppp_book_rows(путь_к_книге)}
        расхождения = [
            номер
            for номер, оценка in расчёт_гчп.оценки.items()
            if УРОВНИ_КНИГИ[витрина[str(номер)].level] is not оценка.level
        ]
        assert расхождения == []

    def test_tip_b_ball_sovpadaet_postrochno(
        self, путь_к_книге: Path, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        витрина = {
            normalize_registration_number(строка.key): строка
            for строка in read_expertise_book_rows(путь_к_книге)
        }
        расхождения = [
            номер
            for номер, оценка in расчёт_экспертизы.оценки.items()
            if оценка.score is None
            or abs(round(оценка.score, 1) - round(витрина[номер].score, 1)) > 0.05
        ]
        assert расхождения == []

    def test_tip_b_indikatory_i_k_sovpadayut_postrochno(
        self, путь_к_книге: Path, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        витрина = {
            normalize_registration_number(строка.key): строка
            for строка in read_expertise_book_rows(путь_к_книге)
        }
        расхождения: list[tuple[str, str]] = []
        for строка in расчёт_экспертизы.строки:
            эталон = витрина[строка.registration_number]
            наши = расчёт_экспертизы.выборка.indicator_values(строка)
            if расчёт_экспертизы.выборка.significance_k(строка) != pytest.approx(
                эталон.significance_k
            ):
                расхождения.append((строка.registration_number, "K"))
            расхождения.extend(
                (строка.registration_number, код)
                for код in EXPERTISE_CODES
                if значения_расходятся(наши[код].value, эталон.indicator_values[код])
            )
        assert расхождения == []

    def test_tip_b_uroven_sovpadaet_postrochno(
        self, путь_к_книге: Path, расчёт_экспертизы: РасчётЭкспертизы
    ) -> None:
        витрина = {
            normalize_registration_number(строка.key): строка
            for строка in read_expertise_book_rows(путь_к_книге)
        }
        расхождения = [
            номер
            for номер, оценка in расчёт_экспертизы.оценки.items()
            if УРОВНИ_КНИГИ[витрина[номер].level] is not оценка.level
        ]
        assert расхождения == []

    def test_koeffitsient_znachimosti_prinimaet_tri_znacheniya(
        self, расчёт_гчп: РасчётГЧП
    ) -> None:
        """K: 1.00 → 991, 1.15 → 317, 1.30 → 15."""
        значения = Counter(
            расчёт_гчп.выборка.significance_k(строка) for строка in расчёт_гчп.строки
        )
        assert значения == Counter({1.0: 991, 1.15: 317, 1.30: 15})

    def test_uroven_kazhdoy_stroki_sootvetstvuet_porogam(self, расчёт_гчп: РасчётГЧП) -> None:
        """Никакой уровень не назначен в обход порогов 25/50/75.

        Исключение ровно одно и оно объявлено: серый уровень при недостаточной
        полноте либо при ошибке в данных.
        """
        нарушители = [
            номер
            for номер, оценка in расчёт_гчп.оценки.items()
            if уровень_вне_порогов(оценка)
        ]
        assert нарушители == []
