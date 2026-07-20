"""Мастер импорта данных: приём файла, сопоставление колонок, сухой прогон, откат.

Модуль закрывает раздел 15 ТЗ и трёхшаговый мастер с референса. Решения,
которые определили его устройство, стоит назвать явно — они не очевидны.

**Загруженный файл никогда не подменяет исходники.** Комплект книг ДЭР
(`settings.source_data_dir`) открыт только на чтение. Всё, что приходит через
мастер, ложится в отдельный каталог `settings.data_dir / "uploads"` под именем,
равным SHA-256 содержимого. Имя-по-содержимому даёт бесплатную дедупликацию и
исключает ситуацию, когда два пользователя затирают файлы друг друга
одинаковыми именами.

**Идемпотентность держится на детерминированном ключе, а не на проверке перед
вставкой.** Идентификатор строки — `stable_id(таблица, естественный ключ)`, тот
же механизм, что у загрузчиков книг. Повторное подтверждение того же файла
попадает в `ON CONFLICT (id) DO UPDATE` и обновляет ту же строку. Версия данных
в ключ намеренно **не** входит: половина целевых таблиц имеет ограничение
уникальности по естественному ключу без версии (`uq_supplier_bin`,
`uq_contract_source_id`, `uq_organization_bin`), и версионирование ключа
привело бы там не к новой строке, а к отказу вставки.

**Версия данных считается от содержимого файла, а не от числа запусков.** Тот
же файл, поданный дважды, получает ту же логическую версию — иначе «повторный
запуск не создаёт дублей» выполнялось бы на уровне строк и нарушалось на
уровне версий. Новое содержимое (другой SHA-256) поднимает версию слоя на
единицу.

**Откат не удаляет ничего.** Он снимает `is_current` со строк, записанных
заданием, и переводит задание в статус `rolled_back`. Строки, замечания к
качеству и сама запись задания остаются на месте: оценка, показанная вчера,
должна оставаться объяснимой и после отката.

**В обновление попадают только сопоставленные колонки.** Набор колонок
строится из фактического сопоставления, а не из полного перечня полей. Иначе
импорт двух колонок затирал бы NULL-ами всё, что до него загрузил
полноценный загрузчик книги.

**Значения по умолчанию проставляются только при вставке.** Обязательные
колонки, которых нет в сопоставлении (`model_version`, `territory_status`
и подобные), передаются в `immutable_columns` — так новая строка получает
корректное значение, а существующая не теряет посчитанное загрузчиком книги.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import unicodedata
import uuid
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from pathlib import Path
from typing import Any, Final

from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.base import utcnow
from app.db.models.access import AuditAction, User
from app.db.models.budget import BudgetProgram
from app.db.models.infrastructure import PppProject, ProjectEntity, ProjectEntityKind
from app.db.models.organization import Organization
from app.db.models.procurement import Contract, Supplier
from app.db.models.source import (
    DataQualityIssue,
    ImportJob,
    ImportStatus,
    IssueSeverity,
    SourceDataset,
    SourceFile,
)
from app.db.models.subsidy import SubsidyRecipient
from app.db.models.territory import PopulationStat
from app.importers.persistence import (
    IssueRecord,
    LayerJob,
    TerritoryIndex,
    bulk_upsert,
    jsonable,
    load_territory_index,
    stable_id,
    table_of,
)
from app.services import audit
from app.services.audit import RequestContext

# --- Пределы -----------------------------------------------------------------

ACCEPTED_EXTENSIONS: Final[frozenset[str]] = frozenset(
    {".xlsx", ".xlsm", ".xls", ".csv", ".json", ".geojson"}
)
"""Форматы из ТЗ 15.1 и из подписи под зоной перетаскивания на референсе."""

PREVIEW_ROWS: Final[int] = 20
"""Сколько строк показывается в предпросмотре шага 2."""

MAX_PERSISTED_ISSUES: Final[int] = 2_000
"""Потолок числа замечаний, записываемых к одному заданию.

Файл с ошибкой в каждой строке дал бы сотни тысяч записей журнала качества, из
которых пользователь прочитает первые двадцать. Остаток сворачивается в одно
итоговое замечание с числом отброшенных.
"""

BACKGROUND_ROW_THRESHOLD: Final[int] = 5_000
"""С какого объёма мастер предлагает фоновую обработку.

Порог, а не жёсткое правило: решение принимает вызывающая сторона, а число
здесь — то, начиная с которого запись перестаёт укладываться в отведённые ТЗ
пять секунд на запрос.
"""

PROGRESS_CHUNK: Final[int] = 500
"""Через сколько строк обновляется прогресс фоновой обработки."""

_KZ_BBOX: Final[tuple[float, float, float, float]] = (46.0, 40.0, 88.0, 56.0)
"""Приблизительная рамка Казахстана: запад, юг, восток, север.

Координата вне рамки — не обязательно ошибка, но почти всегда перепутанные
местами широта и долгота, поэтому уровень замечания — предупреждение.
"""


class ImportWizardError(Exception):
    """Отказ мастера, который нужно показать пользователю дословно.

    Несёт машинный код: интерфейсу нужно различать «файл слишком велик» и
    «формат не поддерживается», чтобы подсказать разное, а не показывать одну
    и ту же фразу «ошибка загрузки».
    """

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


# --- Канонические поля -------------------------------------------------------


class FieldType(StrEnum):
    """Тип канонического поля.

    От типа зависит и приведение значения, и проверка. Разведены `NUMBER` и
    `MONEY`, потому что деньги хранятся в `Decimal`: перевод суммы договора
    через `float` теряет тиын на больших значениях.
    """

    TEXT = "text"
    INTEGER = "integer"
    NUMBER = "number"
    MONEY = "money"
    DATE = "date"
    BOOL = "bool"
    XIN = "xin"
    """БИН или ИИН: ровно 12 цифр, ведущие нули значимы."""

    TERRITORY = "territory"
    LATITUDE = "latitude"
    LONGITUDE = "longitude"


@dataclass(frozen=True, slots=True)
class CanonicalField:
    """Поле Системы, на которое сопоставляется колонка файла."""

    code: str
    title: str
    type: FieldType
    required: bool = False
    aliases: tuple[str, ...] = ()
    """Написания в источниках. Нужны автосопоставлению на шаге 2."""

    hint: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "title": self.title,
            "type": str(self.type),
            "required": self.required,
            "hint": self.hint,
            "aliases": list(self.aliases),
        }


@dataclass(frozen=True, slots=True)
class TargetSpec:
    """Куда именно ложатся строки одного типа данных.

    Целей у типа данных может быть несколько, и порядок значим: договор
    ссылается на поставщика внешним ключом, поэтому поставщик пишется первым.
    """

    table_name: str
    model: type[Any]
    natural_key_fields: tuple[str, ...]
    columns: Mapping[str, str]
    """Код канонического поля → имя колонки."""

    insert_defaults: Mapping[str, Any] = field(default_factory=dict)
    """Значения обязательных колонок, которых нет в сопоставлении.

    Пишутся только при вставке: у существующей строки эти колонки уже
    заполнены загрузчиком книги, и перетирать их значением-заглушкой нельзя.
    """

    null_fallback: Mapping[str, Any] = field(default_factory=dict)
    """Чем заменить пустое значение в колонке NOT NULL, которую сопоставили."""

    territory_column: str | None = None
    territory_raw_column: str | None = None
    territory_resolution_column: str | None = None
    territory_required: bool = False

    parent_index: int | None = None
    parent_link_column: str | None = None
    """Колонка, в которую кладётся идентификатор родительской цели."""

    shares_parent_id: bool = False
    """Наследование таблиц: подтип делит первичный ключ с супертипом."""


@dataclass(frozen=True, slots=True)
class KindSpec:
    """Одна из шести плиток шага 1."""

    kind: DataKind
    title: str
    description: str
    layer_code: str
    fields: tuple[CanonicalField, ...]
    targets: tuple[TargetSpec, ...]
    territory_field: str | None = None
    note: str = ""
    """Честное предупреждение о границах загрузки — показывается на шаге 1."""

    logic_checks: tuple[LogicCheck, ...] = ()
    """Проверки логических противоречий.

    Уровень возвращает сама проверка: расхождение в разбивке населения на
    единицу — повод предупредить, а окончание строительства раньше начала
    делает строку непригодной. Приравнивать одно к другому значило бы либо
    терять данные, либо пропускать невозможное.
    """

    def field_by_code(self, code: str) -> CanonicalField | None:
        return next((item for item in self.fields if item.code == code), None)

    @property
    def required_codes(self) -> tuple[str, ...]:
        return tuple(item.code for item in self.fields if item.required)

    def as_dict(self) -> dict[str, Any]:
        return {
            "code": str(self.kind),
            "title": self.title,
            "description": self.description,
            "layer_code": self.layer_code,
            "note": self.note,
            "fields": [item.as_dict() for item in self.fields],
            "targets": [target.table_name for target in self.targets],
        }


class DataKind(StrEnum):
    """Шесть типов загружаемых данных — плитки шага 1 с референса."""

    PROCUREMENT = "procurement"
    BUDGET = "budget"
    SUBSIDIES = "subsidies"
    ORGANIZATIONS = "organizations"
    INFRASTRUCTURE = "infrastructure"
    SOCIOECONOMIC = "socioeconomic"


# --- Логические проверки -----------------------------------------------------

#: Проверка получает приведённые значения строки и возвращает
#: (уровень, код, сообщение) либо None, если противоречия нет.
LogicCheck = Callable[[Mapping[str, Any]], tuple[IssueSeverity, str, str] | None]


def _check_negative_money(values: Mapping[str, Any]) -> tuple[IssueSeverity, str, str] | None:
    """Отрицательная сумма — почти всегда перепутанный знак возврата."""
    for code, value in values.items():
        if isinstance(value, Decimal) and value < 0:
            return (
                IssueSeverity.ERROR,
                "negative_amount",
                f"Поле «{code}»: отрицательная сумма {value}.",
            )
    return None


def _check_construction_dates(values: Mapping[str, Any]) -> tuple[IssueSeverity, str, str] | None:
    start = values.get("construction_start")
    end = values.get("construction_end")
    if isinstance(start, date) and isinstance(end, date) and end < start:
        return (
            IssueSeverity.ERROR,
            "logical_contradiction",
            f"Окончание строительства {end:%d.%m.%Y} раньше начала {start:%d.%m.%Y}.",
        )
    return None


def _check_payments_consistency(values: Mapping[str, Any]) -> tuple[IssueSeverity, str, str] | None:
    total = values.get("total_amount")
    payments = values.get("payments_count")
    if isinstance(total, Decimal) and total > 0 and payments == 0:
        return (
            IssueSeverity.ERROR,
            "logical_contradiction",
            "Сумма поддержки больше нуля при нулевом числе выплат.",
        )
    return None


def _check_population_split(values: Mapping[str, Any]) -> tuple[IssueSeverity, str, str] | None:
    """Разбивка по полу должна складываться в общее число.

    Уровень — предупреждение, а не ошибка: статистика пересматривается, и
    расхождение в единицы человек встречается в законных выгрузках. Прятать
    его нельзя, но и отбрасывать строку из-за него — потеря данных.
    """
    total = values.get("total")
    male = values.get("male")
    female = values.get("female")
    if (
        isinstance(total, int)
        and isinstance(male, int)
        and isinstance(female, int)
        and male + female != total
    ):
        return (
            IssueSeverity.WARNING,
            "logical_contradiction",
            f"Мужчины ({male}) и женщины ({female}) в сумме дают {male + female}, "
            f"а всего указано {total}.",
        )
    return None


def _check_registration_future(values: Mapping[str, Any]) -> tuple[IssueSeverity, str, str] | None:
    reg = values.get("reg_date")
    if isinstance(reg, date) and reg > datetime.now(tz=UTC).date():
        return (
            IssueSeverity.ERROR,
            "logical_contradiction",
            f"Дата регистрации {reg:%d.%m.%Y} в будущем.",
        )
    return None


# --- Описание шести типов данных ---------------------------------------------


_PROCUREMENT = KindSpec(
    kind=DataKind.PROCUREMENT,
    title="Государственные закупки",
    description="Договоры и поставщики слоя 8.4.",
    layer_code="8.4",
    territory_field="territory_name",
    fields=(
        CanonicalField(
            "contract_id",
            "Идентификатор договора",
            FieldType.TEXT,
            required=True,
            aliases=("id договора", "номер договора", "contract id", "договор"),
        ),
        CanonicalField(
            "supplier_bin",
            "БИН поставщика",
            FieldType.XIN,
            required=True,
            aliases=("бин поставщика", "бин", "supplier bin"),
            hint="Ровно 12 цифр. Ведущие нули значимы.",
        ),
        CanonicalField(
            "supplier_name",
            "Наименование поставщика",
            FieldType.TEXT,
            required=True,
            aliases=("поставщик", "наименование поставщика", "supplier"),
        ),
        CanonicalField(
            "brief_content_ru",
            "Предмет договора",
            FieldType.TEXT,
            aliases=("предмет закупки", "краткое содержание", "наименование договора"),
        ),
        CanonicalField(
            "final_amount",
            "Сумма договора",
            FieldType.MONEY,
            aliases=("сумма", "сумма договора", "итоговая сумма", "цена договора"),
        ),
        CanonicalField(
            "planned_exec_date",
            "Плановый срок исполнения",
            FieldType.DATE,
            aliases=("плановая дата", "срок исполнения", "план исполнения"),
        ),
        CanonicalField(
            "actual_exec_date",
            "Фактический срок исполнения",
            FieldType.DATE,
            aliases=("фактическая дата", "факт исполнения"),
        ),
        CanonicalField(
            "contract_status",
            "Статус договора",
            FieldType.TEXT,
            aliases=("статус", "состояние договора"),
        ),
        CanonicalField(
            "territory_name",
            "Территория",
            FieldType.TERRITORY,
            aliases=("район", "регион", "область", "территория"),
        ),
    ),
    targets=(
        TargetSpec(
            table_name="suppliers",
            model=Supplier,
            natural_key_fields=("supplier_bin",),
            columns={"supplier_bin": "bin", "supplier_name": "name"},
            insert_defaults={
                "in_rnu_gz": False,
                "in_lzhepred_list": False,
                "no_physical_activity": False,
                "high_oked_diversity": False,
                "mass_address": False,
                "nominal_director": False,
            },
            territory_column="territory_id",
            territory_raw_column="district_source_name",
        ),
        TargetSpec(
            table_name="contracts",
            model=Contract,
            natural_key_fields=("contract_id",),
            columns={
                "contract_id": "contract_id",
                "brief_content_ru": "brief_content_ru",
                "final_amount": "final_amount",
                "planned_exec_date": "planned_exec_date",
                "actual_exec_date": "actual_exec_date",
                "contract_status": "contract_status",
            },
            insert_defaults={
                "is_terminated": False,
                "is_preliminary": False,
                "model_code": "8.4",
                # Версия модели обязательна, а риск при ручной загрузке не
                # считается: пометка честнее, чем чужой номер версии.
                "model_version": "manual-import",
            },
            territory_column="territory_id",
            territory_raw_column="district_source_name",
            parent_index=0,
            parent_link_column="supplier_id",
        ),
    ),
    logic_checks=(_check_negative_money,),
)


_BUDGET = KindSpec(
    kind=DataKind.BUDGET,
    title="Бюджетные данные",
    description="Справочник бюджетных программ слоя 8.3.",
    layer_code="8.3",
    note=(
        "Загружается справочник программ. Факты исполнения требуют восьми "
        "обязательных сумм (утверждено, уточнено, план, обязательства, "
        "кассовое исполнение); подставлять вместо отсутствующих сумм нули "
        "запрещено, поэтому они грузятся загрузчиком книги 8.3."
    ),
    fields=(
        CanonicalField(
            "program_code",
            "Код программы",
            FieldType.TEXT,
            aliases=("код", "код программы", "бп"),
        ),
        CanonicalField(
            "program_name",
            "Наименование программы",
            FieldType.TEXT,
            required=True,
            aliases=("наименование", "программа", "наименование программы"),
        ),
        CanonicalField(
            "program_level",
            "Уровень в иерархии",
            FieldType.INTEGER,
            aliases=("уровень", "level"),
            hint="1 — корневая программа, глубже — подпрограммы и специфики.",
        ),
    ),
    targets=(
        TargetSpec(
            table_name="budget_programs",
            model=BudgetProgram,
            natural_key_fields=("program_code", "program_name"),
            columns={
                "program_code": "code",
                "program_name": "name",
                "program_level": "level",
            },
            insert_defaults={"is_leaf": False},
            null_fallback={"level": 1},
        ),
    ),
)


_SUBSIDIES = KindSpec(
    kind=DataKind.SUBSIDIES,
    title="Субсидии и поддержка",
    description="Получатели субсидий слоя 8.5.",
    layer_code="8.5",
    territory_field="territory_name",
    fields=(
        CanonicalField(
            "xin",
            "ИИН/БИН получателя",
            FieldType.XIN,
            required=True,
            aliases=("иин", "бин", "иин/бин", "иин получателя"),
            hint="Персональные данные: показывается по роли, в журнал не пишется.",
        ),
        CanonicalField(
            "name",
            "Наименование получателя",
            FieldType.TEXT,
            required=True,
            aliases=("получатель", "наименование", "фио"),
        ),
        CanonicalField(
            "director_name",
            "Руководитель",
            FieldType.TEXT,
            aliases=("руководитель", "директор"),
        ),
        CanonicalField(
            "territory_name",
            "Территория",
            FieldType.TERRITORY,
            aliases=("район", "область", "территория"),
        ),
        CanonicalField(
            "total_amount",
            "Сумма поддержки",
            FieldType.MONEY,
            required=True,
            aliases=("сумма", "сумма субсидий", "итого"),
        ),
        CanonicalField(
            "payments_count",
            "Число выплат",
            FieldType.INTEGER,
            aliases=("выплат", "количество выплат"),
        ),
        CanonicalField(
            "programs_count",
            "Число программ",
            FieldType.INTEGER,
            aliases=("программ", "количество программ"),
        ),
    ),
    targets=(
        TargetSpec(
            table_name="subsidy_recipients",
            model=SubsidyRecipient,
            natural_key_fields=("xin",),
            columns={
                "xin": "xin",
                "name": "name",
                "director_name": "director_name",
                "total_amount": "total_amount",
                "payments_count": "payments_count",
                "programs_count": "programs_count",
            },
            insert_defaults={
                "model_code": "8.5",
                "model_version": "manual-import",
                # Уровень риска при ручной загрузке не рассчитан. «Нет данных» —
                # честное состояние; ноль означал бы «риска нет».
                "risk_level": "unknown",
                "risk_completeness": 0.0,
            },
            null_fallback={"payments_count": 0, "programs_count": 0},
            territory_column="territory_id",
            territory_raw_column="territory_name_raw",
            territory_resolution_column="territory_resolution",
        ),
    ),
    logic_checks=(_check_negative_money, _check_payments_consistency),
)


_ORGANIZATIONS = KindSpec(
    kind=DataKind.ORGANIZATIONS,
    title="Хозяйствующие субъекты",
    description="Организации слоя 8.7.",
    layer_code="8.7",
    territory_field="territory_name",
    fields=(
        CanonicalField(
            "bin",
            "БИН",
            FieldType.XIN,
            required=True,
            aliases=("бин", "bin", "бизнес-идентификационный номер"),
        ),
        CanonicalField(
            "name",
            "Наименование",
            FieldType.TEXT,
            required=True,
            aliases=("наименование", "организация", "название"),
        ),
        CanonicalField(
            "full_name",
            "Полное наименование",
            FieldType.TEXT,
            aliases=("полное наименование",),
        ),
        CanonicalField(
            "reg_date",
            "Дата регистрации",
            FieldType.DATE,
            aliases=("дата регистрации", "зарегистрирован"),
        ),
        CanonicalField(
            "oked_main",
            "Основной ОКЭД",
            FieldType.TEXT,
            aliases=("окэд", "оквэд", "основной окэд"),
        ),
        CanonicalField(
            "employees_count",
            "Численность работников",
            FieldType.INTEGER,
            aliases=("численность", "работников", "сотрудников"),
        ),
        CanonicalField(
            "territory_name",
            "Территория",
            FieldType.TERRITORY,
            aliases=("район", "область", "территория"),
        ),
    ),
    targets=(
        TargetSpec(
            table_name="organizations",
            model=Organization,
            natural_key_fields=("bin",),
            columns={
                "bin": "bin",
                "name": "name",
                "full_name": "full_name",
                "reg_date": "reg_date",
                "oked_main": "oked_main",
                "employees_count": "employees_count",
            },
            insert_defaults={
                "territory_status": "not_determined",
                "is_category_a": False,
                "risk_is_preliminary": False,
            },
            territory_column="territory_id",
        ),
    ),
    logic_checks=(_check_registration_future,),
)


_INFRASTRUCTURE = KindSpec(
    kind=DataKind.INFRASTRUCTURE,
    title="Инфраструктурные проекты",
    description="Проекты ГЧП слоя 8.6.",
    layer_code="8.6",
    territory_field="territory_name",
    note=(
        "Плитка на референсе одна, а в книге 8.6 две несвязанные популяции. "
        "Мастер грузит проекты ГЧП; заключения экспертизы имеют другой "
        "естественный ключ и грузятся отдельно."
    ),
    fields=(
        CanonicalField(
            "registry_number",
            "Номер в реестре",
            FieldType.INTEGER,
            required=True,
            aliases=("номер", "реестровый номер", "№"),
        ),
        CanonicalField(
            "title",
            "Наименование проекта",
            FieldType.TEXT,
            required=True,
            aliases=("наименование", "проект", "название проекта"),
        ),
        CanonicalField(
            "sector",
            "Отрасль",
            FieldType.TEXT,
            aliases=("отрасль", "сфера", "сектор"),
        ),
        CanonicalField(
            "object_kind",
            "Вид объекта",
            FieldType.TEXT,
            aliases=("вид объекта", "тип объекта"),
        ),
        CanonicalField(
            "status_raw",
            "Статус проекта",
            FieldType.TEXT,
            aliases=("статус", "состояние"),
        ),
        CanonicalField(
            "contract_date",
            "Дата договора",
            FieldType.DATE,
            aliases=("дата договора", "заключён"),
        ),
        CanonicalField(
            "construction_start",
            "Начало строительства",
            FieldType.DATE,
            aliases=("начало строительства", "старт строительства"),
        ),
        CanonicalField(
            "construction_end",
            "Окончание строительства",
            FieldType.DATE,
            aliases=("окончание строительства", "завершение строительства"),
        ),
        CanonicalField(
            "cost_initial",
            "Первоначальная стоимость",
            FieldType.MONEY,
            aliases=("стоимость", "стоимость проекта", "сумма"),
        ),
        CanonicalField(
            "territory_name",
            "Территория",
            FieldType.TERRITORY,
            aliases=("район", "область", "регион", "территория"),
        ),
    ),
    targets=(
        TargetSpec(
            table_name="project_entities",
            model=ProjectEntity,
            natural_key_fields=("registry_number",),
            columns={"title": "title"},
            insert_defaults={
                "kind": str(ProjectEntityKind.PPP_PROJECT),
                "has_data_error": False,
                "risk_is_preliminary": False,
                "territory_precision": "none",
            },
            territory_column="territory_id",
            territory_raw_column="territory_raw",
        ),
        TargetSpec(
            table_name="ppp_projects",
            model=PppProject,
            natural_key_fields=("registry_number",),
            columns={
                "registry_number": "registry_number",
                "sector": "sector",
                "object_kind": "object_kind",
                "status_raw": "status_raw",
                "contract_date": "contract_date",
                "construction_start": "construction_start",
                "construction_end": "construction_end",
                "cost_initial": "cost_initial",
            },
            insert_defaults={"is_terminated": False},
            parent_index=0,
            shares_parent_id=True,
        ),
    ),
    logic_checks=(_check_negative_money, _check_construction_dates),
)


_SOCIOECONOMIC = KindSpec(
    kind=DataKind.SOCIOECONOMIC,
    title="Соц.-экон. показатели",
    description="Численность населения территорий слоя 8.1.",
    layer_code="8.1",
    territory_field="territory_name",
    note=(
        "Показатель привязывается к территории справочника. Строка с "
        "неопознанным названием не загружается: угадывать территорию "
        "запрещено."
    ),
    fields=(
        CanonicalField(
            "territory_name",
            "Территория",
            FieldType.TERRITORY,
            required=True,
            aliases=("район", "область", "регион", "территория", "наименование"),
        ),
        CanonicalField(
            "as_of_date",
            "Дата актуальности",
            FieldType.DATE,
            required=True,
            aliases=("дата", "на дату", "период", "дата актуальности"),
        ),
        CanonicalField(
            "total",
            "Численность, всего",
            FieldType.INTEGER,
            required=True,
            aliases=("всего", "население", "численность", "итого"),
        ),
        CanonicalField("male", "Мужчины", FieldType.INTEGER, aliases=("мужчины", "муж")),
        CanonicalField("female", "Женщины", FieldType.INTEGER, aliases=("женщины", "жен")),
        CanonicalField(
            "urban_total",
            "Городское население",
            FieldType.INTEGER,
            aliases=("городское", "город"),
        ),
        CanonicalField(
            "rural_total",
            "Сельское население",
            FieldType.INTEGER,
            aliases=("сельское", "село"),
        ),
    ),
    targets=(
        TargetSpec(
            table_name="population_stats",
            model=PopulationStat,
            natural_key_fields=("territory_name", "as_of_date"),
            columns={
                "as_of_date": "as_of_date",
                "total": "total",
                "male": "male",
                "female": "female",
                "urban_total": "urban_total",
                "rural_total": "rural_total",
            },
            territory_column="territory_id",
            territory_required=True,
        ),
    ),
    logic_checks=(_check_population_split,),
)


KIND_SPECS: Final[dict[DataKind, KindSpec]] = {
    spec.kind: spec
    for spec in (
        _PROCUREMENT,
        _BUDGET,
        _SUBSIDIES,
        _ORGANIZATIONS,
        _INFRASTRUCTURE,
        _SOCIOECONOMIC,
    )
}


def kind_spec(kind: DataKind | str) -> KindSpec:
    """Описание типа данных. Неизвестный код — отказ, а не молчаливый пропуск."""
    try:
        resolved = DataKind(str(kind))
    except ValueError as exc:
        raise ImportWizardError(
            "unknown_data_kind", f"Неизвестный тип данных: {kind}"
        ) from exc
    return KIND_SPECS[resolved]


def describe_kinds() -> list[dict[str, Any]]:
    """Шесть плиток шага 1 для интерфейса."""
    return [spec.as_dict() for spec in KIND_SPECS.values()]


# --- Разбор файла ------------------------------------------------------------


@dataclass(slots=True)
class TableData:
    """Прочитанное содержимое файла в виде таблицы."""

    columns: list[str]
    rows: list[dict[str, Any]]
    sheet_name: str
    row_refs: list[str]
    """Адрес каждой строки в источнике — попадает в замечание и в провенанс."""

    geometries: list[Any] = field(default_factory=list)
    """Геометрии GeoJSON, если файл геопространственный."""


def _normalize_header(value: object) -> str:
    """Свернуть написание заголовка к сравнимому виду.

    NFC обязателен: часть выгрузок приходит в NFD, и «й» из двух кодовых точек
    не совпадает с «й» из одной, хотя выглядит так же.
    """
    text = unicodedata.normalize("NFC", str(value or "")).strip().casefold()
    text = text.replace("ё", "е")
    return re.sub(r"[\s_\-–—.,:;()\[\]/\\]+", " ", text).strip()


def _read_excel(content: bytes, file_name: str) -> TableData:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        if file_name.casefold().endswith(".xls"):
            raise ImportWizardError(
                "legacy_xls",
                "Формат .xls (старый BIFF) не читается. Пересохраните книгу как .xlsx.",
            ) from exc
        raise ImportWizardError(
            "unreadable_workbook", f"Книга не открывается: {exc}"
        ) from exc

    try:
        sheet = workbook.worksheets[0]
        raw_rows = list(sheet.iter_rows(values_only=True))
        sheet_name = str(sheet.title)
    finally:
        workbook.close()

    if not raw_rows:
        raise ImportWizardError("empty_file", "В книге нет ни одной строки.")

    header_index = _detect_header_row(raw_rows)
    header = raw_rows[header_index]
    columns = _unique_columns(header)

    rows: list[dict[str, Any]] = []
    refs: list[str] = []
    for offset, values in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        rows.append(dict(zip(columns, list(values) + [None] * len(columns), strict=False)))
        refs.append(f"{sheet_name}!строка {offset}")

    return TableData(columns=columns, rows=rows, sheet_name=sheet_name, row_refs=refs)


def _detect_header_row(raw_rows: Sequence[Sequence[Any]]) -> int:
    """Найти строку заголовка.

    Первая строка заголовком бывает не всегда: в книге 8.4 он третий, а выше
    лежит шапка отчёта. Признак — наибольшее число непустых текстовых ячеек
    среди первых пяти строк.
    """
    best_index = 0
    best_score = -1
    for index, row in enumerate(raw_rows[:5]):
        score = sum(1 for value in row if value is not None and str(value).strip())
        if score > best_score:
            best_score, best_index = score, index
    return best_index


def _unique_columns(header: Sequence[Any]) -> list[str]:
    """Имена колонок без повторов.

    Повтор заголовка встречается в книгах регулярно. Молча слить две колонки в
    одну значило бы потерять данные, поэтому второй экземпляр получает суффикс.
    """
    seen: dict[str, int] = {}
    columns: list[str] = []
    for position, value in enumerate(header, start=1):
        text = "" if value is None else str(value).strip()
        name = text or f"Колонка {position}"
        count = seen.get(name, 0)
        seen[name] = count + 1
        columns.append(name if count == 0 else f"{name} ({count + 1})")
    return columns


def _decode_text(content: bytes) -> str:
    """Раскодировать текстовый файл.

    Порядок попыток не случаен: выгрузки из 1С и старого Excel приходят в
    cp1251, а всё остальное — в UTF-8. BOM у UTF-8 срезается `utf-8-sig`.
    """
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportWizardError(
        "unknown_encoding",
        "Кодировка файла не распознана. Сохраните файл в UTF-8.",
    )


def _read_csv(content: bytes) -> TableData:
    text = _decode_text(content)
    sample = text[:8192]
    try:
        dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        # Одна колонка без разделителей — законный случай, а не сбой.
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    raw_rows = [row for row in reader if row]
    if not raw_rows:
        raise ImportWizardError("empty_file", "В файле нет ни одной строки.")

    columns = _unique_columns(raw_rows[0])
    rows: list[dict[str, Any]] = []
    refs: list[str] = []
    for offset, values in enumerate(raw_rows[1:], start=2):
        padded = list(values) + [None] * (len(columns) - len(values))
        rows.append(dict(zip(columns, padded, strict=False)))
        refs.append(f"строка {offset}")
    return TableData(columns=columns, rows=rows, sheet_name="CSV", row_refs=refs)


def _read_json(content: bytes, *, geo: bool) -> TableData:
    text = _decode_text(content)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ImportWizardError(
            "invalid_json", f"Файл не является корректным JSON: {exc.msg} (позиция {exc.pos})."
        ) from exc

    records: list[dict[str, Any]] = []
    geometries: list[Any] = []

    if geo or (isinstance(payload, dict) and payload.get("type") == "FeatureCollection"):
        features = payload.get("features") if isinstance(payload, dict) else None
        if not isinstance(features, list):
            raise ImportWizardError(
                "invalid_geojson", "В GeoJSON нет массива features."
            )
        for feature in features:
            properties = feature.get("properties") if isinstance(feature, dict) else None
            records.append(dict(properties) if isinstance(properties, dict) else {})
            geometries.append(feature.get("geometry") if isinstance(feature, dict) else None)
    elif isinstance(payload, list):
        records = [dict(item) for item in payload if isinstance(item, dict)]
    elif isinstance(payload, dict):
        # Объект-обёртка с единственным списком внутри — частый вид выгрузки.
        lists = [value for value in payload.values() if isinstance(value, list)]
        if len(lists) == 1:
            records = [dict(item) for item in lists[0] if isinstance(item, dict)]
        else:
            records = [payload]
    if not records:
        raise ImportWizardError("empty_file", "В файле нет ни одной записи.")

    columns: list[str] = []
    for record in records:
        for key in record:
            if key not in columns:
                columns.append(str(key))

    rows = [{column: record.get(column) for column in columns} for record in records]
    refs = [f"запись {index}" for index in range(1, len(rows) + 1)]
    return TableData(
        columns=columns,
        rows=rows,
        sheet_name="GeoJSON" if geometries else "JSON",
        row_refs=refs,
        geometries=geometries,
    )


def read_table(content: bytes, file_name: str) -> TableData:
    """Прочитать файл любого поддерживаемого формата в таблицу."""
    suffix = Path(file_name).suffix.casefold()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return _read_excel(content, file_name)
    if suffix == ".csv":
        return _read_csv(content)
    if suffix == ".geojson":
        return _read_json(content, geo=True)
    if suffix == ".json":
        return _read_json(content, geo=False)
    raise ImportWizardError(
        "unsupported_format",
        f"Формат «{suffix or file_name}» не поддерживается. "
        f"Допустимы: {', '.join(sorted(ACCEPTED_EXTENSIONS))}.",
    )


# --- Приём файла -------------------------------------------------------------


def uploads_dir() -> Path:
    """Каталог загруженных файлов.

    Отдельный от `source_data_dir` намеренно: комплект книг ДЭР неизменяем, и
    запись в него — не «перезапись файла», а утрата исходника, по которому
    воспроизводятся все оценки.
    """
    directory = get_settings().data_dir / "uploads"
    directory.mkdir(parents=True, exist_ok=True)
    return directory


@dataclass(frozen=True, slots=True)
class UploadResult:
    """Результат шага 1: файл принят, структура прочитана."""

    source_file_id: uuid.UUID
    sha256: str
    file_name: str
    size_bytes: int
    stored_path: Path
    table: TableData
    suggestions: dict[str, str]

    def as_dict(self) -> dict[str, Any]:
        return {
            "upload_id": self.sha256,
            "source_file_id": str(self.source_file_id),
            "file_name": self.file_name,
            "size_bytes": self.size_bytes,
            "sheet_name": self.table.sheet_name,
            "row_count": len(self.table.rows),
            "columns": list(self.table.columns),
            "preview": [
                {column: jsonable(row.get(column)) for column in self.table.columns}
                for row in self.table.rows[:PREVIEW_ROWS]
            ],
            "suggested_mapping": dict(self.suggestions),
            "background_recommended": len(self.table.rows) >= BACKGROUND_ROW_THRESHOLD,
        }


def accept_upload(
    session: Session,
    *,
    file_name: str,
    content: bytes,
    kind: DataKind | str,
    user: User | None = None,
) -> UploadResult:
    """Шаг 1: принять файл, зафиксировать его и прочитать структуру.

    Файл регистрируется в `source_files` по SHA-256 — тому же признаку
    тождества, которым пользуются загрузчики книг. Повторная загрузка того же
    содержимого переиспользует запись, а не заводит вторую.
    """
    spec = kind_spec(kind)

    suffix = Path(file_name).suffix.casefold()
    if suffix not in ACCEPTED_EXTENSIONS:
        raise ImportWizardError(
            "unsupported_format",
            f"Формат «{suffix or '(без расширения)'}» не поддерживается. "
            f"Допустимы: {', '.join(sorted(ACCEPTED_EXTENSIONS))}.",
        )

    limit_bytes = get_settings().max_upload_mb * 1024 * 1024
    if len(content) > limit_bytes:
        raise ImportWizardError(
            "file_too_large",
            f"Файл больше {get_settings().max_upload_mb} МБ "
            f"({len(content) / 1024 / 1024:.1f} МБ).",
        )
    if not content:
        raise ImportWizardError("empty_file", "Файл пустой.")

    table = read_table(content, file_name)

    digest = hashlib.sha256(content).hexdigest()
    stored_path = uploads_dir() / f"{digest}{suffix}"
    if not stored_path.exists():
        stored_path.write_bytes(content)

    source_file = session.scalars(
        select(SourceFile).where(SourceFile.sha256 == digest)
    ).one_or_none()
    if source_file is None:
        source_file = SourceFile(
            id=stable_id("source_files", digest),
            file_name=file_name,
            normalized_name=unicodedata.normalize("NFC", file_name).casefold(),
            sha256=digest,
            size_bytes=len(content),
            origin="upload",
            uploaded_by_id=user.id if user is not None else None,
        )
        session.add(source_file)
        session.flush()

    return UploadResult(
        source_file_id=source_file.id,
        sha256=digest,
        file_name=file_name,
        size_bytes=len(content),
        stored_path=stored_path,
        table=table,
        suggestions=suggest_mapping(spec, table.columns),
    )


def source_file_for(session: Session, sha256: str) -> SourceFile:
    """Запись файла-источника по хешу принятой загрузки."""
    found = session.scalars(select(SourceFile).where(SourceFile.sha256 == sha256)).one_or_none()
    if found is None:
        raise ImportWizardError(
            "upload_not_found",
            "Загруженный файл не зарегистрирован — начните мастер заново с шага 1.",
        )
    return found


def stored_upload(sha256: str) -> Path:
    """Найти ранее принятый файл по его хешу."""
    for candidate in uploads_dir().glob(f"{sha256}.*"):
        return candidate
    raise ImportWizardError(
        "upload_not_found",
        "Загруженный файл не найден — начните мастер заново с шага 1.",
    )


# --- Сопоставление колонок ---------------------------------------------------


def suggest_mapping(spec: KindSpec, columns: Sequence[str]) -> dict[str, str]:
    """Предложить сопоставление «поле Системы → колонка файла».

    Направление именно такое, а не обратное: у поля Системы может быть лишь
    одна колонка-источник, а одна колонка теоретически способна попасть в два
    поля. Словарь с однозначным ключом избавляет от разбора этого случая.

    Совпадение ищется по нормализованному написанию: сначала точное, затем
    вхождение алиаса в заголовок. Догадки по частичному совпадению одного
    слова не делаются — ошибочное автосопоставление хуже отсутствующего,
    потому что его не замечают.
    """
    normalized = {column: _normalize_header(column) for column in columns}
    taken: set[str] = set()
    mapping: dict[str, str] = {}

    for candidate in spec.fields:
        variants = {_normalize_header(candidate.title), _normalize_header(candidate.code)}
        variants.update(_normalize_header(alias) for alias in candidate.aliases)
        variants.discard("")

        exact = next(
            (col for col, text in normalized.items() if text in variants and col not in taken),
            None,
        )
        if exact is not None:
            mapping[candidate.code] = exact
            taken.add(exact)
            continue

        partial = next(
            (
                col
                for col, text in normalized.items()
                if col not in taken and any(v and v in text for v in variants)
            ),
            None,
        )
        if partial is not None:
            mapping[candidate.code] = partial
            taken.add(partial)

    return mapping


def validate_mapping(spec: KindSpec, mapping: Mapping[str, str], columns: Sequence[str]) -> None:
    """Проверить сопоставление до чтения строк.

    Отсутствие обязательного поля — отказ на шаге 2, а не сто тысяч построчных
    ошибок на шаге 3: пользователю нужно сказать, что именно он не сопоставил,
    а не показать последствия.
    """
    known = {item.code for item in spec.fields}
    unknown = sorted(set(mapping) - known)
    if unknown:
        raise ImportWizardError(
            "unknown_field",
            f"Полей Системы с кодами {', '.join(unknown)} не существует.",
        )

    missing_columns = sorted({col for col in mapping.values() if col not in set(columns)})
    if missing_columns:
        raise ImportWizardError(
            "unknown_column",
            f"В файле нет колонок: {', '.join(missing_columns)}.",
        )

    missing_required = [
        item.title for item in spec.fields if item.required and item.code not in mapping
    ]
    if missing_required:
        raise ImportWizardError(
            "required_field_unmapped",
            f"Не сопоставлены обязательные поля: {', '.join(missing_required)}.",
        )

    if spec.territory_field is not None:
        needs_territory = any(target.territory_required for target in spec.targets)
        if needs_territory and spec.territory_field not in mapping:
            raise ImportWizardError(
                "required_field_unmapped",
                "Не сопоставлено поле территории, без которого запись некуда привязать.",
            )


# --- Шаблоны сопоставления ---------------------------------------------------

_TEMPLATE_ROLE: Final[str] = "mapping_template"
_TEMPLATE_PREFIX: Final[str] = "шаблон:"


def save_mapping_template(
    session: Session,
    *,
    name: str,
    kind: DataKind | str,
    mapping: Mapping[str, str],
    source_file_id: uuid.UUID,
) -> SourceDataset:
    """Сохранить сопоставление, чтобы не повторять его при следующей выгрузке.

    Шаблон хранится как `SourceDataset` с ролью `mapping_template`, а не в
    отдельной таблице: набор данных внутри файла — ровно то, чем шаблон и
    является, а `columns_meta` уже предназначен для описания колонок. Заводить
    ради этого новую таблицу значило бы дублировать существующую сущность.
    """
    spec = kind_spec(kind)
    cleaned = name.strip()
    if not cleaned:
        raise ImportWizardError("invalid_template_name", "У шаблона должно быть название.")

    sheet_name = f"{_TEMPLATE_PREFIX}{cleaned}"
    existing = session.scalars(
        select(SourceDataset).where(
            SourceDataset.source_file_id == source_file_id,
            SourceDataset.sheet_name == sheet_name,
        )
    ).one_or_none()

    dataset = existing or SourceDataset(
        id=stable_id("source_datasets", f"{source_file_id}|{sheet_name}"),
        source_file_id=source_file_id,
        sheet_name=sheet_name,
        role=_TEMPLATE_ROLE,
    )
    dataset.role = _TEMPLATE_ROLE
    dataset.layer_code = spec.layer_code
    dataset.columns_meta = {
        "data_kind": str(spec.kind),
        "name": cleaned,
        "mapping": dict(mapping),
    }
    if existing is None:
        session.add(dataset)
    session.flush()
    return dataset


def list_mapping_templates(
    session: Session, kind: DataKind | str | None = None
) -> list[dict[str, Any]]:
    """Сохранённые шаблоны сопоставления — для выпадающего списка шага 2."""
    stmt = select(SourceDataset).where(SourceDataset.role == _TEMPLATE_ROLE)
    if kind is not None:
        stmt = stmt.where(SourceDataset.layer_code == kind_spec(kind).layer_code)

    templates: list[dict[str, Any]] = []
    for dataset in session.scalars(stmt.order_by(SourceDataset.created_at.desc())):
        meta = dataset.columns_meta or {}
        templates.append(
            {
                "id": str(dataset.id),
                "name": str(meta.get("name") or dataset.sheet_name or ""),
                "data_kind": str(meta.get("data_kind") or ""),
                "mapping": dict(meta.get("mapping") or {}),
                "created_at": dataset.created_at.isoformat() if dataset.created_at else None,
            }
        )
    return templates


# --- Приведение значений -----------------------------------------------------


def _blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and not value.strip()


_DATE_FORMATS: Final[tuple[str, ...]] = (
    "%d.%m.%Y",
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%d-%m-%Y",
)

_EXCEL_EPOCH: Final[date] = date(1899, 12, 30)
"""Начало отсчёта дат в Excel.

30 декабря, а не 31: Excel считает 1900 год високосным, и смещение на день
компенсирует эту его ошибку.
"""


class _ConversionError(Exception):
    """Значение не приводится к типу поля. Несёт текст для пользователя."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


def _to_decimal(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        raise _ConversionError("invalid_number", "Логическое значение вместо числа.")
    if isinstance(value, int | float):
        if isinstance(value, float) and not math.isfinite(value):
            raise _ConversionError("invalid_number", "Не число (nan/inf).")
        return Decimal(str(value))
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    text = re.sub(r"[₸тг]", "", text, flags=re.IGNORECASE)
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise _ConversionError("invalid_number", f"«{value}» не является числом.") from exc


def _to_int(value: object) -> int:
    number = _to_decimal(value)
    if number != number.to_integral_value():
        raise _ConversionError(
            "invalid_type", f"«{value}» — дробное число там, где ожидается целое."
        )
    return int(number)


def _to_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, int | float) and not isinstance(value, bool):
        # Excel отдаёт даты числом, если ячейка не размечена как дата.
        return _EXCEL_EPOCH + timedelta(days=int(value))
    text = str(value).strip()
    for pattern in _DATE_FORMATS:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError as exc:
        raise _ConversionError(
            "invalid_date",
            f"«{value}» не распознано как дата. Ожидается ДД.ММ.ГГГГ или ГГГГ-ММ-ДД.",
        ) from exc


_TRUE_WORDS: Final[frozenset[str]] = frozenset({"да", "true", "1", "yes", "истина", "+"})
_FALSE_WORDS: Final[frozenset[str]] = frozenset({"нет", "false", "0", "no", "ложь", "-"})


def _to_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value).strip().casefold()
    if text in _TRUE_WORDS:
        return True
    if text in _FALSE_WORDS:
        return False
    raise _ConversionError("invalid_type", f"«{value}» не распознано как да/нет.")


def _to_xin(value: object) -> tuple[str, str | None]:
    """Привести ИИН/БИН к 12 цифрам и сообщить о восстановленных нулях.

    Ведущий ноль теряется всякий раз, когда книгу открывали в Excel: колонка
    распознаётся как число. Восстановить ноль можно, но молчать об этом
    нельзя — значение изменилось по сравнению с файлом.
    """
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    elif isinstance(value, int) and not isinstance(value, bool):
        text = str(value)
    else:
        text = str(value).strip()

    digits = re.sub(r"\D", "", text)
    if not digits:
        raise _ConversionError("invalid_type", f"«{value}» не содержит цифр ИИН/БИН.")
    if len(digits) > 12:
        raise _ConversionError(
            "invalid_type", f"«{value}»: в ИИН/БИН {len(digits)} цифр вместо 12."
        )
    if len(digits) < 12:
        restored = digits.rjust(12, "0")
        return restored, (
            f"Восстановлены ведущие нули: «{text}» → «{restored}». "
            "Скорее всего, колонка была прочитана Excel как число."
        )
    return digits, None


def _to_coordinate(value: object, *, is_latitude: bool) -> float:
    number = float(_to_decimal(value))
    limit = 90.0 if is_latitude else 180.0
    if not -limit <= number <= limit:
        raise _ConversionError(
            "invalid_coordinates",
            f"{'Широта' if is_latitude else 'Долгота'} {number} вне допустимого "
            f"диапазона [-{limit:g}; {limit:g}].",
        )
    return number


def convert_value(value: object, field_type: FieldType) -> tuple[Any, str | None]:
    """Привести значение к типу поля. Возвращает значение и, если было, замечание."""
    if _blank(value):
        return None, None
    if field_type is FieldType.TEXT or field_type is FieldType.TERRITORY:
        return str(value).strip(), None
    if field_type is FieldType.INTEGER:
        return _to_int(value), None
    if field_type is FieldType.NUMBER:
        return float(_to_decimal(value)), None
    if field_type is FieldType.MONEY:
        return _to_decimal(value), None
    if field_type is FieldType.DATE:
        return _to_date(value), None
    if field_type is FieldType.BOOL:
        return _to_bool(value), None
    if field_type is FieldType.XIN:
        return _to_xin(value)
    if field_type is FieldType.LATITUDE:
        return _to_coordinate(value, is_latitude=True), None
    return _to_coordinate(value, is_latitude=False), None


# --- Проверка строк ----------------------------------------------------------


@dataclass(slots=True)
class ValidatedRow:
    """Строка, прошедшая приведение типов."""

    index: int
    row_ref: str
    values: dict[str, Any]
    natural_keys: dict[str, str]
    territory_id: uuid.UUID | None = None
    territory_resolution: str = "not_determined"
    territory_raw: str | None = None
    valid: bool = True


@dataclass(slots=True)
class ValidationOutcome:
    """Итог проверки всего файла."""

    rows: list[ValidatedRow]
    issues: list[IssueRecord]
    rows_read: int = 0
    rows_failed: int = 0
    duplicates_in_file: int = 0
    duplicates_in_db: int = 0
    territory_report: dict[str, Any] = field(default_factory=dict)

    @property
    def valid_rows(self) -> list[ValidatedRow]:
        return [row for row in self.rows if row.valid]

    def counts_by_severity(self) -> dict[str, int]:
        counts = {str(level): 0 for level in IssueSeverity}
        for item in self.issues:
            counts[str(item.severity)] += 1
        return counts


def _check_geometry(geometry: object, row_ref: str) -> IssueRecord | None:
    """Проверить координаты объекта GeoJSON.

    Проверяются две разные вещи: формальная допустимость (широта в [-90; 90])
    и правдоподобие (точка внутри рамки Казахстана). Первое — ошибка, второе —
    предупреждение: точка за рамкой чаще всего означает перепутанные местами
    широту и долготу, но бывает и законной.
    """
    if geometry is None:
        return None
    if not isinstance(geometry, Mapping) or "coordinates" not in geometry:
        return IssueRecord(
            severity=IssueSeverity.ERROR,
            code="invalid_coordinates",
            message="Геометрия объекта не содержит координат.",
            source_row_ref=row_ref,
            column_name="geometry",
        )

    points: list[tuple[float, float]] = []

    def collect(node: Any) -> None:
        if (
            isinstance(node, list | tuple)
            and len(node) >= 2
            and all(isinstance(part, int | float) for part in node[:2])
        ):
            points.append((float(node[0]), float(node[1])))
            return
        if isinstance(node, list | tuple):
            for item in node:
                collect(item)

    collect(geometry["coordinates"])
    if not points:
        return IssueRecord(
            severity=IssueSeverity.ERROR,
            code="invalid_coordinates",
            message="В геометрии объекта нет ни одной точки.",
            source_row_ref=row_ref,
            column_name="geometry",
        )

    for longitude, latitude in points:
        if not -180.0 <= longitude <= 180.0 or not -90.0 <= latitude <= 90.0:
            return IssueRecord(
                severity=IssueSeverity.ERROR,
                code="invalid_coordinates",
                message=f"Координата ({longitude}; {latitude}) вне допустимого диапазона.",
                source_row_ref=row_ref,
                column_name="geometry",
            )

    west, south, east, north = _KZ_BBOX
    outside = [
        point
        for point in points
        if not (west <= point[0] <= east and south <= point[1] <= north)
    ]
    if outside:
        return IssueRecord(
            severity=IssueSeverity.WARNING,
            code="coordinates_outside_country",
            message=(
                f"Точек за пределами рамки Казахстана: {len(outside)}. "
                "Обычно это перепутанные местами широта и долгота."
            ),
            source_row_ref=row_ref,
            column_name="geometry",
            context={"sample": list(outside[0])},
        )
    return None


def validate_rows(
    session: Session,
    *,
    spec: KindSpec,
    mapping: Mapping[str, str],
    table: TableData,
    territory_index: TerritoryIndex | None = None,
) -> ValidationOutcome:
    """Построчная проверка: обязательные поля, типы, координаты, дубликаты, логика.

    Замечание всегда несёт адрес строки и имя колонки — этого требует ТЗ и без
    этого замечание бесполезно: «неверный формат даты» без указания места
    заставляет искать ошибку глазами по всему файлу.
    """
    index = territory_index if territory_index is not None else load_territory_index(session)
    outcome = ValidationOutcome(rows=[], issues=[], rows_read=len(table.rows))

    seen_keys: dict[str, int] = {}

    for position, raw_row in enumerate(table.rows):
        row_ref = (
            table.row_refs[position]
            if position < len(table.row_refs)
            else f"строка {position + 1}"
        )
        values: dict[str, Any] = {}
        row_valid = True

        for candidate in spec.fields:
            column = mapping.get(candidate.code)
            if column is None:
                continue
            raw_value = raw_row.get(column)
            try:
                converted, note = convert_value(raw_value, candidate.type)
            except _ConversionError as exc:
                row_valid = False
                outcome.issues.append(
                    IssueRecord(
                        severity=IssueSeverity.ERROR,
                        code=exc.code,
                        message=f"Поле «{candidate.title}»: {exc.message}",
                        source_row_ref=row_ref,
                        column_name=column,
                        raw_value=None if raw_value is None else str(raw_value),
                    )
                )
                continue

            if note is not None:
                outcome.issues.append(
                    IssueRecord(
                        severity=IssueSeverity.WARNING,
                        code="leading_zeros_lost",
                        message=f"Поле «{candidate.title}»: {note}",
                        source_row_ref=row_ref,
                        column_name=column,
                        raw_value=None if raw_value is None else str(raw_value),
                    )
                )

            if converted is None and candidate.required:
                row_valid = False
                outcome.issues.append(
                    IssueRecord(
                        severity=IssueSeverity.ERROR,
                        code="required_field_missing",
                        message=f"Обязательное поле «{candidate.title}» не заполнено.",
                        source_row_ref=row_ref,
                        column_name=column,
                    )
                )
            values[candidate.code] = converted

        # Логические противоречия проверяются на уже приведённых значениях:
        # сравнивать даты, пока они строки, значит сравнивать написание.
        for check in spec.logic_checks:
            found = check(values)
            if found is not None:
                severity, code, message = found
                outcome.issues.append(
                    IssueRecord(
                        severity=severity,
                        code=code,
                        message=message,
                        source_row_ref=row_ref,
                    )
                )
                if severity is IssueSeverity.ERROR:
                    row_valid = False

        if position < len(table.geometries):
            geometry_issue = _check_geometry(table.geometries[position], row_ref)
            if geometry_issue is not None:
                outcome.issues.append(geometry_issue)
                if geometry_issue.severity is IssueSeverity.ERROR:
                    row_valid = False

        validated = ValidatedRow(
            index=position, row_ref=row_ref, values=values, natural_keys={}, valid=row_valid
        )

        if spec.territory_field is not None and spec.territory_field in values:
            raw_name = values.get(spec.territory_field)
            validated.territory_raw = None if raw_name is None else str(raw_name)
            territory_id, resolution = index.lookup(raw_name, row_ref=row_ref)
            validated.territory_id = territory_id
            validated.territory_resolution = (
                "resolved" if territory_id is not None else str(resolution.status)
            )
            needs_territory = any(target.territory_required for target in spec.targets)
            if territory_id is None and needs_territory:
                row_valid = False
                validated.valid = False
                outcome.issues.append(
                    IssueRecord(
                        severity=IssueSeverity.ERROR,
                        code="territory_not_resolved",
                        message=(
                            f"Территория «{raw_name}» не найдена в справочнике, "
                            "а без неё запись некуда привязать. Угадывание запрещено."
                        ),
                        source_row_ref=row_ref,
                        column_name=mapping.get(spec.territory_field),
                        raw_value=None if raw_name is None else str(raw_name),
                    )
                )

        # Естественные ключи считаются для каждой цели: у договора он свой,
        # у поставщика — свой, и повтор одного не означает повтора другого.
        for target in spec.targets:
            key = _natural_key(target, values)
            if key is not None:
                validated.natural_keys[target.table_name] = key

        primary = spec.targets[-1]
        primary_key = validated.natural_keys.get(primary.table_name)
        if row_valid and primary_key is not None:
            previous = seen_keys.get(primary_key)
            if previous is not None:
                outcome.duplicates_in_file += 1
                outcome.issues.append(
                    IssueRecord(
                        severity=IssueSeverity.WARNING,
                        code="duplicate_in_file",
                        message=(
                            f"Ключ «{primary_key}» уже встречался в строке {previous + 1}. "
                            "В базу попадёт последнее вхождение."
                        ),
                        source_row_ref=row_ref,
                        raw_value=primary_key,
                    )
                )
            seen_keys[primary_key] = position

        if not row_valid:
            validated.valid = False
            outcome.rows_failed += 1

        outcome.rows.append(validated)

    outcome.duplicates_in_db = _count_existing(session, spec, outcome.valid_rows)
    if outcome.duplicates_in_db:
        outcome.issues.append(
            IssueRecord(
                severity=IssueSeverity.INFO,
                code="duplicate_in_db",
                message=(
                    f"Записей с такими ключами уже есть в базе: {outcome.duplicates_in_db}. "
                    "Они будут обновлены, а не продублированы."
                ),
            )
        )

    outcome.issues.extend(index.issues())
    outcome.territory_report = index.report()
    return outcome


def _natural_key(target: TargetSpec, values: Mapping[str, Any]) -> str | None:
    """Естественный ключ строки для одной цели.

    `None`, если хотя бы одна составляющая пуста: ключ из пустых частей
    склеил бы в одну запись все строки с пропусками.
    """
    parts: list[str] = []
    for code in target.natural_key_fields:
        value = values.get(code)
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        parts.append(str(value).strip())
    return "|".join(parts)


def _count_existing(
    session: Session, spec: KindSpec, rows: Sequence[ValidatedRow]
) -> int:
    """Сколько строк файла уже есть в базе.

    Считается по первичному ключу, который тот же самый детерминированный
    `stable_id`: это ровно те строки, которые импорт обновит, а не создаст.
    """
    target = spec.targets[-1]
    ids = {
        stable_id(target.table_name, key)
        for key in (row.natural_keys.get(target.table_name) for row in rows)
        if key is not None
    }
    if not ids:
        return 0
    table = table_of(target.model)
    found = session.execute(
        select(func.count()).select_from(table).where(table.c.id.in_(ids))
    ).scalar_one()
    return int(found)


# --- Версионирование ---------------------------------------------------------


def next_data_version(session: Session, *, spec: KindSpec, sha256: str) -> int:
    """Логическая версия данных для этой загрузки.

    Тот же файл — та же версия: иначе двойное подтверждение одной и той же
    выгрузки порождало бы две версии, отличающиеся только номером, и
    «повторный запуск не создаёт дублей» перестало бы выполняться на уровне
    версий. Новое содержимое поднимает версию слоя на единицу.
    """
    previous = session.execute(
        select(ImportJob.data_version)
        .join(SourceFile, SourceFile.id == ImportJob.source_file_id)
        .where(
            SourceFile.sha256 == sha256,
            ImportJob.layer_code == spec.layer_code,
            ImportJob.is_dry_run.is_(False),
        )
        .order_by(ImportJob.data_version.desc())
        .limit(1)
    ).scalar_one_or_none()
    if previous is not None:
        return int(previous)

    highest = session.execute(
        select(func.max(ImportJob.data_version)).where(
            ImportJob.layer_code == spec.layer_code,
            ImportJob.is_dry_run.is_(False),
        )
    ).scalar_one_or_none()
    return int(highest or 0) + 1


# --- Сухой прогон ------------------------------------------------------------


def _persist_issues(session: Session, job_id: uuid.UUID, issues: Sequence[IssueRecord]) -> int:
    """Записать замечания к заданию, свернув хвост при переполнении."""
    kept = list(issues[:MAX_PERSISTED_ISSUES])
    dropped = len(issues) - len(kept)
    if dropped > 0:
        kept.append(
            IssueRecord(
                severity=IssueSeverity.INFO,
                code="issues_truncated",
                message=(
                    f"Показаны первые {MAX_PERSISTED_ISSUES} замечаний, "
                    f"ещё {dropped} не сохранены."
                ),
            )
        )
    if not kept:
        return 0

    rows = [
        {
            "id": stable_id("data_quality_issues", f"{job_id}|{position}"),
            "import_job_id": job_id,
            "severity": item.severity,
            "code": item.code,
            "message": item.message,
            "source_row_ref": item.source_row_ref,
            "column_name": item.column_name,
            "raw_value": item.raw_value,
            "context": jsonable(item.context) if item.context is not None else None,
        }
        for position, item in enumerate(kept)
    ]
    bulk_upsert(session, table_of(DataQualityIssue), rows)
    return len(kept)


def _wizard_payload(
    *,
    spec: KindSpec,
    mapping: Mapping[str, str],
    upload_sha: str,
    file_name: str,
    outcome: ValidationOutcome,
    dry_run_job_id: uuid.UUID | None = None,
) -> dict[str, Any]:
    """Состояние мастера, сохраняемое в задании.

    Кладётся в `ImportJob.reconciliation` — свободное JSONB-поле задания.
    Отдельная колонка не заводилась намеренно: шаг подтверждения должен уметь
    восстановить контекст сухого прогона после перезапуска приложения, а
    заводить ради этого таблицу состояния мастера — лишняя сущность.
    """
    return {
        "wizard": {
            "data_kind": str(spec.kind),
            "mapping": dict(mapping),
            "upload_id": upload_sha,
            "file_name": file_name,
            "dry_run_job_id": str(dry_run_job_id) if dry_run_job_id else None,
        },
        "summary": {
            "rows_read": outcome.rows_read,
            "rows_valid": len(outcome.valid_rows),
            "rows_failed": outcome.rows_failed,
            "duplicates_in_file": outcome.duplicates_in_file,
            "duplicates_in_db": outcome.duplicates_in_db,
            "issues": outcome.counts_by_severity(),
        },
    }


def dry_run(
    session: Session,
    *,
    upload_id: str,
    kind: DataKind | str,
    mapping: Mapping[str, str],
    user: User | None = None,
    context: RequestContext | None = None,
) -> ImportJob:
    """Шаг 3: показать, что произойдёт, ничего не записав.

    Задание создаётся и здесь — со статусом `dry_run`. Так история загрузок
    честно показывает и проверочные прогоны, а построчные замечания остаются
    доступны по ссылке после закрытия мастера. В доменные таблицы при этом не
    уходит ни одной строки.
    """
    spec = kind_spec(kind)
    path = stored_upload(upload_id)
    table = read_table(path.read_bytes(), path.name)
    validate_mapping(spec, mapping, table.columns)

    outcome = validate_rows(session, spec=spec, mapping=mapping, table=table)

    source_file = session.scalars(
        select(SourceFile).where(SourceFile.sha256 == upload_id)
    ).one_or_none()

    started = utcnow()
    job = ImportJob(
        id=uuid.uuid4(),
        source_file_id=source_file.id if source_file else None,
        layer_code=spec.layer_code,
        importer=f"wizard:{spec.kind}",
        status=ImportStatus.DRY_RUN,
        is_dry_run=True,
        data_version=next_data_version(session, spec=spec, sha256=upload_id),
        started_at=started,
        finished_at=utcnow(),
        rows_read=outcome.rows_read,
        rows_created=0,
        rows_updated=0,
        rows_skipped=outcome.rows_failed,
        rows_failed=outcome.rows_failed,
        started_by_id=user.id if user is not None else None,
        territory_match_report=jsonable(outcome.territory_report) or None,
        reconciliation=_wizard_payload(
            spec=spec,
            mapping=mapping,
            upload_sha=upload_id,
            file_name=source_file.file_name if source_file else path.name,
            outcome=outcome,
        ),
    )
    session.add(job)
    session.flush()
    _persist_issues(session, job.id, outcome.issues)
    session.flush()

    audit.record(
        AuditAction.IMPORT_STARTED,
        session=session,
        user=user,
        context=context,
        entity_type="import_job",
        entity_id=job.id,
        details={"mode": "dry_run", "data_kind": str(spec.kind), "rows": outcome.rows_read},
    )
    return job


# --- Подтверждение и запись --------------------------------------------------


def _row_payload(
    target: TargetSpec,
    row: ValidatedRow,
    *,
    columns: Mapping[str, str],
    row_id: uuid.UUID,
    parent_id: uuid.UUID | None,
    provenance: Mapping[str, Any],
) -> dict[str, Any] | None:
    """Собрать словарь колонок для одной строки одной цели.

    Набор ключей одинаков у всех строк — этого требует массовая вставка, и это
    же обеспечивает главную гарантию: `columns` содержит только сопоставленные
    поля, поэтому колонки, которых не было в сопоставлении, в словарь не
    попадают и при обновлении остаются нетронутыми. Импорт двух колонок не
    имеет права обнулить всё остальное, что загрузчик книги посчитал раньше.
    """
    payload: dict[str, Any] = {"id": row_id}

    for field_code, column in columns.items():
        value = row.values.get(field_code)
        if value is None and column in target.null_fallback:
            value = target.null_fallback[column]
        payload[column] = value

    payload.update(dict(target.insert_defaults))

    if target.territory_column is not None:
        payload[target.territory_column] = row.territory_id
    if target.territory_raw_column is not None:
        payload[target.territory_raw_column] = row.territory_raw
    if target.territory_resolution_column is not None:
        payload[target.territory_resolution_column] = row.territory_resolution
    if target.parent_link_column is not None:
        if parent_id is None:
            return None
        payload[target.parent_link_column] = parent_id

    payload.update(provenance)
    return payload


@dataclass(slots=True)
class ConfirmResult:
    """Что записало подтверждение."""

    job: ImportJob
    created: int = 0
    updated: int = 0
    skipped: int = 0


ProgressCallback = Callable[[ImportJob, int, int], None]
"""Отчёт о прогрессе: задание, обработано строк, всего строк.

Задание передаётся первым параметром, потому что до вызова `confirm` его не
существует: фоновому обработчику неоткуда взять идентификатор заранее, а без
идентификатора прогресс некуда записать.
"""


def confirm(
    session: Session,
    *,
    upload_id: str,
    kind: DataKind | str,
    mapping: Mapping[str, str],
    user: User | None = None,
    context: RequestContext | None = None,
    progress: ProgressCallback | None = None,
) -> ConfirmResult:
    """Шаг 3, подтверждение: записать данные новой логической версией.

    Строки, не прошедшие проверку, не пишутся — но и не роняют импорт целиком:
    файл из десяти тысяч строк с тремя битыми датами должен загрузиться на
    9997 строк с тремя видимыми замечаниями, а не отвергнуться полностью.
    """
    spec = kind_spec(kind)
    path = stored_upload(upload_id)
    content = path.read_bytes()
    table = read_table(content, path.name)
    validate_mapping(spec, mapping, table.columns)

    territory_index = load_territory_index(session)
    outcome = validate_rows(
        session, spec=spec, mapping=mapping, table=table, territory_index=territory_index
    )

    version = next_data_version(session, spec=spec, sha256=upload_id)

    job_runner = LayerJob(
        session,
        layer_code=spec.layer_code,
        importer=f"wizard:{spec.kind}",
        dry_run=False,
    )
    job = job_runner.start()
    job.data_version = version
    job.started_by_id = user.id if user is not None else None

    source_file = job_runner.source_file(path, origin="upload")
    dataset = job_runner.dataset(
        source_file,
        sheet_name=table.sheet_name,
        role="raw",
        row_count=len(table.rows),
    )

    audit.record(
        AuditAction.IMPORT_STARTED,
        session=session,
        user=user,
        context=context,
        entity_type="import_job",
        entity_id=job.id,
        details={"mode": "confirm", "data_kind": str(spec.kind), "version": version},
    )

    valid = outcome.valid_rows
    job_runner.count_read(outcome.rows_read)
    job_runner.count_skipped(outcome.rows_failed)
    job_runner.extend_issues(outcome.issues)

    result = ConfirmResult(job=job, skipped=outcome.rows_failed)
    total = len(valid)
    parent_ids: dict[int, uuid.UUID] = {}

    for target_index, target in enumerate(spec.targets):
        rows_payload: list[dict[str, Any]] = []
        current_ids: dict[int, uuid.UUID] = {}
        # Только сопоставленные поля. Несопоставленные колонки не попадают ни
        # в INSERT, ни в UPDATE — см. `_row_payload`.
        columns = {
            field_code: column
            for field_code, column in target.columns.items()
            if field_code in mapping
        }

        for processed, row in enumerate(valid, start=1):
            key = row.natural_keys.get(target.table_name)
            if key is None:
                continue

            if target.shares_parent_id:
                parent_row_id = parent_ids.get(row.index)
                if parent_row_id is None:
                    continue
                row_id = parent_row_id
            else:
                row_id = stable_id(target.table_name, key)

            provenance = job_runner.provenance(
                dataset,
                natural_key=key,
                source_row_ref=row.row_ref,
                data_as_of=None,
                data_version=version,
            )
            payload = _row_payload(
                target,
                row,
                columns=columns,
                row_id=row_id,
                parent_id=parent_ids.get(row.index) if target.parent_index is not None else None,
                provenance=provenance,
            )
            if payload is None:
                continue
            rows_payload.append(payload)
            current_ids[row.index] = row_id

            if progress is not None and processed % PROGRESS_CHUNK == 0:
                progress(job, processed, total)

        if rows_payload:
            # Значения по умолчанию неизменяемы: у существующей строки они уже
            # посчитаны загрузчиком книги, и заглушка не должна их затирать.
            immutable = ("created_at", *target.insert_defaults.keys())
            counts = bulk_upsert(
                session,
                table_of(target.model),
                rows_payload,
                immutable_columns=immutable,
            )
            job_runner.report.tables[target.table_name] = counts
            if target_index == len(spec.targets) - 1:
                result.created = counts.created
                result.updated = counts.updated

        if target_index == 0:
            parent_ids = current_ids

    if progress is not None:
        progress(job, total, total)

    final_payload = _wizard_payload(
        spec=spec,
        mapping=mapping,
        upload_sha=upload_id,
        file_name=source_file.file_name,
        outcome=outcome,
    )
    # Прогресс дописывается в итоговую сводку, иначе завершённое задание
    # осталось бы с последним промежуточным процентом, а не со ста.
    final_payload["progress"] = {"processed": total, "total": total, "percent": 100}
    job_runner.finish(final_payload)

    audit.record(
        AuditAction.IMPORT_FINISHED,
        session=session,
        user=user,
        context=context,
        entity_type="import_job",
        entity_id=job.id,
        details={
            "data_kind": str(spec.kind),
            "version": version,
            "created": result.created,
            "updated": result.updated,
            "skipped": result.skipped,
        },
    )
    session.flush()
    return result


def write_progress(session: Session, job: ImportJob, processed: int, total: int) -> None:
    """Записать прогресс в задание и зафиксировать его.

    Фиксация обязательна: смысл прогресса в том, чтобы его видел *другой*
    запрос, опрашивающий состояние задания. Незакоммиченное значение видно
    только той транзакции, которая его записала, то есть никому.

    Словарь присваивается целиком, а не правится по месту: SQLAlchemy не
    отслеживает изменения внутри JSONB, и правка по месту тихо не сохранилась
    бы.
    """
    payload = dict(job.reconciliation or {})
    payload["progress"] = {
        "processed": processed,
        "total": total,
        "percent": round(processed * 100 / total) if total else 100,
    }
    job.reconciliation = payload
    session.commit()


def confirm_in_background(
    *,
    upload_id: str,
    kind: DataKind | str,
    mapping: Mapping[str, str],
    user_id: uuid.UUID | None,
    session_factory: Callable[[], Session] | None = None,
) -> uuid.UUID:
    """Фоновая обработка большого файла с записью прогресса.

    Собственная сессия обязательна: транзакция HTTP-запроса к этому моменту уже
    закрыта. Фабрика сессий передаётся параметром ради тестов — иначе проверить
    фоновый путь можно было бы только запустив настоящий сервер.
    """
    if session_factory is None:
        from app.db.session import get_session_factory

        session_factory = get_session_factory()

    with session_factory() as session:
        user = session.get(User, user_id) if user_id is not None else None
        try:
            result = confirm(
                session,
                upload_id=upload_id,
                kind=kind,
                mapping=mapping,
                user=user,
                progress=lambda job, processed, total: write_progress(
                    session, job, processed, total
                ),
            )
            session.commit()
            return result.job.id
        except Exception:
            session.rollback()
            raise


# --- Откат -------------------------------------------------------------------


def rollback(
    session: Session,
    *,
    job_id: uuid.UUID,
    user: User | None = None,
    context: RequestContext | None = None,
    reason: str = "",
) -> ImportJob:
    """Откатить логическую версию, не удаляя данных.

    Откат снимает `is_current` со строк, помеченных этим заданием импорта, и
    переводит задание в статус `rolled_back`. Ни одна строка, ни одно
    замечание качества и ни одна запись журнала не удаляются: показанная вчера
    оценка обязана остаться объяснимой, а журнал — доказательством.

    Чего откат намеренно **не** делает: он не восстанавливает значения полей,
    какими они были до импорта. Половина целевых таблиц имеет ограничение
    уникальности по естественному ключу без версии, поэтому новая версия
    обновляет строку на месте, а не заводит рядом вторую. Полноценный возврат
    значений требовал бы построчной истории (SCD-2), которой в схеме нет.
    Правило поэтому сформулировано честно: откат отзывает *актуальность*
    версии, а не переписывает прошлое.
    """
    job = session.get(ImportJob, job_id)
    if job is None:
        raise ImportWizardError("job_not_found", "Задание импорта не найдено.")
    if job.is_dry_run:
        raise ImportWizardError(
            "dry_run_rollback",
            "Сухой прогон ничего не записал — откатывать нечего.",
        )
    if job.status == ImportStatus.ROLLED_BACK:
        raise ImportWizardError("already_rolled_back", "Эта версия уже откачена.")

    spec = _spec_for_job(job)
    affected = 0

    for target in spec.targets if spec else ():
        table = table_of(target.model)
        result = session.execute(
            update(table)
            .where(table.c.import_job_id == job_id, table.c.is_current.is_(True))
            .values(is_current=False)
        )
        # `rowcount` объявлен не на базовом `Result`, а на курсорном подтипе,
        # который возвращает UPDATE. Читаем через getattr, чтобы не подавлять
        # проверку типов приведением, обещающим больше, чем известно.
        affected += int(getattr(result, "rowcount", 0) or 0)

    job.status = ImportStatus.ROLLED_BACK
    job.error_message = reason or "Версия отозвана оператором"
    session.flush()

    audit.record(
        AuditAction.IMPORT_ROLLED_BACK,
        session=session,
        user=user,
        context=context,
        entity_type="import_job",
        entity_id=job.id,
        details={
            "layer_code": job.layer_code,
            "version": job.data_version,
            "rows_deactivated": affected,
            "reason": reason,
        },
    )
    return job


def _spec_for_job(job: ImportJob) -> KindSpec | None:
    """Восстановить тип данных задания.

    Сначала по сохранённому состоянию мастера, потом по имени импортёра:
    задание могло быть создано и загрузчиком книги, у которого состояния
    мастера нет.
    """
    payload = job.reconciliation or {}
    wizard = payload.get("wizard") if isinstance(payload, dict) else None
    code = wizard.get("data_kind") if isinstance(wizard, dict) else None
    if code is None and job.importer.startswith("wizard:"):
        code = job.importer.split(":", 1)[1]
    if code is None:
        return next(
            (spec for spec in KIND_SPECS.values() if spec.layer_code == job.layer_code), None
        )
    try:
        return KIND_SPECS[DataKind(str(code))]
    except (ValueError, KeyError):
        return None


# --- История загрузок --------------------------------------------------------


def _badge(job: ImportJob, severities: Mapping[str, int]) -> str:
    """Статусный бейдж карточки истории: ОК или Предупреждение.

    Ровно два состояния успеха с референса плюс «ошибка» для упавших заданий.
    Замечание уровня warning не отменяет успех загрузки, но и не должно
    выглядеть как безоблачный результат.
    """
    if job.status in {ImportStatus.FAILED}:
        return "error"
    if job.status == ImportStatus.ROLLED_BACK:
        return "rolled_back"
    if severities.get(str(IssueSeverity.ERROR), 0) or job.rows_failed:
        return "warning"
    if severities.get(str(IssueSeverity.WARNING), 0):
        return "warning"
    return "ok"


def job_payload(session: Session, job: ImportJob, *, with_issues: bool = False) -> dict[str, Any]:
    """Задание импорта в виде, пригодном для истории загрузок и карточки."""
    counts_rows = session.execute(
        select(DataQualityIssue.severity, func.count())
        .where(DataQualityIssue.import_job_id == job.id)
        .group_by(DataQualityIssue.severity)
    ).all()
    severities = {str(level): int(count) for level, count in counts_rows}

    source_name: str | None = None
    if job.source_file_id is not None:
        source_name = session.execute(
            select(SourceFile.file_name).where(SourceFile.id == job.source_file_id)
        ).scalar_one_or_none()

    payload: dict[str, Any] = dict(job.reconciliation or {})
    raw_wizard = payload.get("wizard")
    wizard: dict[str, Any] = dict(raw_wizard) if isinstance(raw_wizard, dict) else {}

    result: dict[str, Any] = {
        "id": str(job.id),
        "layer_code": job.layer_code,
        "data_kind": wizard.get("data_kind"),
        "importer": job.importer,
        "status": str(job.status),
        "is_dry_run": job.is_dry_run,
        "data_version": job.data_version,
        "file_name": source_name or wizard.get("file_name"),
        "started_at": job.started_at.isoformat() if job.started_at else None,
        "finished_at": job.finished_at.isoformat() if job.finished_at else None,
        "rows_read": job.rows_read,
        "rows_created": job.rows_created,
        "rows_updated": job.rows_updated,
        "rows_skipped": job.rows_skipped,
        "rows_failed": job.rows_failed,
        "issues": severities,
        "badge": _badge(job, severities),
        "summary": payload.get("summary"),
        "progress": payload.get("progress"),
        "territory": job.territory_match_report,
        "error_message": job.error_message,
        "can_rollback": job.status == ImportStatus.SUCCEEDED and not job.is_dry_run,
    }

    if with_issues:
        rows = session.scalars(
            select(DataQualityIssue)
            .where(DataQualityIssue.import_job_id == job.id)
            .order_by(DataQualityIssue.severity, DataQualityIssue.id)
            .limit(500)
        ).all()
        result["issue_list"] = [
            {
                "severity": str(item.severity),
                "code": item.code,
                "message": item.message,
                "row": item.source_row_ref,
                "column": item.column_name,
                "raw_value": item.raw_value,
                "context": item.context,
            }
            for item in rows
        ]
    return result


def job_history(
    session: Session, *, limit: int = 20, layer_code: str | None = None
) -> list[dict[str, Any]]:
    """Правая колонка мастера: последние загрузки со статусами."""
    stmt = select(ImportJob).order_by(ImportJob.started_at.desc().nulls_last()).limit(limit)
    if layer_code is not None:
        stmt = stmt.where(ImportJob.layer_code == layer_code)
    return [job_payload(session, job) for job in session.scalars(stmt)]


__all__ = [
    "ACCEPTED_EXTENSIONS",
    "BACKGROUND_ROW_THRESHOLD",
    "KIND_SPECS",
    "PREVIEW_ROWS",
    "CanonicalField",
    "ConfirmResult",
    "DataKind",
    "FieldType",
    "ImportWizardError",
    "KindSpec",
    "TableData",
    "TargetSpec",
    "UploadResult",
    "ValidatedRow",
    "ValidationOutcome",
    "accept_upload",
    "confirm",
    "confirm_in_background",
    "convert_value",
    "describe_kinds",
    "dry_run",
    "job_history",
    "job_payload",
    "kind_spec",
    "list_mapping_templates",
    "next_data_version",
    "read_table",
    "rollback",
    "save_mapping_template",
    "source_file_for",
    "stored_upload",
    "suggest_mapping",
    "uploads_dir",
    "validate_mapping",
    "validate_rows",
    "write_progress",
]
