"""Импортёр слоя 8.4 «Госзакупки».

Книга `Слой_8_4_госзакупки_риски_v1.xlsx`, 15 листов. Расчётный лист —
статический экспорт: **формул в нём ноль**, вся арифметика выполнена вне
Excel. Проверить методику внутри книги невозможно, поэтому импортёр читает и
сырьё, и результат, а расчёт выполняется заново в `app.risk.layers.procurement`.

Ловушки этой книги стоит перечислить сразу — почти каждая строчка разбора
ниже написана против одной из них.

* **Заголовки в строке 3**, данные с четвёртой. Наивное чтение даёт мусорные
  имена колонок и две лишние строки.
* **`contract_id` разного типа**: `str` в расчётном листе, `int` в сырых.
  Join без приведения к строке даёт ноль совпадений.
* **Числа записаны строками с пробелами**: `'11 953 000.00'`, причём пробел
  местами неразрывный. `float()` на таком падает.
* **Даты в `contract_additions` — Excel-серийные числа** (`45415.549363…`),
  тогда как в `contract_details` даты настоящие. В одной колонке смешаны int
  и float.
* **БИН потерял ведущие нули** у 763 организаций из 3 668: `440010133` вместо
  `000440010133`. Лечится `zfill(12)` на обеих сторонах join.
* **Заглушки `'—'` и `'nan'` — это строки, а не пустые ячейки.** Фильтр по
  «пусто» их не поймает, и `'nan'` превратится в осмысленный способ закупки.
* **Имя заказчика в расчётном листе обрезано** до 60 знаков, из-за чего разные
  заказчики схлопываются в одного. Для группировок B3 и B4 берётся полное имя
  из листа `lots`.

Геопривязка идёт по юридическому адресу поставщика: КАТО места поставки
заполнен у 191 лота из 381, а привязка по заказчику покрыла бы 129 договоров
из 355. Адрес поставщика даёт 355 из 355. Ограничение при этом остаётся и
названо в книге прямо: юридический адрес — место регистрации, а не место
исполнения договора.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook

from app.risk.layers.procurement import (
    ContractRiskInputs,
    ContractRiskResult,
    SupplierRiskProfile,
    derive_b1,
    derive_b2,
    derive_b3,
    derive_b4,
    derive_b5,
    derive_b6,
    derive_b7,
    derive_b9,
    evaluate_contract,
)

LAYER_CODE: Final[str] = "8.4"
SOURCE_FILE_NAME: Final[str] = "Слой_8_4_госзакупки_риски_v1.xlsx"

SHEET_CALC: Final[str] = "Расчёт по договорам"
SHEET_CONTRACTS: Final[str] = "contract_details"
SHEET_ADDITIONS: Final[str] = "contract_additions"
SHEET_ORGANIZATIONS: Final[str] = "organization_profile"
SHEET_LOTS: Final[str] = "lots"
SHEET_LOT_DETAILS: Final[str] = "lots_details"
SHEET_REGISTRY: Final[str] = "registry"

CALC_HEADER_ROW: Final[int] = 3
"""Заголовки расчётного листа не в первой строке, а в третьей."""

CALC_FIRST_DATA_ROW: Final[int] = 4

EXPECTED_SHEETS: Final[int] = 15
EXPECTED_CONTRACTS: Final[int] = 355
EXPECTED_SUPPLIERS: Final[int] = 26
EXPECTED_DISTRICTS: Final[int] = 9
EXPECTED_ORGANIZATIONS: Final[int] = 3_668

BIN_LENGTH: Final[int] = 12

# Строки-заглушки, которые источник записал вместо пустых ячеек. Каждая из них
# обязана превращаться в «нет данных», а не в значение.
PLACEHOLDERS: Final[frozenset[str]] = frozenset({"", "-", "—", "–", "nan", "NaN", "None", "null"})

# Excel считает дни от 1899-12-30 (сдвиг учитывает несуществующее 29.02.1900).
_EXCEL_EPOCH: Final[datetime] = datetime(1899, 12, 30)

_ADDRESS_REGION = re.compile(r"Область:\s*([^,]+)")
_ADDRESS_DISTRICT = re.compile(r"Район:\s*([^,]+)")
_ADDRESS_CITY = re.compile(r"Город:\s*([^,]+)")

TERMINATED_MARKER: Final[str] = "асторгнут"
"""Подстрока без первой буквы: статус пишется и «Расторгнут», и «расторгнут»."""


def clean_text(value: object) -> str | None:
    """Текст ячейки или `None`, если это пустота или заглушка.

    Отдельная функция, потому что заглушек в книге три вида и встречаются они
    в разных колонках: `'—'` у 131 заказчика, `'nan'` у 3 способов закупки,
    пустые ячейки — везде.
    """
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return None if text.casefold() in {p.casefold() for p in PLACEHOLDERS} else text


def clean_number(value: object) -> float | None:
    """Число из ячейки, где оно может быть записано строкой с пробелами.

    `'11 953 000.00'` — обычный вид денежной суммы в этой книге, причём
    разделителем бывает и неразрывный пробел U+00A0.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, int | float):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(" ", "").strip()
    if text.casefold() in {p.casefold() for p in PLACEHOLDERS}:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def clean_bool(value: object) -> bool:
    """Булев признак из ячейки, где `True`/`False` записаны строками."""
    if isinstance(value, bool):
        return value
    return str(value).strip().casefold() == "true"


def normalize_bin(value: object) -> str | None:
    """БИН из 12 знаков с восстановленными ведущими нулями.

    В источниках БИН хранится числом, из-за чего у 763 организаций из 3 668
    ведущие нули потеряны: `ТОО ASIA GRAND GROUP` записан как `440010133`
    вместо `000440010133`. Восстанавливать обязательно на **обеих** сторонах
    join, иначе связь распадается ровно на этих 763 записях.
    """
    text = clean_text(value)
    if text is None:
        return None
    digits = text.split(".")[0].strip()
    return digits.zfill(BIN_LENGTH) if digits.isdigit() else digits


def excel_serial_to_date(value: object) -> date | None:
    """Дата из Excel-серийного числа.

    Нужна только для `contract_additions`: там все три даты — числа
    (`45415.549363425926`), причём в одной колонке смешаны int и float.
    В `contract_details` даты настоящие, и эта функция к ним не применяется.
    """
    number = clean_number(value)
    if number is None or number <= 0:
        return None
    return (_EXCEL_EPOCH + timedelta(days=number)).date()


def cell_to_date(value: object) -> date | None:
    """Дата из ячейки `contract_details`, где даты хранятся как datetime."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return excel_serial_to_date(value)


@dataclass(frozen=True, slots=True)
class LegalAddress:
    """Разобранный юридический адрес из реестра налогоплательщиков.

    Формат строки — структурированный, с метками:

        Страна: Казахстан, Область: Алматинская, Район: Карасайский, …

    Город республиканского значения приезжает сюда в поле «Область», и это
    приходится принимать как есть: `г. Алматы` в книге отнесён к Алматинской
    области именно потому, что так записан юридический адрес поставщика.
    """

    raw: str
    region: str | None
    district: str | None
    city: str | None

    @property
    def territory_name(self) -> str | None:
        """Название единицы, к которой привязывается договор.

        Район приоритетнее города: в адресе города областного значения район
        не указывается, и подмена одного другим создала бы несуществующую
        территорию.
        """
        return self.district or self.city


def parse_legal_address(raw: object) -> LegalAddress:
    """Разобрать юридический адрес по меткам «Область:» и «Район:»."""
    text = clean_text(raw) or ""
    region = _ADDRESS_REGION.search(text)
    district = _ADDRESS_DISTRICT.search(text)
    city = _ADDRESS_CITY.search(text)
    return LegalAddress(
        raw=text,
        region=region.group(1).strip() if region else None,
        district=district.group(1).strip() if district else None,
        city=city.group(1).strip() if city else None,
    )


@dataclass(frozen=True, slots=True)
class OrganizationRow:
    """Строка профиля организации — стыковка со слоем 8.7."""

    bin: str
    name: str
    in_rnu_gz: bool
    in_lzhepred_list: bool
    no_physical_activity: bool
    high_oked_diversity: bool
    mass_address: bool
    nominal_director: bool
    n_contracts: float | None
    max_direct_one_customer: float | None
    pct_terminated: float | None
    final_points_v2: int | None
    final_risk_level_v2: str | None

    @property
    def risk_profile(self) -> SupplierRiskProfile:
        return SupplierRiskProfile(
            bin=self.bin,
            name=self.name,
            in_rnu_gz=self.in_rnu_gz,
            in_lzhepred_list=self.in_lzhepred_list,
        )


@dataclass(frozen=True, slots=True)
class ContractRow:
    """Строка `contract_details` — ядро выборки."""

    contract_id: str
    supplier_bin: str
    announcement_number: str | None
    brief_content_ru: str | None
    subject_type: str | None
    planned_amount: float | None
    final_amount: float | None
    actual_amount: float | None
    planned_method: str | None
    actual_method: str | None
    planned_exec_date: date | None
    actual_exec_date: date | None
    contract_status: str | None

    @property
    def is_terminated(self) -> bool:
        return TERMINATED_MARKER in (self.contract_status or "")


@dataclass(frozen=True, slots=True)
class AdditionRow:
    """Строка `contract_additions` — версия договора."""

    contract_id: str
    creation_date: date | None
    conclusion_date: date | None
    planned_exec_date: date | None
    final_total_amount: float | None
    actual_total_amount: float | None
    justification: str | None

    @property
    def changes_term(self) -> bool:
        """Продлевает ли соглашение срок — признак для B5."""
        return "срок" in (self.justification or "").casefold()


@dataclass(frozen=True, slots=True)
class LotRow:
    """Строка листа `lots` — лот объявления."""

    announcement_id: str
    announcement_number: str | None
    submitted_bids: float | None
    lot_number: str | None
    customer: str | None
    lot_name: str | None
    planned_sum: float | None
    lot_status: str | None


@dataclass(frozen=True, slots=True)
class LotDetailRow:
    """Строка листа `lots_details` — детали лота, включая КАТО места поставки."""

    announcement_id: str
    lot_id: str | None
    customer_bin: str | None
    customer_name: str | None
    tru_code: str | None
    delivery_kato: str | None
    delivery_address: str | None


@dataclass(frozen=True, slots=True)
class CalcRow:
    """Строка расчётного листа — эталон книги для сверки.

    Хранится отдельно от входных данных: как только эти значения начнут
    подставляться в расчёт, сверка перестанет что-либо доказывать.
    """

    contract_id: str
    supplier_bin: str
    customer_truncated: str | None
    region: str
    district: str
    method: str | None
    final_amount: float | None
    indicators: dict[str, float | None]
    s_raw: float
    w_avail: int
    s_norm: float
    k: float
    risk_score: float
    level: str
    source_row_ref: str


def resolve_workbook(source_dir: Path) -> Path:
    """Найти книгу 8.4 в каталоге источников устойчиво к NFD."""
    from scripts.source_manifest import resolve_source

    return resolve_source(source_dir, SOURCE_FILE_NAME)


def _rows(workbook: Any, sheet_name: str, *, min_row: int) -> list[tuple[Any, ...]]:
    sheet = workbook[sheet_name]
    return [
        row for row in sheet.iter_rows(min_row=min_row, values_only=True) if row[0] is not None
    ]


@dataclass(slots=True)
class ProcurementWorkbook:
    """Разобранная книга 8.4 целиком.

    Все листы читаются за один проход: книга открывается дорого, а связи между
    листами нужны почти всем расчётам сразу.
    """

    path: Path
    sheet_names: list[str] = field(default_factory=list)
    contracts: dict[str, ContractRow] = field(default_factory=dict)
    additions: dict[str, list[AdditionRow]] = field(default_factory=dict)
    organizations: dict[str, OrganizationRow] = field(default_factory=dict)
    lots: list[LotRow] = field(default_factory=list)
    lot_details: list[LotDetailRow] = field(default_factory=list)
    addresses: dict[str, LegalAddress] = field(default_factory=dict)
    calc_rows: list[CalcRow] = field(default_factory=list)

    # --- производные связи ---------------------------------------------------

    def customer_of(self, contract_id: str) -> str | None:
        """Полное имя заказчика по номеру объявления.

        В расчётном листе имя обрезано до 60 знаков, из-за чего разные
        заказчики становятся неразличимы и группировки B3/B4 «съезжают».
        Здесь берётся полное имя из листа `lots`.
        """
        contract = self.contracts.get(contract_id)
        if contract is None or contract.announcement_number is None:
            return None
        return self._customer_by_announcement.get(contract.announcement_number)

    def bids_of(self, contract_id: str) -> float | None:
        """Число поданных заявок по объявлению договора."""
        contract = self.contracts.get(contract_id)
        if contract is None or contract.announcement_number is None:
            return None
        return self._bids_by_announcement.get(contract.announcement_number)

    @property
    def _customer_by_announcement(self) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for lot in self.lots:
            if lot.announcement_number and lot.customer:
                mapping.setdefault(lot.announcement_number, lot.customer)
        return mapping

    @property
    def _bids_by_announcement(self) -> dict[str, float]:
        mapping: dict[str, float] = {}
        for lot in self.lots:
            if lot.announcement_number and lot.submitted_bids is not None:
                mapping.setdefault(lot.announcement_number, lot.submitted_bids)
        return mapping


def load_workbook_8_4(path: Path) -> ProcurementWorkbook:
    """Прочитать книгу 8.4 целиком."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        parsed = ProcurementWorkbook(path=path, sheet_names=list(workbook.sheetnames))

        for row in _rows(workbook, SHEET_CONTRACTS, min_row=2):
            contract_id = clean_text(row[0])
            supplier_bin = normalize_bin(row[1])
            if contract_id is None or supplier_bin is None:
                continue
            parsed.contracts[contract_id] = ContractRow(
                contract_id=contract_id,
                supplier_bin=supplier_bin,
                announcement_number=clean_text(row[2]),
                brief_content_ru=clean_text(row[3]),
                subject_type=clean_text(row[4]),
                planned_amount=clean_number(row[5]),
                final_amount=clean_number(row[6]),
                actual_amount=clean_number(row[7]),
                planned_method=clean_text(row[8]),
                actual_method=clean_text(row[9]),
                planned_exec_date=cell_to_date(row[10]),
                actual_exec_date=cell_to_date(row[11]),
                contract_status=clean_text(row[12]),
            )

        additions: dict[str, list[AdditionRow]] = defaultdict(list)
        for row in _rows(workbook, SHEET_ADDITIONS, min_row=2):
            contract_id = clean_text(row[0])
            if contract_id is None:
                continue
            additions[contract_id].append(
                AdditionRow(
                    contract_id=contract_id,
                    creation_date=excel_serial_to_date(row[1]),
                    conclusion_date=excel_serial_to_date(row[2]),
                    planned_exec_date=excel_serial_to_date(row[3]),
                    final_total_amount=clean_number(row[4]),
                    actual_total_amount=clean_number(row[5]),
                    justification=clean_text(row[6]),
                )
            )
        parsed.additions = dict(additions)

        for row in _rows(workbook, SHEET_ORGANIZATIONS, min_row=2):
            org_bin = normalize_bin(row[0])
            if org_bin is None:
                continue
            parsed.organizations[org_bin] = OrganizationRow(
                bin=org_bin,
                name=clean_text(row[1]) or "",
                in_rnu_gz=clean_bool(row[2]),
                in_lzhepred_list=clean_bool(row[3]),
                no_physical_activity=clean_bool(row[4]),
                high_oked_diversity=clean_bool(row[5]),
                mass_address=clean_bool(row[6]),
                nominal_director=clean_bool(row[7]),
                n_contracts=clean_number(row[8]),
                max_direct_one_customer=clean_number(row[9]),
                pct_terminated=clean_number(row[10]),
                final_points_v2=int(clean_number(row[11]) or 0),
                final_risk_level_v2=clean_text(row[12]),
            )

        for row in _rows(workbook, SHEET_LOTS, min_row=2):
            announcement_id = clean_text(row[0])
            if announcement_id is None:
                continue
            parsed.lots.append(
                LotRow(
                    announcement_id=announcement_id,
                    announcement_number=clean_text(row[1]),
                    submitted_bids=clean_number(row[2]),
                    lot_number=clean_text(row[3]),
                    customer=clean_text(row[4]),
                    lot_name=clean_text(row[5]),
                    planned_sum=clean_number(row[10]),
                    lot_status=clean_text(row[11]),
                )
            )

        for row in _rows(workbook, SHEET_LOT_DETAILS, min_row=2):
            announcement_id = clean_text(row[0])
            if announcement_id is None:
                continue
            kato, address = _split_kato(row[14])
            parsed.lot_details.append(
                LotDetailRow(
                    announcement_id=announcement_id,
                    lot_id=clean_text(row[1]),
                    customer_bin=normalize_bin(row[4]),
                    customer_name=clean_text(row[5]),
                    tru_code=clean_text(row[6]),
                    delivery_kato=kato,
                    delivery_address=address,
                )
            )

        for row in _rows(workbook, SHEET_REGISTRY, min_row=2):
            registry_bin = normalize_bin(row[0])
            if registry_bin is None:
                continue
            parsed.addresses[registry_bin] = parse_legal_address(row[3])

        for offset, row in enumerate(_rows(workbook, SHEET_CALC, min_row=CALC_FIRST_DATA_ROW)):
            parsed.calc_rows.append(_build_calc_row(row, excel_row=CALC_FIRST_DATA_ROW + offset))

        return parsed
    finally:
        workbook.close()


def _split_kato(value: object) -> tuple[str | None, str | None]:
    """Отделить код КАТО от адреса: он склеен с ним до первой запятой.

    `'101010000, область Абай, г.Семей…'` — код и адрес в одной ячейке.
    В геопривязке слоя 8.4 КАТО не используется (заполнен у 191 лота из 381),
    но выбрасывать его при импорте нельзя: это единственный официальный код в
    книге, и он понадобится, когда появится справочник соответствий.
    """
    text = clean_text(value)
    if text is None:
        return None, None
    head, _, tail = text.partition(",")
    head = head.strip()
    if head.isdigit():
        return head, tail.strip() or None
    return None, text


def _build_calc_row(row: tuple[Any, ...], *, excel_row: int) -> CalcRow:
    indicators: dict[str, float | None] = {}
    for index in range(1, 10):
        raw = row[7 + index - 1]
        indicators[f"B{index}"] = None if raw is None else float(raw)
    return CalcRow(
        contract_id=str(row[0]).strip(),
        supplier_bin=normalize_bin(row[1]) or "",
        customer_truncated=clean_text(row[2]),
        region=str(row[3]).strip(),
        district=str(row[4]).strip(),
        method=clean_text(row[5]),
        final_amount=clean_number(row[6]),
        indicators=indicators,
        s_raw=float(clean_number(row[16]) or 0.0),
        w_avail=int(clean_number(row[17]) or 0),
        s_norm=float(clean_number(row[18]) or 0.0),
        k=float(clean_number(row[19]) or 0.0),
        risk_score=float(clean_number(row[20]) or 0.0),
        level=str(row[21]).strip(),
        source_row_ref=f"{SHEET_CALC}!A{excel_row}",
    )


def derive_indicators(book: ProcurementWorkbook) -> dict[str, dict[str, float | None]]:
    """Независимо вывести B1…B9 из сырых листов книги.

    Возвращает `contract_id → {код метрики: значение или None}`. B8 в словарь
    не попадает: он требует справочника ОКЭД (`oked.csv`, 78 375 строк),
    который в книгу не вложен. Заполнять его нулём нельзя — это ровно та
    подстановка «нет данных = нет риска», которую методика запрещает.

    Функция нужна для проверки, а не для замены значений книги: расчётный лист
    формул не содержит, и без независимого пересчёта утверждать, что методика
    воспроизводится, было бы нечем.
    """
    customer_contracts: Counter[str] = Counter()
    pair_contracts: Counter[tuple[str, str]] = Counter()
    pair_one_source: Counter[tuple[str, str]] = Counter()

    for calc in book.calc_rows:
        customer = book.customer_of(calc.contract_id)
        if customer is None:
            continue
        customer_contracts[customer] += 1
        pair_contracts[(customer, calc.supplier_bin)] += 1
        if calc.method and "одного источника" in calc.method.casefold():
            pair_one_source[(customer, calc.supplier_bin)] += 1

    derived: dict[str, dict[str, float | None]] = {}
    for calc in book.calc_rows:
        contract_id = calc.contract_id
        customer = book.customer_of(contract_id)
        organization = book.organizations.get(calc.supplier_bin)
        additions = sorted(
            book.additions.get(contract_id, []),
            key=lambda a: a.creation_date or date.min,
        )

        values: dict[str, float | None] = {
            # Способ закупки берётся из расчётного листа, а не из
            # `contract_details`: у шести договоров сырьё пусто, а способ в
            # расчётном листе есть — генератор располагал источником, которого
            # в книге нет. Заглушка 'nan' при этом остаётся «не измерено».
            "B1": derive_b1(calc.method),
            "B2": derive_b2(book.bids_of(contract_id)),
            "B3": (
                derive_b3(pair_contracts[(customer, calc.supplier_bin)],
                          customer_contracts[customer])
                if customer is not None
                else None
            ),
            "B4": (
                derive_b4(pair_one_source[(customer, calc.supplier_bin)])
                if customer is not None
                else None
            ),
            "B5": derive_b5(sum(1 for a in additions if a.changes_term)),
            "B6": derive_b6([a.final_total_amount for a in additions if a.final_total_amount]),
            "B7": (
                derive_b7(organization.no_physical_activity, organization.n_contracts)
                if organization is not None
                else None
            ),
            "B8": None,
            "B9": (
                derive_b9(
                    organization.nominal_director,
                    organization.mass_address,
                    organization.high_oked_diversity,
                )
                if organization is not None
                else None
            ),
        }
        derived[contract_id] = values
    return derived


def build_contract_inputs(book: ProcurementWorkbook) -> list[ContractRiskInputs]:
    """Собрать входные данные расчёта по каждому договору.

    Значения метрик берутся из расчётного листа — он единственный содержит B8,
    невыводимый без отсутствующего справочника ОКЭД. Исключение — B1: его
    значение в листе противоречит собственному же `W_avail` книги у трёх
    договоров, поэтому метрика выводится из колонки «Способ» заново. Подробный
    разбор — в `docs/assumptions-and-gaps.md`.
    """
    inputs: list[ContractRiskInputs] = []
    for calc in book.calc_rows:
        contract = book.contracts.get(calc.contract_id)
        organization = book.organizations.get(calc.supplier_bin)
        supplier = (
            organization.risk_profile
            if organization is not None
            else SupplierRiskProfile(bin=calc.supplier_bin)
        )

        indicators = dict(calc.indicators)
        indicators["B1"] = derive_b1(calc.method)

        inputs.append(
            ContractRiskInputs(
                contract_id=calc.contract_id,
                supplier=supplier,
                district=calc.district,
                region=calc.region,
                indicators=indicators,
                final_amount=calc.final_amount,
                is_terminated=contract.is_terminated if contract is not None else False,
                customer=book.customer_of(calc.contract_id),
            )
        )
    return inputs


def evaluate_contracts(book: ProcurementWorkbook) -> list[ContractRiskResult]:
    """Посчитать риск по всем договорам книги."""
    return [evaluate_contract(item) for item in build_contract_inputs(book)]


def supplier_territories(book: ProcurementWorkbook) -> dict[str, str | None]:
    """Район поставщика по юридическому адресу — `БИН → название района`."""
    territories: dict[str, str | None] = {}
    for calc in book.calc_rows:
        address = book.addresses.get(calc.supplier_bin)
        territories[calc.supplier_bin] = address.territory_name if address else None
    return territories


__all__ = [
    "BIN_LENGTH",
    "CALC_FIRST_DATA_ROW",
    "CALC_HEADER_ROW",
    "EXPECTED_CONTRACTS",
    "EXPECTED_DISTRICTS",
    "EXPECTED_ORGANIZATIONS",
    "EXPECTED_SHEETS",
    "EXPECTED_SUPPLIERS",
    "LAYER_CODE",
    "PLACEHOLDERS",
    "SHEET_ADDITIONS",
    "SHEET_CALC",
    "SHEET_CONTRACTS",
    "SHEET_LOTS",
    "SHEET_ORGANIZATIONS",
    "SHEET_REGISTRY",
    "SOURCE_FILE_NAME",
    "AdditionRow",
    "CalcRow",
    "ContractRow",
    "LegalAddress",
    "LotDetailRow",
    "LotRow",
    "OrganizationRow",
    "ProcurementWorkbook",
    "build_contract_inputs",
    "cell_to_date",
    "clean_bool",
    "clean_number",
    "clean_text",
    "derive_indicators",
    "evaluate_contracts",
    "excel_serial_to_date",
    "load_workbook_8_4",
    "normalize_bin",
    "parse_legal_address",
    "resolve_workbook",
    "supplier_territories",
]
