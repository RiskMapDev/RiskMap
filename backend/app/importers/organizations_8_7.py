"""Импорт слоя 8.7 — хозяйствующие субъекты.

Источник: книга `Слой_8.7_организации_риски_v1.xlsx`, лист `Расчёт рисков`,
заголовок в третьей строке, 3668 строк данных.

Две ловушки, ради которых этот модуль вообще существует отдельно от общего
чтения Excel.

**Имя файла хранится в Unicode NFD.** Буква «й» записана как «и» + U+0306, и
прямое сравнение строкового литерала с именем файла даёт «файл не найден», хотя
файл на месте. Поэтому путь собирается не конкатенацией, а поиском по каталогу
через :func:`scripts.source_manifest.resolve_source` — единственную функцию, где
это правило записано.

**БИН хранится целым числом.** 763 значения из 3668 (20.8 %) потеряли ведущие
нули: `90340012684` вместо `090340012684`. Джойн с другими слоями без
восстановления теряет пятую часть связей. Восстановление выполняет
:func:`normalize_bin`, и факт восстановления сохраняется — он должен быть виден
в отчёте о качестве данных, а не исчезнуть после `zfill`.

Про сам расчёт. Лист `Расчёт рисков` содержит уже приведённые значения v по
четырём индикаторам, но не содержит их исходных величин: числа организаций по
адресу, числа секций ОКЭД и числа компаний у руководителя в книге нет — они
остались в CSV `organization_profile_MASTER (2).csv`, которого в комплекте
исходников нет. Поэтому импортёр берёт v как измерение, а не пересчитывает
его, и это ограничение зафиксировано явно, а не спрятано.
"""

from __future__ import annotations

import unicodedata
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.risk.core import IndicatorValue
from app.risk.layers.organizations import (
    NOT_CONNECTED,
    ORGANIZATION_MODEL,
    category_a_fact,
    unmeasured,
)
from scripts.source_manifest import resolve_source

SOURCE_FILE_NAME = "Слой_8.7_организации_риски_v1.xlsx"
"""Имя книги в NFC. Сравнивать с именем на диске напрямую нельзя — см. модульный docstring."""

RISK_SHEET = "Расчёт рисков"
HEADER_ROW = 3

BIN_LENGTH = 12

EXPECTED_ROW_COUNT = 3668
"""Контрольное число строк. Расхождение означает, что подменили книгу."""

MEASURED_CODES: tuple[str, ...] = ("B3", "B5", "B6", "B8")
"""Индикаторы, у которых источник подключён. Суммарный вес 45 из 110."""

_UNCONNECTED_CODES: tuple[str, ...] = ("B1", "B2", "B4", "B7", "B9")


def normalize_bin(value: object) -> str:
    """Привести БИН к канонической форме из 12 знаков.

    Источник хранит БИН целым числом, поэтому ведущие нули теряются. Дополнение
    слева нулями согласуется с JSON-схемой книги, где `bin` объявлен строкой с
    шаблоном `^[0-9]{12}$`.

    Значение длиннее 12 знаков или содержащее не цифры — не «плохо
    отформатированный БИН», а другая сущность, и молча подрезать его нельзя.
    """
    text = str(value).strip()
    if not text or not text.isdigit():
        raise ValueError(f"БИН {value!r} не является последовательностью цифр")
    if len(text) > BIN_LENGTH:
        raise ValueError(f"БИН {text!r} длиннее {BIN_LENGTH} знаков")
    return text.zfill(BIN_LENGTH)


def bin_leading_zeros_lost(value: object) -> bool:
    """Потерял ли исходный БИН ведущие нули при выгрузке."""
    return len(str(value).strip()) < BIN_LENGTH


@dataclass(frozen=True, slots=True)
class OrganizationRow:
    """Строка витрины риска слоя 8.7.

    Поля `book_*` — величины, посчитанные самой книгой. Они нужны не для
    расчёта, а для сверки: собственный расчёт обязан воспроизводить их
    построчно, и расхождение должно быть видно как расхождение, а не
    исправлено подгонкой.
    """

    row_number: int
    """Номер строки на листе — для трассировки до источника."""

    bin: str
    bin_raw: str
    leading_zeros_restored: bool

    name: str

    b3_value: float | None
    b5_value: float | None
    b6_value: float | None
    b8_value: float | None

    is_category_a: bool

    book_raw_score: float
    book_available_weight: float
    book_score: float
    book_completeness_percent: float
    book_level_preliminary: str
    book_level_strict: str
    book_explanation: str

    @property
    def source_row_ref(self) -> str:
        return f"{RISK_SHEET}!A{self.row_number}"


def _cell_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    return float(value)  # type: ignore[arg-type]


def _rows(path: Path, sheet: str, header_row: int) -> Iterator[tuple[int, dict[str, object]]]:
    """Прочитать лист, отдавая номер строки вместе со словарём значений.

    Номер строки нужен всегда: без него замечание о качестве данных невозможно
    привязать к месту в книге.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        header: list[str] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index < header_row:
                continue
            if index == header_row:
                header = ["" if cell is None else str(cell) for cell in row]
                continue
            if all(cell is None for cell in row):
                continue
            yield index, dict(zip(header, row, strict=False))
    finally:
        workbook.close()


def read_organizations(path: Path) -> list[OrganizationRow]:
    """Прочитать витрину риска организаций."""
    result: list[OrganizationRow] = []
    for index, row in _rows(path, RISK_SHEET, HEADER_ROW):
        raw_bin = str(row["БИН"]).strip()
        result.append(
            OrganizationRow(
                row_number=index,
                bin=normalize_bin(raw_bin),
                bin_raw=raw_bin,
                leading_zeros_restored=bin_leading_zeros_lost(raw_bin),
                name=str(row["Наименование"]).strip(),
                b3_value=_cell_float(row["B3 v"]),
                b5_value=_cell_float(row["B5 v"]),
                b6_value=_cell_float(row["B6 v"]),
                b8_value=_cell_float(row["B8 v"]),
                # Колонка заполнена словом «ДА» ровно у 23 организаций и пуста
                # у остальных. Пусто — это «факт не подтверждён по подключённому
                # источнику», а не «неизвестно»: реестр РНУ госзакупок подключён.
                is_category_a=bool(row["Кат. A"]),
                book_raw_score=float(row["S_raw"]),  # type: ignore[arg-type]
                book_available_weight=float(row["W_avail"]),  # type: ignore[arg-type]
                book_score=float(row["Балл 0-100"]),  # type: ignore[arg-type]
                book_completeness_percent=float(row["Полнота %"]),  # type: ignore[arg-type]
                book_level_preliminary=str(row["Уровень (предв.)"]),
                book_level_strict=str(row["Уровень (строгий)"]),
                book_explanation=str(row["Расшифровка (ТЗ п.14)"]),
            )
        )
    return result


def read_organizations_from_source_dir(source_dir: Path) -> list[OrganizationRow]:
    """Найти книгу в каталоге источников и прочитать её.

    Путь ищется по нормализованному имени: имя файла на диске записано в NFD.
    """
    return read_organizations(resolve_source(source_dir, SOURCE_FILE_NAME))


def indicator_values(row: OrganizationRow) -> dict[str, IndicatorValue]:
    """Собрать значения индикаторов для расчёта.

    Неподключённые индикаторы передаются явно, с причиной. Можно было бы их не
    передавать вовсе — ядро всё равно посчитает их неизмеренными, — но тогда
    причина «нет публичного API» не дошла бы до карточки риска, и пользователь
    увидел бы «не измерено» без объяснения.
    """
    values: dict[str, IndicatorValue] = {
        "A1": category_a_fact("A1", confirmed=row.is_category_a),
        "A2": category_a_fact("A2", confirmed=None),
        "A3": category_a_fact("A3", confirmed=None),
        "A4": category_a_fact("A4", confirmed=None),
        "B3": IndicatorValue(code="B3", value=row.b3_value),
        "B5": IndicatorValue(code="B5", value=row.b5_value),
        "B6": (
            IndicatorValue(code="B6", value=row.b6_value)
            if row.b6_value is not None
            else unmeasured("B6", "сведения об ОКЭД отсутствуют")
        ),
        "B8": (
            IndicatorValue(code="B8", value=row.b8_value)
            if row.b8_value is not None
            else unmeasured("B8", "ИИН руководителя неизвестен")
        ),
    }
    for code in _UNCONNECTED_CODES:
        values[code] = unmeasured(code, NOT_CONNECTED)
    return values


def bin_index(rows: Sequence[OrganizationRow]) -> Mapping[str, OrganizationRow]:
    """Указатель по каноническому БИН — точка стыковки с остальными слоями.

    БИН здесь единственный ключ: ни адреса, ни КАТО, ни координат в слое нет.
    """
    index: dict[str, OrganizationRow] = {}
    for row in rows:
        if row.bin in index:
            raise ValueError(
                f"Дубль БИН {row.bin} в строках "
                f"{index[row.bin].row_number} и {row.row_number}"
            )
        index[row.bin] = row
    return index


def normalize_file_name(name: str) -> str:
    """Свёртка имени файла — та же, что в манифесте источников."""
    return unicodedata.normalize("NFC", name).casefold()


MODEL = ORGANIZATION_MODEL
"""Модель, которой считается этот слой. Ссылка нужна, чтобы версия была одна."""


__all__ = [
    "BIN_LENGTH",
    "EXPECTED_ROW_COUNT",
    "HEADER_ROW",
    "MEASURED_CODES",
    "MODEL",
    "RISK_SHEET",
    "SOURCE_FILE_NAME",
    "OrganizationRow",
    "bin_index",
    "bin_leading_zeros_lost",
    "indicator_values",
    "normalize_bin",
    "normalize_file_name",
    "read_organizations",
    "read_organizations_from_source_dir",
]
