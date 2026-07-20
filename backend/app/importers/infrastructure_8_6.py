"""Импорт слоя 8.6 — инфраструктурные и инвестиционные проекты.

Источник: книга `Слой_8_6_инфраструктурные_проекты_риски_v1.xlsx`, 14 листов.
Импортёр читает **две несвязанные популяции по отдельности** и намеренно не
предоставляет никакой функции, которая свела бы их в один список: общего ключа
между проектами ГЧП и заключениями экспертизы не существует, и любая такая
функция была бы приглашением к ошибке.

Ловушки этой книги, каждая из которых уже ломала расчёт.

**Имя файла в Unicode NFD** — путь собирается поиском по каталогу, а не
конкатенацией (:func:`scripts.source_manifest.resolve_source`).

**Два пробела в имени листа** `Данные  Экспертиза инфр проект`. Обращение по
«очевидному» имени с одним пробелом даёт `KeyError`.

**Двухуровневая шапка** в `Данные Проекты ГЧП`: заголовок размазан по строкам
3–4, а строка 7 — служебная нумерация колонок 1..22, которую нужно отбросить.
Поэтому лист читается по номерам колонок, а не по заголовкам: собирать имена из
двух строк ради четырёх безымянных колонок — способ получить тихий сдвиг.

**Срезанные ведущие нули** в регистрационном номере витрины экспертизы. Прямое
пересечение с сырым реестром даёт 0 совпадений из 4842, после дополнения нулями
до шести знаков — 4842 из 4842.

**Ноль как «не заполнено»** в объёме инвестиций: при нуле индикатор роста
инвестзатрат недоступен, а не равен нулю.

**Строковые даты-композиты** вида `04.11.2019 (осн.договор); 10.10.2024 (ДС
№004…)`. Разобрать их однозначно нельзя, и индикатор просрочки для таких строк
недоступен.

Отдельно про ключи группировки. Индикаторы концентрации требуют устойчивого
имени контрагента, но нормализация не универсальна: для частного партнёра нужна
свёртка до букв и цифр (иначе A2 и A3 расходятся с книгой на 12 и 4 строках), а
для государственного партнёра, заказчика и генпроектировщика — наоборот, строка
как она лежит в источнике, включая хвостовые пробелы (иначе расходится A4). Это
несогласованность самой книги, и воспроизводить её приходится буквально;
попытка «причесать» ключи меняет результат.
"""

from __future__ import annotations

import datetime as dt
import re
import statistics
from collections import Counter, defaultdict
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.risk.core import IndicatorValue
from app.risk.layers.infrastructure import (
    A2_SCALE,
    A3_SCALE,
    A4_SCALE,
    A5_OVERDUE_DAYS_CRITICAL,
    A6_SCALE,
    B2_SCALE,
    B3_ABSENT_VALUE,
    B3_NOT_AGREED_VALUE,
    B5_SCALE,
    B6_SCALE,
    expertise_significance_k,
    graded,
    ppp_significance_k,
)
from scripts.source_manifest import resolve_source

SOURCE_FILE_NAME = "Слой_8_6_инфраструктурные_проекты_риски_v1.xlsx"

SHEET_PPP_RAW = "Данные Проекты ГЧП"
SHEET_EXPERTISE_RAW = "Данные  Экспертиза инфр проект"
"""Два пробела после «Данные» — так в книге. Ошибка в имени даёт KeyError."""

SHEET_PPP_RISK = "Расчёт — Проекты ГЧП"
SHEET_EXPERTISE_RISK = "Расчёт — Объекты строит."
SHEET_CONTESTS = "Данные Конкурсы ГЧП"
SHEET_CONTRACTS = "Данные Договоры ГЧП"

PPP_FIRST_DATA_ROW = 8
"""Строки 1–6 — заголовки и объединённая шапка, строка 7 — нумерация колонок."""

EXPERTISE_HEADER_ROW = 1
RISK_HEADER_ROW = 3
CONTESTS_HEADER_ROW = 1
CONTRACTS_HEADER_ROW = 2

EXPECTED_PPP_COUNT = 1323
EXPECTED_EXPERTISE_COUNT = 4842
EXPECTED_CONTEST_COUNT = 514
EXPECTED_CONTRACT_COUNT = 12

REGISTRATION_NUMBER_LENGTH = 6

BOOK_CALCULATION_DATE = dt.date(2026, 7, 17)
"""Дата расчёта книги. Нужна для воспроизводимости индикатора просрочки."""

TERMINATED_STATUS = "расторгнут"
OPERATION_STATUS_MARKER = "эксплуатац"

_SQUASH_PATTERN = re.compile(r"[^0-9a-zа-яё]")
_HAZARD_PATTERN = re.compile(r"^\s*[12]\s+класс опасности")
_CORRECTION_MARKER = "корректировк"

# Колонки листа `Данные Проекты ГЧП`, нумерация с нуля.
_C_NUMBER = 0
_C_REGION = 1
_C_LEVEL = 2
_C_TITLE = 3
_C_OBJECT_KIND = 4
_C_STATUS = 5
_C_CAPACITY = 6
_C_SECTOR = 7
_C_INITIATIVE = 8
_C_CONTRACT_DATE = 9
_C_BUILD_START = 10
_C_BUILD_END = 11
_C_OPERATION_START = 12
_C_OPERATION_END = 13
_C_CONTRACT_KIND = 14
_C_GOV_PARTNER = 15
_C_PRIVATE_PARTNER = 16
_C_COST = 17
_C_INVESTMENTS = 18
_C_GOV_FORM = 19
_C_CONTACTS = 20
_C_URL = 21


# --- Нормализация ------------------------------------------------------------


def squash(value: object) -> str:
    """Свернуть наименование до букв и цифр в нижнем регистре.

    Применяется там, где книга сравнивает контрагентов по существу: частный
    партнёр в индикаторах A2 и A3, а также пара «наименование + заказчик»,
    определяющая тождество объекта экспертизы. Именно такая свёртка
    воспроизводит значения книги на всех строках без расхождений.
    """
    return _SQUASH_PATTERN.sub("", str(value).casefold())


def as_stored(value: object) -> str:
    """Ключ группировки «как в источнике», включая хвостовые пробелы.

    Выглядит небрежно, но это буквальное поведение книги для государственного
    партнёра, заказчика и генпроектировщика. Обрезка пробелов меняет значение
    индикатора A4 у проекта № 908: госпартнёр с хвостовыми пробелами образует
    отдельную группу, в которой меньше трёх проектов, и индикатор становится
    недоступным.
    """
    return "" if value is None else str(value)


def clean_text(value: object) -> str:
    """Текст без пустых суррогатов.

    Строка «nan» появляется при выгрузке из pandas и означает пустую ячейку,
    а не наименование.
    """
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.casefold() == "nan" else text


def normalize_registration_number(value: object) -> str:
    """Восстановить шестизначный регистрационный номер заключения.

    В витрине книги ведущие нули срезаны (`5617`), в сыром реестре они на месте
    (`005617`). Без восстановления джойн даёт ноль совпадений из 4842.
    """
    text = str(value).strip()
    if not text.isdigit():
        raise ValueError(f"Регистрационный номер {value!r} не является числом")
    if len(text) > REGISTRATION_NUMBER_LENGTH:
        raise ValueError(f"Регистрационный номер {text!r} длиннее {REGISTRATION_NUMBER_LENGTH}")
    return text.zfill(REGISTRATION_NUMBER_LENGTH)


def parse_source_date(value: object) -> dt.date | None:
    """Разобрать дату периода реализации проекта ГЧП.

    Источник смешивает три вида значений: настоящие даты, голый год числом и
    строки-композиты вроде «04.11.2019 (осн.договор); 10.10.2024 (ДС №004…)».
    Первые два разбираются, третий — нет: выбрать одну дату из композита можно
    только гаданием, а гадание здесь стоит дороже, чем недоступный индикатор.
    """
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, int) and 1900 < value < 2100:
        # Год без месяца и дня. Берётся конец года: плановый срок «до конца
        # 2024» просрочен не с первого января.
        return dt.date(value, 12, 31)
    return None


def parse_number(value: object) -> float | None:
    """Число из ячейки. Ноль считается незаполненным значением.

    В реестре проектов ноль в стоимости и в объёме инвестиций стоит там, где
    значения нет. Отличить «инвестиций ноль» от «поле не заполнено» по данным
    нельзя, и книга трактует ноль как отсутствие: у 348 проектов индикатор
    роста инвестзатрат недоступен именно поэтому.
    """
    if isinstance(value, (int, float)) and not isinstance(value, bool) and value != 0:
        return float(value)
    return None


def is_terminated(status: object) -> bool:
    """Расторгнут ли договор.

    Регистр в источнике плавает: «Расторгнут» 150 раз и «расторгнут» 2 раза.
    Сравнение без свёртки регистра даёт 150 вместо 152.
    """
    return clean_text(status).casefold() == TERMINATED_STATUS


# --- Чтение листов -----------------------------------------------------------


def _sheet_rows(
    path: Path, sheet: str, header_row: int
) -> Iterator[tuple[int, dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        header: list[str] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index < header_row:
                continue
            if index == header_row:
                header = ["" if cell is None else str(cell) for cell in row]
                continue
            if all(cell is None for cell in row):
                continue
            yield index, dict(zip(header, row, strict=False))
    finally:
        workbook.close()


def _positional_rows(path: Path, sheet: str, first_row: int) -> Iterator[tuple[int, list[object]]]:
    """Читать лист по номерам колонок — для двухуровневой шапки."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index < first_row:
                continue
            if all(cell is None for cell in row):
                continue
            yield index, list(row)
    finally:
        workbook.close()


@dataclass(frozen=True, slots=True)
class PppProjectRow:
    """Проект ГЧП из сырого реестра."""

    row_number: int
    registry_number: int

    region_raw: str
    project_level: str
    title: str
    object_kind: str
    status_raw: str
    capacity: str
    sector: str
    initiative_kind: str

    contract_date: dt.date | None
    construction_start: dt.date | None
    construction_end: dt.date | None
    operation_start: dt.date | None
    operation_end: dt.date | None

    contract_kind: str
    government_partner_raw: str
    private_partner_raw: str

    cost_initial: float | None
    investments: float | None
    government_participation_form: str
    source_url: str

    @property
    def private_partner_key(self) -> str:
        return squash(self.private_partner_raw)

    @property
    def government_partner_key(self) -> str:
        return as_stored(self.government_partner_raw)

    @property
    def is_terminated(self) -> bool:
        return is_terminated(self.status_raw)

    @property
    def is_republican(self) -> bool:
        return "республиканск" in self.project_level.casefold()

    @property
    def has_date_error(self) -> bool:
        """Окончание строительства раньше начала — логическая ошибка в данных."""
        start, end = self.construction_start, self.construction_end
        return start is not None and end is not None and end < start

    @property
    def source_row_ref(self) -> str:
        return f"{SHEET_PPP_RAW}!A{self.row_number}"


def read_ppp_projects(path: Path) -> list[PppProjectRow]:
    """Прочитать сырой реестр проектов ГЧП."""
    rows: list[PppProjectRow] = []
    for index, cells in _positional_rows(path, SHEET_PPP_RAW, PPP_FIRST_DATA_ROW):
        number = cells[_C_NUMBER]
        if not isinstance(number, int):
            continue
        rows.append(
            PppProjectRow(
                row_number=index,
                registry_number=number,
                region_raw=clean_text(cells[_C_REGION]),
                project_level=clean_text(cells[_C_LEVEL]),
                title=clean_text(cells[_C_TITLE]),
                object_kind=clean_text(cells[_C_OBJECT_KIND]),
                status_raw=clean_text(cells[_C_STATUS]),
                capacity=clean_text(cells[_C_CAPACITY]),
                sector=clean_text(cells[_C_SECTOR]),
                initiative_kind=clean_text(cells[_C_INITIATIVE]),
                contract_date=parse_source_date(cells[_C_CONTRACT_DATE]),
                construction_start=parse_source_date(cells[_C_BUILD_START]),
                construction_end=parse_source_date(cells[_C_BUILD_END]),
                operation_start=parse_source_date(cells[_C_OPERATION_START]),
                operation_end=parse_source_date(cells[_C_OPERATION_END]),
                contract_kind=clean_text(cells[_C_CONTRACT_KIND]),
                # Госпартнёр сохраняется без обрезки пробелов — см. as_stored().
                government_partner_raw=as_stored(cells[_C_GOV_PARTNER]),
                private_partner_raw=clean_text(cells[_C_PRIVATE_PARTNER]),
                cost_initial=parse_number(cells[_C_COST]),
                investments=parse_number(cells[_C_INVESTMENTS]),
                government_participation_form=clean_text(cells[_C_GOV_FORM]),
                source_url=clean_text(cells[_C_URL]),
            )
        )
    return rows


@dataclass(frozen=True, slots=True)
class ExpertiseConclusionRow:
    """Заключение государственной экспертизы ПСД из сырого реестра.

    Единица — заключение, а не объект. Тождество объекта задаётся
    :attr:`object_identity_key`, и только по нему можно считать объекты.
    """

    row_number: int
    registration_number: str

    author_supervision_status: str
    work_kind: str
    design_stage: str
    industry: str
    object_kind: str
    title: str
    customer_raw: str
    designer_raw: str
    location_raw: str
    conclusion_number: str
    issue_date_raw: str
    capacity: str
    capacity_unit: str
    has_cost_estimate: bool | None
    technological_complexity: str
    responsibility_level: str
    hazard_class: str
    category: str
    efficiency_class: str
    funding_source: str
    expertise_place: str
    full_set_cost: str
    external_id: int | None

    @property
    def object_identity_key(self) -> tuple[str, str]:
        return squash(self.title), squash(self.customer_raw)

    @property
    def customer_key(self) -> str:
        return as_stored(self.customer_raw)

    @property
    def designer_key(self) -> str:
        return as_stored(self.designer_raw)

    @property
    def has_correction(self) -> bool:
        """Наименование содержит признак корректировки ПСД.

        Проверять нужно на полном наименовании из сырого реестра: витрина
        обрезает его до 180 знаков, и на обрезанном тексте признак теряется у
        172 строк из 4842.
        """
        return _CORRECTION_MARKER in self.title.casefold()

    @property
    def is_high_hazard(self) -> bool:
        return bool(_HAZARD_PATTERN.match(self.hazard_class))

    @property
    def is_first_responsibility_level(self) -> bool:
        return self.responsibility_level.strip().startswith("1 уровень")

    @property
    def source_row_ref(self) -> str:
        return f"{SHEET_EXPERTISE_RAW}!A{self.row_number}"


def _parse_yes_no(value: object) -> bool | None:
    text = clean_text(value).casefold()
    if text == "да":
        return True
    if text == "нет":
        return False
    return None


def read_expertise_conclusions(path: Path) -> list[ExpertiseConclusionRow]:
    """Прочитать сырой реестр заключений экспертизы."""
    rows: list[ExpertiseConclusionRow] = []
    for index, row in _sheet_rows(path, SHEET_EXPERTISE_RAW, EXPERTISE_HEADER_ROW):
        external_id = row.get("ID")
        rows.append(
            ExpertiseConclusionRow(
                row_number=index,
                registration_number=normalize_registration_number(row["Регистрационный номер"]),
                author_supervision_status=clean_text(row["Статус авторского договора"]),
                work_kind=clean_text(row["Вид работ"]),
                design_stage=clean_text(row["Стадии проектирования"]),
                industry=clean_text(row["Отрасль строительства"]),
                object_kind=clean_text(row["Вид объекта"]),
                title=clean_text(row["Наименование объекта"]),
                customer_raw=as_stored(row["Заказчик строительства"]),
                designer_raw=as_stored(row["Генеральный проектировщик"]),
                location_raw=clean_text(row["Местоположение объекта"]),
                conclusion_number=clean_text(row["Номер заключения"]),
                issue_date_raw=clean_text(row["Дата выдачи заключения"]),
                capacity=clean_text(row["Мощность"]),
                capacity_unit=clean_text(row["Единица измерения мощности"]),
                has_cost_estimate=_parse_yes_no(row["Имеется сметная документация?"]),
                technological_complexity=clean_text(row["Технологическая сложность"]),
                responsibility_level=clean_text(row["Уровень ответственности"]),
                hazard_class=clean_text(row["Класс опасности"]),
                category=clean_text(row["Категория"]),
                efficiency_class=clean_text(row["Класс эффективности"]),
                funding_source=clean_text(row["Источник финансирования"]),
                expertise_place=clean_text(row["Место проведения экспертизы"]),
                full_set_cost=clean_text(row["Стоимость полного комплекта в тенге, без НДС"]),
                external_id=external_id if isinstance(external_id, int) else None,
            )
        )
    return rows


@dataclass(frozen=True, slots=True)
class BookRiskRow:
    """Строка витрины риска книги — эталон для сверки.

    Хранится отдельно от предметных данных: это не источник, а контрольное
    значение. Собственный расчёт обязан его воспроизводить, и расхождение
    должно оставаться видимым, а не исправляться подгонкой.
    """

    key: str
    indicator_values: dict[str, float | None]
    raw_score: float
    available_weight: float
    normalized_score: float
    significance_k: float
    score: float
    level: str
    explanation: str


def _read_book_risk_rows(
    path: Path, sheet: str, key_column: str, codes: Sequence[str]
) -> list[BookRiskRow]:
    rows: list[BookRiskRow] = []
    for _index, row in _sheet_rows(path, sheet, RISK_HEADER_ROW):
        values: dict[str, float | None] = {}
        for code in codes:
            cell = row[code]
            values[code] = None if cell is None else float(cell)  # type: ignore[arg-type]
        rows.append(
            BookRiskRow(
                key=str(row[key_column]).strip(),
                indicator_values=values,
                raw_score=float(row["S_raw"]),  # type: ignore[arg-type]
                available_weight=float(row["W_avail"]),  # type: ignore[arg-type]
                normalized_score=float(row["S_norm"]),  # type: ignore[arg-type]
                significance_k=float(row["K"]),  # type: ignore[arg-type]
                score=float(row["Risk Score"]),  # type: ignore[arg-type]
                level=str(row["Уровень"]),
                explanation=str(row["Расшифровка факторов"]),
            )
        )
    return rows


PPP_CODES: tuple[str, ...] = ("A1", "A2", "A3", "A4", "A5", "A6", "A7")
EXPERTISE_CODES: tuple[str, ...] = ("B1", "B2", "B3", "B4", "B5", "B6")


def read_ppp_book_rows(path: Path) -> list[BookRiskRow]:
    """Витрина риска типа A — контрольные значения книги."""
    return _read_book_risk_rows(path, SHEET_PPP_RISK, "№", PPP_CODES)


def read_expertise_book_rows(path: Path) -> list[BookRiskRow]:
    """Витрина риска типа B. Ключ хранится со срезанными нулями — как в книге."""
    return _read_book_risk_rows(path, SHEET_EXPERTISE_RISK, "Рег. №", EXPERTISE_CODES)


@dataclass(frozen=True, slots=True)
class ContestRow:
    """Конкурс ГЧП. Организатор записан как «БИН Наименование» у всех 514 строк."""

    contest_number: str
    organizer_raw: str
    government_partner_raw: str
    contest_name: str
    method: str
    status: str
    url: str

    @property
    def organizer_bin(self) -> str | None:
        """БИН из префикса наименования организатора.

        Единственное место во всей книге, где БИН доступен систематически.
        """
        head = self.organizer_raw.strip().split(maxsplit=1)
        if head and head[0].isdigit() and len(head[0]) == 12:
            return head[0]
        return None


@dataclass(frozen=True, slots=True)
class ContractRow:
    """Договор ГЧП. Всего 12 строк — для расчёта риска объём непригоден."""

    contract_number: str
    contest_number: str
    contract_type: str
    status: str
    selection_method: str
    financial_year: int | None
    amount: float | None
    government_partner: str
    private_partner: str


def read_contests(path: Path) -> list[ContestRow]:
    return [
        ContestRow(
            contest_number=clean_text(row["contest_number"]),
            organizer_raw=clean_text(row["organizer"]),
            government_partner_raw=clean_text(row["government_partner"]),
            contest_name=clean_text(row["contest_name"]),
            method=clean_text(row["method"]),
            status=clean_text(row["status"]),
            url=clean_text(row["url"]),
        )
        for _index, row in _sheet_rows(path, SHEET_CONTESTS, CONTESTS_HEADER_ROW)
    ]


def read_contracts(path: Path) -> list[ContractRow]:
    rows: list[ContractRow] = []
    for _index, row in _sheet_rows(path, SHEET_CONTRACTS, CONTRACTS_HEADER_ROW):
        year = row["Финансовый год"]
        rows.append(
            ContractRow(
                contract_number=clean_text(row["Номер договора"]),
                contest_number=clean_text(row["Номер конкурса"]),
                contract_type=clean_text(row["Тип договора"]),
                status=clean_text(row["Статус договора"]),
                selection_method=clean_text(row["Способ определения частного партнера"]),
                financial_year=year if isinstance(year, int) else None,
                amount=parse_number(row["Сумма договора"]),
                government_partner=clean_text(row["Наименование государственного партнера"]),
                private_partner=clean_text(row["Наименование потенциального частного партнера"]),
            )
        )
    return rows


def link_contracts_to_contests(
    contracts: Sequence[ContractRow], contests: Sequence[ContestRow]
) -> tuple[dict[str, str], list[str]]:
    """Связать договоры с конкурсами по номеру конкурса.

    Единственная работающая связка во всей книге, и она целиком внутри контура
    ГЧП: 12 договоров из 12 находят свой конкурс. Наружу, к заключениям
    экспертизы, она не ведёт — общего ключа с ними нет.

    Возвращает пару «связанное» и «номера договоров без конкурса».
    """
    known = {contest.contest_number for contest in contests}
    linked: dict[str, str] = {}
    orphans: list[str] = []
    for contract in contracts:
        if contract.contest_number in known:
            linked[contract.contract_number] = contract.contest_number
        else:
            orphans.append(contract.contract_number)
    return linked, orphans


def read_source_dir(source_dir: Path) -> Path:
    """Найти книгу слоя 8.6 в каталоге источников с учётом NFD в имени."""
    return resolve_source(source_dir, SOURCE_FILE_NAME)


# --- Популяции: индикаторы, зависящие от всей выборки ------------------------


class PppPopulation:
    """Выборка проектов ГЧП целиком.

    Индикаторы концентрации по определению считаются не по одной строке, а по
    всей выборке: доля расторгнутых у партнёра, доля партнёра в регионе, доля
    топ-1 партнёра у госпартнёра. Поэтому они живут здесь, а не в функции
    «посчитай строку» — иначе пришлось бы каждый раз незаметно пересчитывать
    агрегаты и получать разный результат при разной нарезке данных.
    """

    def __init__(
        self, rows: Sequence[PppProjectRow], *, as_of: dt.date = BOOK_CALCULATION_DATE
    ) -> None:
        self.rows = tuple(rows)
        self.as_of = as_of

        self._partner_projects: dict[str, list[PppProjectRow]] = defaultdict(list)
        self._region_totals: Counter[str] = Counter()
        self._region_partner: Counter[tuple[str, str]] = Counter()
        self._gov_partner_counts: dict[str, Counter[str]] = defaultdict(Counter)

        for row in self.rows:
            partner = row.private_partner_key
            region = row.region_raw
            self._region_totals[region] += 1
            if partner:
                self._partner_projects[partner].append(row)
                self._region_partner[(region, partner)] += 1
                self._gov_partner_counts[row.government_partner_key][partner] += 1

        costs = [row.cost_initial for row in self.rows if row.cost_initial is not None]
        # Квантиль с линейной интерполяцией по включающему методу: именно он
        # воспроизводит коэффициент значимости книги на всех 1323 строках.
        # Исключающий метод даёт другую границу и расходится на проекте № 889.
        self.cost_top_quartile = (
            statistics.quantiles(costs, n=4, method="inclusive")[2] if len(costs) > 1 else None
        )

    def a1(self, row: PppProjectRow) -> IndicatorValue:
        return IndicatorValue(
            code="A1", value=1.0 if row.is_terminated else 0.0, raw_value=row.status_raw
        )

    def a2(self, row: PppProjectRow) -> IndicatorValue:
        partner = row.private_partner_key
        if not partner:
            return IndicatorValue(code="A2", value=None, note="частный партнёр не указан")
        projects = self._partner_projects[partner]
        if len(projects) < 3:
            # Меньше трёх проектов — доля неустойчива, и методика приравнивает
            # её к нулю, а не объявляет индикатор неизмеренным: партнёр известен.
            return IndicatorValue(code="A2", value=0.0, raw_value=len(projects))
        share = sum(1 for project in projects if project.is_terminated) / len(projects)
        return IndicatorValue(code="A2", value=graded(share, A2_SCALE), raw_value=round(share, 4))

    def a3(self, row: PppProjectRow) -> IndicatorValue:
        partner = row.private_partner_key
        if not partner:
            return IndicatorValue(code="A3", value=None, note="частный партнёр не указан")
        total = self._region_totals[row.region_raw]
        share = self._region_partner[(row.region_raw, partner)] / total if total else 0.0
        return IndicatorValue(code="A3", value=graded(share, A3_SCALE), raw_value=round(share, 4))

    def a4(self, row: PppProjectRow) -> IndicatorValue:
        counts = self._gov_partner_counts.get(row.government_partner_key)
        total = sum(counts.values()) if counts else 0
        if total < 3:
            return IndicatorValue(
                code="A4",
                value=None,
                note="у государственного партнёра меньше трёх проектов с известным партнёром",
            )
        share = counts.most_common(1)[0][1] / total if counts else 0.0
        return IndicatorValue(code="A4", value=graded(share, A4_SCALE), raw_value=round(share, 4))

    def a5(self, row: PppProjectRow) -> IndicatorValue:
        end = row.construction_end
        if end is None:
            return IndicatorValue(
                code="A5", value=None, note="плановое окончание строительства не разобрано"
            )
        if row.has_date_error:
            return IndicatorValue(
                code="A5", value=None, note="окончание строительства раньше начала"
            )
        if OPERATION_STATUS_MARKER in row.status_raw.casefold():
            return IndicatorValue(code="A5", value=0.0, raw_value=row.status_raw)
        overdue_days = (self.as_of - end).days
        if overdue_days > A5_OVERDUE_DAYS_CRITICAL:
            value = 1.0
        elif overdue_days > 0:
            value = 0.5
        else:
            value = 0.0
        return IndicatorValue(code="A5", value=value, raw_value=overdue_days)

    def a6(self, row: PppProjectRow) -> IndicatorValue:
        cost, investments = row.cost_initial, row.investments
        if cost is None or investments is None:
            return IndicatorValue(
                code="A6", value=None, note="стоимость или объём инвестиций не заполнены"
            )
        ratio = investments / cost
        return IndicatorValue(code="A6", value=graded(ratio, A6_SCALE), raw_value=round(ratio, 4))

    def a7(self, row: PppProjectRow) -> IndicatorValue:
        initiative = row.initiative_kind.casefold()
        if "прямые переговоры" in initiative:
            value = 1.0
        elif "чфи" in initiative:
            value = 0.5
        else:
            value = 0.0
        return IndicatorValue(code="A7", value=value, raw_value=row.initiative_kind)

    def indicator_values(self, row: PppProjectRow) -> dict[str, IndicatorValue]:
        return {
            "A1": self.a1(row),
            "A2": self.a2(row),
            "A3": self.a3(row),
            "A4": self.a4(row),
            "A5": self.a5(row),
            "A6": self.a6(row),
            "A7": self.a7(row),
        }

    def is_top_quartile_cost(self, row: PppProjectRow) -> bool:
        return (
            self.cost_top_quartile is not None
            and row.cost_initial is not None
            and row.cost_initial >= self.cost_top_quartile
        )

    def significance_k(self, row: PppProjectRow) -> float:
        return ppp_significance_k(
            top_quartile_cost=self.is_top_quartile_cost(row),
            republican_level=row.is_republican,
        )


class ExpertisePopulation:
    """Выборка заключений экспертизы целиком.

    Здесь же лежит ответ на ловушку «111 объектов с повторной экспертизой».
    Строк с признаком повторной экспертизы действительно 111, но различных
    объектов за ними — 52. Метод :meth:`distinct_objects_with_repeated_expertise`
    считает объекты, а не строки; считать их по числу строк значит завысить
    результат вдвое.
    """

    def __init__(self, rows: Sequence[ExpertiseConclusionRow]) -> None:
        self.rows = tuple(rows)

        self._object_counts: Counter[tuple[str, str]] = Counter()
        self._customer_designers: dict[str, Counter[str]] = defaultdict(Counter)
        self._customer_totals: Counter[str] = Counter()
        self._customer_corrections: Counter[str] = Counter()

        for row in self.rows:
            self._object_counts[row.object_identity_key] += 1
            customer = row.customer_key
            self._customer_totals[customer] += 1
            if row.has_correction:
                self._customer_corrections[customer] += 1
            if customer.strip() and row.designer_key.strip():
                self._customer_designers[customer][row.designer_key] += 1

    def distinct_objects(self) -> int:
        """Число различных физических объектов за 4842 заключениями."""
        return len(self._object_counts)

    def distinct_objects_with_repeated_expertise(self) -> int:
        return sum(1 for count in self._object_counts.values() if count >= 2)

    def rows_with_repeated_expertise(self) -> int:
        """Число строк, а не объектов. Именно эти две величины и путают."""
        return sum(count for count in self._object_counts.values() if count >= 2)

    def b1(self, row: ExpertiseConclusionRow) -> IndicatorValue:
        return IndicatorValue(
            code="B1", value=1.0 if row.has_correction else 0.0, raw_value=row.has_correction
        )

    def b2(self, row: ExpertiseConclusionRow) -> IndicatorValue:
        count = self._object_counts[row.object_identity_key]
        return IndicatorValue(code="B2", value=graded(float(count), B2_SCALE), raw_value=count)

    def b3(self, row: ExpertiseConclusionRow) -> IndicatorValue:
        status = row.author_supervision_status.casefold()
        if "отсутствует согласование" in status:
            value = B3_ABSENT_VALUE
        elif "не согласован" in status:
            value = B3_NOT_AGREED_VALUE
        else:
            value = 0.0
        return IndicatorValue(code="B3", value=value, raw_value=row.author_supervision_status)

    def b4(self, row: ExpertiseConclusionRow) -> IndicatorValue:
        if row.has_cost_estimate is None:
            return IndicatorValue(
                code="B4", value=None, note="нет сведений о наличии сметной документации"
            )
        return IndicatorValue(
            code="B4",
            value=0.0 if row.has_cost_estimate else 1.0,
            raw_value=row.has_cost_estimate,
        )

    def b5(self, row: ExpertiseConclusionRow) -> IndicatorValue:
        counts = self._customer_designers.get(row.customer_key)
        total = sum(counts.values()) if counts else 0
        if total < 3:
            return IndicatorValue(
                code="B5", value=None, note="у заказчика меньше трёх объектов с проектировщиком"
            )
        share = counts.most_common(1)[0][1] / total if counts else 0.0
        return IndicatorValue(code="B5", value=graded(share, B5_SCALE), raw_value=round(share, 4))

    def b6(self, row: ExpertiseConclusionRow) -> IndicatorValue:
        customer = row.customer_key
        total = self._customer_totals[customer]
        if total < 5:
            return IndicatorValue(code="B6", value=None, note="у заказчика меньше пяти объектов")
        share = self._customer_corrections[customer] / total
        return IndicatorValue(code="B6", value=graded(share, B6_SCALE), raw_value=round(share, 4))

    def indicator_values(self, row: ExpertiseConclusionRow) -> dict[str, IndicatorValue]:
        return {
            "B1": self.b1(row),
            "B2": self.b2(row),
            "B3": self.b3(row),
            "B4": self.b4(row),
            "B5": self.b5(row),
            "B6": self.b6(row),
        }

    def significance_k(self, row: ExpertiseConclusionRow) -> float:
        return expertise_significance_k(
            hazard_class_1_2=row.is_high_hazard,
            responsibility_level_1=row.is_first_responsibility_level,
        )


__all__ = [
    "BOOK_CALCULATION_DATE",
    "EXPECTED_CONTEST_COUNT",
    "EXPECTED_CONTRACT_COUNT",
    "EXPECTED_EXPERTISE_COUNT",
    "EXPECTED_PPP_COUNT",
    "EXPERTISE_CODES",
    "PPP_CODES",
    "REGISTRATION_NUMBER_LENGTH",
    "SHEET_EXPERTISE_RAW",
    "SHEET_EXPERTISE_RISK",
    "SHEET_PPP_RAW",
    "SHEET_PPP_RISK",
    "SOURCE_FILE_NAME",
    "BookRiskRow",
    "ContestRow",
    "ContractRow",
    "ExpertiseConclusionRow",
    "ExpertisePopulation",
    "PppPopulation",
    "PppProjectRow",
    "as_stored",
    "clean_text",
    "is_terminated",
    "link_contracts_to_contests",
    "normalize_registration_number",
    "parse_number",
    "parse_source_date",
    "read_contests",
    "read_contracts",
    "read_expertise_book_rows",
    "read_expertise_conclusions",
    "read_ppp_book_rows",
    "read_ppp_projects",
    "read_source_dir",
    "squash",
]
