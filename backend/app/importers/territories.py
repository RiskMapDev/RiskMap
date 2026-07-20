"""Загрузка справочника территорий: границы, иерархия, алиасы, население.

Этот импортёр наполняет фундамент всей системы. Всё остальное — бюджет,
закупки, субсидии — привязывается к территории, и если справочник собран
небрежно, ошибка расходится по каждому слою. Поэтому здесь несколько решений
приняты жёстче, чем требует схема.

**Границы не грузятся без лицензии.** `BoundaryVersion` заполняется из
`data/boundaries/PROVENANCE.md`: источник, дата выгрузки, SHA-256, лицензия и
дословный текст атрибуции. Хеш файла сверяется с зафиксированным в PROVENANCE
при каждом запуске: если файл изменился, происхождение больше ничего не
доказывает, и импорт останавливается, а не грузит «похожие» данные.

**Наборов границ два, потому что файлов два.** У `kazakhstan-regions-osm.geojson`
и `almaty-oblast-osm.geojson` разные SHA-256 и разное время выгрузки. Свести их
в одну версию значило бы записать в поле `sha256` хеш одного файла и выдать его
за хеш обоих. Отношение `relation/215718` (Алматинская область) присутствует в
обоих файлах — оно загружается один раз, из набора регионов, а повтор
фиксируется замечанием, а не вторым полигоном.

**Прочерк в книге населения превращается в ноль только под расчёт.** Аудит
(`docs/audit/05-naselenie.md`) утверждает, что «-» означает отсутствие категории.
Импортёр это не принимает на веру, а проверяет арифметикой самой строки: если
«город + село = всё население» сходится только при нуле, значит источник
действительно фиксирует отсутствие. Если не сходится — это пропуск измерения,
и в базу идёт NULL с замечанием. Разница принципиальна: ноль участвует в
суммах и в расчёте долей, NULL — нет.

**Названия не исправляются.** Четыре области книги 8.3 написаны с опечатками
и вариантами транслитерации. Они заводятся алиасами `SOURCE_SPELLING` ровно так,
как написаны в источнике.
"""

from __future__ import annotations

import json
import uuid
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session

from app.db.base import utcnow
from app.db.models.source import (
    DataQualityIssue,
    ImportJob,
    ImportStatus,
    IssueSeverity,
    SourceDataset,
    SourceFile,
)
from app.db.models.territory import (
    AliasKind,
    BoundaryVersion,
    PopulationStat,
    Territory,
    TerritoryAlias,
    TerritoryLevel,
)
from app.services.territory_resolver import TerritoryResolver, normalize_territory_name

IMPORTER_NAME: Final[str] = "territories"

# --- Лицензия --------------------------------------------------------------
# Дословно из PROVENANCE.md, § «Лицензия и обязательная атрибуция». Текст
# атрибуции обязан показываться рядом с картой, поэтому он хранится как данные,
# а не собирается на лету в шаблоне: иначе его однажды забудут вывести.
LICENSE_NAME: Final[str] = "Open Database License (ODbL) v1.0"
LICENSE_URL: Final[str] = "https://opendatacommons.org/licenses/odbl/1-0/"
ATTRIBUTION_TEXT: Final[str] = "© OpenStreetMap contributors, ODbL 1.0"
SOURCE_NAME: Final[str] = "OpenStreetMap contributors, via Overpass API"
SOURCE_URL: Final[str] = "https://overpass-api.de/api/interpreter"

# --- Допуски упрощения -----------------------------------------------------
# В градусах, потому что геометрия хранится в EPSG:4326. На широте Казахстана
# (~43–55°) 0.001° ≈ 70–110 м, 0.01° ≈ 0,7–1,1 км. Первый допуск рассчитан на
# средние масштабы (район целиком в экране), второй — на обзорную карту страны,
# где деталь мельче километра всё равно не различима.
SIMPLIFY_TOLERANCE_MID: Final[float] = 0.001
SIMPLIFY_TOLERANCE_LOW: Final[float] = 0.01

# Порог расхождения заявленной и вычисленной площади, при котором пишется
# замечание. 5 % — граница, за которой расхождение уже нельзя списать на
# генерализацию контура (см. PROVENANCE.md, § 6).
AREA_MISMATCH_THRESHOLD: Final[float] = 0.05

POPULATION_FILE_NAME: Final[str] = "Численность_населения_Алматинской_области.xlsx"
POPULATION_SHEET: Final[str] = "Sheet1"
POPULATION_AS_OF: Final[date] = date(2026, 4, 1)
POPULATION_FIRST_ROW: Final[int] = 7
POPULATION_LAST_ROW: Final[int] = 31

# Прочерк в книге приходит строкой. Три варианта тире — потому что при
# выгрузке из Excel дефис легко превращается в тире, и различать их незачем.
DASH_VALUES: Final[frozenset[str]] = frozenset({"-", "–", "—"})

EXPECTED_REGIONS: Final[int] = 20
EXPECTED_ALMATY_UNITS: Final[int] = 11
EXPECTED_POPULATION_ROWS: Final[int] = 12
"""11 единиц второго уровня + итоговая строка области."""


# --- Описание наборов границ ------------------------------------------------


@dataclass(frozen=True, slots=True)
class BoundarySetSpec:
    """Паспорт файла границ, переписанный из PROVENANCE.md.

    Значения зашиты в код намеренно: PROVENANCE.md — документ для человека, и
    парсить его регулярками означало бы поставить юридически значимые поля в
    зависимость от вёрстки таблицы. Хеш при этом проверяется по файлу.
    """

    code: str
    title: str
    file_name: str
    sha256: str
    downloaded_at: date
    osm_data_as_of: date
    administrative_division_as_of: date
    notes: str


REGIONS_SET: Final[BoundarySetSpec] = BoundarySetSpec(
    code="osm-kz-regions-2026-07-20",
    title="Регионы Республики Казахстан, OSM admin_level=4",
    file_name="kazakhstan-regions-osm.geojson",
    sha256="c2de3b742b392c516690587be2b5a3e576d0391ec5a9b973e6d9d436ca5d2e09",
    downloaded_at=date(2026, 7, 20),
    osm_data_as_of=date(2026, 7, 20),
    # Последний акт, который набор заведомо отражает: Указ № 887 от 03.05.2022
    # (в силе с 08.06.2022) — в наборе присутствуют Абайская, Жетысуская и
    # Улытауская области. Более поздних изменений верхнего уровня не было.
    administrative_division_as_of=date(2022, 6, 8),
    notes=(
        "20 объектов: 17 областей + 3 города республиканского значения. "
        "Площади включают казахстанский сектор Каспия — границы по береговой "
        "линии не обрезались. Тег kato есть только у 2 из 20 регионов."
    ),
)

ALMATY_SET: Final[BoundarySetSpec] = BoundarySetSpec(
    code="osm-almaty-oblast-2026-07-20",
    title="Алматинская область и её единицы второго уровня, OSM admin_level=6",
    file_name="almaty-oblast-osm.geojson",
    sha256="02ed8cae0da93e427613d1b9a24542c2fdcd4a941acbb8f9fa1d9d8243a210bb",
    downloaded_at=date(2026, 7, 20),
    osm_data_as_of=date(2026, 7, 20),
    # Последнее изменение состава области — выделение города Алатау,
    # в силе с 09.01.2024 (docs/audit/04-geodannye.md, § 5.0).
    administrative_division_as_of=date(2024, 1, 9),
    notes=(
        "12 объектов, из которых загружаются 11 единиц второго уровня: "
        "relation/215718 (сама область) берётся из набора регионов, чтобы не "
        "дублировать полигон. КАТО отсутствует у всех 11 единиц."
    ),
)

ALIASES_8_3_FILE: Final[str] = "region-aliases-8-3.json"
ALIASES_8_3_SHA256: Final[str] = (
    "253c246cb76c89ad2f1481b3a67b6c572ea3a421e1790f73b63cacee82b5856b"
)

# --- Коды территорий --------------------------------------------------------
# Внутренний код — наш, а не источника, поэтому он задан таблицей, а не выведен
# транслитерацией: автоматический слаг молча меняется при правке названия в OSM,
# и тогда повторный импорт создаёт вторую территорию вместо обновления первой.

COUNTRY_CODE: Final[str] = "kz"

REGION_CODES: Final[dict[str, str]] = {
    "KZ-10": "abay-oblast",
    "KZ-11": "akmola-oblast",
    "KZ-15": "aktobe-oblast",
    "KZ-19": "almaty-oblast",
    "KZ-23": "atyrau-oblast",
    "KZ-27": "west-kazakhstan-oblast",
    "KZ-31": "zhambyl-oblast",
    "KZ-33": "zhetysu-oblast",
    "KZ-35": "karaganda-oblast",
    "KZ-39": "kostanay-oblast",
    "KZ-43": "kyzylorda-oblast",
    "KZ-47": "mangistau-oblast",
    "KZ-55": "pavlodar-oblast",
    "KZ-59": "north-kazakhstan-oblast",
    "KZ-61": "turkestan-oblast",
    "KZ-62": "ulytau-oblast",
    "KZ-63": "east-kazakhstan-oblast",
    "KZ-71": "astana-city",
    "KZ-75": "almaty-city",
    "KZ-79": "shymkent-city",
}

ALMATY_UNIT_CODES: Final[dict[int, str]] = {
    5517787: "zhambylskiy",
    5517810: "balkhashskiy",
    5517935: "raiymbekskiy",
    5517948: "uygurskiy",
    5503599: "talgarskiy",
    5518065: "konaev-city",
    5518093: "enbekshikazakhskiy",
    5518195: "iliyskiy",
    5518210: "karasayskiy",
    9159398: "kegenskiy",
    17012094: "alatau-city",
}

# Конаев и Алатау — города областного значения, а не районы. Уровень задан явно
# по нормативному статусу единицы; в OSM у всех одиннадцати admin_level=6, и
# различить их по геоданным невозможно.
ALMATY_CITY_CODES: Final[frozenset[str]] = frozenset({"konaev-city", "alatau-city"})

# Площади, заявленные ведомственным документом «Административно-территориальный
# слой» (см. PROVENANCE.md, § 6). Хранятся отдельно от вычисленных: расхождение
# между ними — самостоятельный факт, а не повод переписать одну из величин.
DOCUMENT_AREA_KM2: Final[dict[str, float]] = {
    "almaty-oblast": 105263.0,
    "balkhashskiy": 37400.0,
    "zhambylskiy": 19300.0,
    "uygurskiy": 8800.0,
    "enbekshikazakhskiy": 8300.0,
    "iliyskiy": 7800.0,
    "raiymbekskiy": 7100.0,
    "kegenskiy": 7100.0,
    "talgarskiy": 3600.0,
    "konaev-city": 3200.0,
    "karasayskiy": 2100.0,
    "alatau-city": 600.0,
}

# По этим двум единицам документ даёт площадь со знаком «~».
APPROXIMATE_AREA_CODES: Final[frozenset[str]] = frozenset({"konaev-city", "alatau-city"})

# Даты вступления в силу — только там, где нормативный акт назван в аудите с
# явной датой введения в действие. Остальным `valid_from` остаётся NULL:
# «единица существует давно» — не дата.
VALID_FROM: Final[dict[str, date]] = {
    "abay-oblast": date(2022, 6, 8),
    "zhetysu-oblast": date(2022, 6, 8),
    "ulytau-oblast": date(2022, 6, 8),
    "alatau-city": date(2024, 1, 9),
}

VALID_FROM_NOTES: Final[dict[str, str]] = {
    "abay-oblast": "Создана Указом Президента РК № 887 от 03.05.2022, в силе с 08.06.2022.",
    "zhetysu-oblast": "Создана Указом Президента РК № 887 от 03.05.2022, в силе с 08.06.2022.",
    "ulytau-oblast": "Создана Указом Президента РК № 887 от 03.05.2022, в силе с 08.06.2022.",
    "alatau-city": (
        "Город областного значения выделен из с. Жетыген Илийского района, "
        "в силе с 09.01.2024. Границы уточняются: контур OSM на 47 % больше "
        "ориентировочной площади документа."
    ),
    "kegenskiy": "Выделен из Райымбекского района Указом № 653 от 31.03.2018.",
}


# --- Разбор GeoJSON ---------------------------------------------------------


@dataclass(frozen=True, slots=True)
class TerritoryFeature:
    """Одна административная единица, вычитанная из GeoJSON."""

    code: str
    level: TerritoryLevel
    parent_code: str | None
    name_ru: str
    name_kk: str | None
    name_en: str | None
    name_osm: str | None
    int_name: str | None
    old_names: tuple[str, ...]
    iso3166_2: str | None
    kato_code: str | None
    osm_relation_id: int
    osm_ref: str
    area_km2_osm: float | None
    geometry: dict[str, Any]
    source_row_ref: str


def _clean(value: object) -> str | None:
    """Строка или None. Пустая строка — это отсутствие значения, а не значение."""
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def _read_feature_collection(path: Path) -> list[dict[str, Any]]:
    payload: dict[str, Any] = json.loads(path.read_text(encoding="utf-8"))
    features: list[dict[str, Any]] = payload["features"]
    return features


def _old_names(properties: dict[str, Any]) -> tuple[str, ...]:
    """Прежние названия из свойств верхнего уровня.

    Берутся только `old_name` и `old_name_ru`: в `osm_tags` прежние названия
    записаны ключами с диапазонами дат (`old_name:ru:1993-2023`), и разбирать
    их по шаблону — значит гадать о периоде действия названия.
    """
    names = [_clean(properties.get("old_name")), _clean(properties.get("old_name_ru"))]
    return tuple(dict.fromkeys(name for name in names if name))


def parse_regions(path: Path) -> list[TerritoryFeature]:
    """Прочитать 20 регионов верхнего уровня."""
    features: list[TerritoryFeature] = []
    for index, raw in enumerate(_read_feature_collection(path)):
        properties: dict[str, Any] = raw["properties"]
        iso = _clean(properties.get("iso3166_2"))
        if iso is None or iso not in REGION_CODES:
            raise ValueError(
                f"{path.name}#{index}: регион без известного ISO 3166-2 ({iso!r}). "
                "Таблицу кодов нужно пополнить осознанно, а не достраивать импортом."
            )
        features.append(
            TerritoryFeature(
                code=REGION_CODES[iso],
                level=TerritoryLevel.REGION,
                parent_code=COUNTRY_CODE,
                name_ru=str(properties["name_ru"]),
                name_kk=_clean(properties.get("name_kk")),
                name_en=_clean(properties.get("name_en")),
                name_osm=_clean(properties.get("name")),
                int_name=_clean((properties.get("osm_tags") or {}).get("int_name")),
                old_names=_old_names(properties),
                iso3166_2=iso,
                kato_code=_clean(properties.get("kato")),
                osm_relation_id=int(properties["osm_id"]),
                osm_ref=str(properties["osm_ref"]),
                area_km2_osm=properties.get("area_km2_wgs84_geodesic"),
                geometry=raw["geometry"],
                source_row_ref=f"{path.name}#features[{index}]",
            )
        )
    return features


def parse_almaty_units(path: Path) -> list[TerritoryFeature]:
    """Прочитать 11 единиц второго уровня Алматинской области.

    Сама область (`admin_level=4`) в файле есть, но пропускается: тот же объект
    OSM уже загружен из набора регионов, и второй его экземпляр стал бы
    дубликатом геометрии с другим кодом набора.
    """
    features: list[TerritoryFeature] = []
    for index, raw in enumerate(_read_feature_collection(path)):
        properties: dict[str, Any] = raw["properties"]
        if int(properties.get("admin_level", 0)) != 6:
            continue
        osm_id = int(properties["osm_id"])
        if osm_id not in ALMATY_UNIT_CODES:
            raise ValueError(
                f"{path.name}#{index}: единица relation/{osm_id} отсутствует в таблице кодов."
            )
        code = ALMATY_UNIT_CODES[osm_id]
        features.append(
            TerritoryFeature(
                code=code,
                level=(
                    TerritoryLevel.CITY
                    if code in ALMATY_CITY_CODES
                    else TerritoryLevel.DISTRICT
                ),
                parent_code=REGION_CODES["KZ-19"],
                name_ru=str(properties["name_ru"]),
                name_kk=_clean(properties.get("name_kk")),
                name_en=_clean(properties.get("name_en")),
                name_osm=_clean(properties.get("name")),
                int_name=_clean(properties.get("int_name")),
                old_names=_old_names(properties),
                iso3166_2=_clean(properties.get("iso3166_2")),
                # КАТО у районов нет ни в одном источнике. Пустое значение —
                # штатное состояние, достраивать код нельзя.
                kato_code=_clean(properties.get("kato")),
                osm_relation_id=osm_id,
                osm_ref=str(properties["osm_ref"]),
                area_km2_osm=properties.get("area_km2_wgs84_geodesic"),
                geometry=raw["geometry"],
                source_row_ref=f"{path.name}#features[{index}]",
            )
        )
    return features


# --- Разбор книги населения -------------------------------------------------


@dataclass(frozen=True, slots=True)
class PopulationRow:
    """Строка книги численности населения.

    `dash_decisions` хранит, что импортёр решил про каждый прочерк и на каком
    основании. Без этого выбор «0 или NULL» становится невидимым, а он влияет
    на все производные показатели.
    """

    excel_row: int
    raw_name: str
    kind: str
    """country | region | unit | center — уровень, восстановленный по тексту."""

    values: dict[str, int | None]
    dash_decisions: dict[str, str]
    center_name: str | None = None

    @property
    def source_row_ref(self) -> str:
        return f"{POPULATION_SHEET}!A{self.excel_row}"


POPULATION_COLUMNS: Final[tuple[str, ...]] = (
    "total",
    "male",
    "female",
    "urban_total",
    "urban_male",
    "urban_female",
    "rural_total",
    "rural_male",
    "rural_female",
)

# Тройки «часть, часть, целое» для проверки прочерков и контрольных сумм.
_URBAN_RURAL_TRIPLES: Final[tuple[tuple[str, str, str], ...]] = (
    ("urban_total", "rural_total", "total"),
    ("urban_male", "rural_male", "male"),
    ("urban_female", "rural_female", "female"),
)


def classify_population_row(raw_name: str) -> str:
    """Определить уровень территории по тексту названия.

    В книге уровень не размечен колонкой — он выводится только из текста.
    Проверка на `г.а.` идёт раньше проверки на префикс `г.`: «Қонаев г.а.» —
    городская администрация, единица второго уровня, а «г.Есик» — райцентр.
    """
    name = raw_name.strip()
    if name == "Республика Казахстан":
        return "country"
    if name.endswith("район") or name.endswith("г.а."):
        return "unit"
    if name.startswith(("город ", "г.", "с.", "пос.")):
        return "center"
    return "region"


def _cell_value(raw: object) -> tuple[int | None, bool]:
    """Значение ячейки: число или признак прочерка."""
    if raw is None:
        return None, False
    if isinstance(raw, int | float) and not isinstance(raw, bool):
        return int(raw), False
    text_value = str(raw).strip()
    if text_value in DASH_VALUES:
        return None, True
    if not text_value:
        return None, False
    return int(float(text_value.replace(" ", "").replace(",", "."))), False


def _resolve_dashes(
    values: dict[str, int | None], dashes: set[str]
) -> tuple[dict[str, int | None], dict[str, str]]:
    """Решить, что означает прочерк в каждой колонке.

    Аудит утверждает, что «-» — это отсутствие категории населения. Импортёр
    проверяет утверждение расчётом: если известны целое и вторая часть, то
    прочерк обязан равняться их разности. Разность ноль — источник
    действительно фиксирует отсутствие, ставим 0. Разность не ноль — за
    прочерком скрыто настоящее число, то есть это пропуск измерения, и в базу
    идёт NULL. Молча подставить ноль во втором случае значило бы занизить
    показатель и при этом сломать сверку сумм незаметно.
    """
    resolved = dict(values)
    decisions: dict[str, str] = {}

    for part_a, part_b, whole in _URBAN_RURAL_TRIPLES:
        for target, other in ((part_a, part_b), (part_b, part_a)):
            if target not in dashes:
                continue
            whole_value = resolved.get(whole)
            other_value = resolved.get(other)
            if whole_value is None or other_value is None or other in dashes:
                decisions[target] = "NULL: вторая часть или итог строки неизвестны"
                continue
            remainder = whole_value - other_value
            if remainder == 0:
                resolved[target] = 0
                decisions[target] = f"0: {whole} − {other} = 0, категория отсутствует"
            else:
                resolved[target] = None
                decisions[target] = (
                    f"NULL: {whole} − {other} = {remainder} ≠ 0, "
                    "прочерк скрывает измерение, а не отсутствие"
                )
    return resolved, decisions


def parse_population(path: Path) -> list[PopulationRow]:
    """Прочитать лист численности населения целиком, включая райцентры.

    Райцентры возвращаются наравне с остальными строками: их численность в
    справочник не грузится (территорий такого уровня в системе нет), но имя
    центра нужно единице второго уровня, а сама строка — контрольным суммам.
    """
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[POPULATION_SHEET]
        rows: list[PopulationRow] = []
        for excel_row, raw in enumerate(
            sheet.iter_rows(
                min_row=POPULATION_FIRST_ROW,
                max_row=POPULATION_LAST_ROW,
                values_only=True,
            ),
            start=POPULATION_FIRST_ROW,
        ):
            name = _clean(raw[0])
            if name is None:
                continue
            values: dict[str, int | None] = {}
            dashes: set[str] = set()
            for offset, column in enumerate(POPULATION_COLUMNS, start=1):
                value, is_dash = _cell_value(raw[offset])
                values[column] = value
                if is_dash:
                    dashes.add(column)
            resolved, decisions = _resolve_dashes(values, dashes)
            rows.append(
                PopulationRow(
                    excel_row=excel_row,
                    raw_name=name,
                    kind=classify_population_row(name),
                    values=resolved,
                    dash_decisions=decisions,
                )
            )
        return _attach_centers(rows)
    finally:
        workbook.close()


def _attach_centers(rows: list[PopulationRow]) -> list[PopulationRow]:
    """Приписать каждой единице её административный центр.

    Центр берётся из самой книги — строкой ниже единицы, — а не из внешнего
    перечня: так название центра остаётся свидетельством источника, а не
    сведением, добавленным импортёром от себя.
    """
    result: list[PopulationRow] = []
    for index, row in enumerate(rows):
        if row.kind != "unit":
            result.append(row)
            continue
        following = rows[index + 1] if index + 1 < len(rows) else None
        center = following.raw_name if following and following.kind == "center" else None
        result.append(
            PopulationRow(
                excel_row=row.excel_row,
                raw_name=row.raw_name,
                kind=row.kind,
                values=row.values,
                dash_decisions=row.dash_decisions,
                center_name=center,
            )
        )
    return result


def check_population_totals(rows: list[PopulationRow]) -> dict[str, Any]:
    """Свести контрольные суммы книги.

    Три независимых контроля: сумма 11 единиц против итога области, «мужчины +
    женщины = всё население» и «город + село = всё население» в каждой строке.
    Возвращается разбор, а не булев ответ: расхождение нужно видеть по
    показателям, иначе непонятно, где именно поехало.
    """
    units = [row for row in rows if row.kind == "unit"]
    oblast = next((row for row in rows if row.kind == "region"), None)

    sums = {
        column: sum(row.values[column] or 0 for row in units) for column in POPULATION_COLUMNS
    }
    oblast_values = dict(oblast.values) if oblast else {}
    per_column = {
        column: {
            "units_sum": sums[column],
            "oblast": oblast_values.get(column),
            "matches": sums[column] == oblast_values.get(column),
        }
        for column in POPULATION_COLUMNS
    }

    gender_mismatches: list[str] = []
    settlement_mismatches: list[str] = []
    for row in rows:
        male, female, total = row.values["male"], row.values["female"], row.values["total"]
        if (
            male is not None
            and female is not None
            and total is not None
            and male + female != total
        ):
            gender_mismatches.append(f"{row.source_row_ref} {row.raw_name}")
        for part_a, part_b, whole in _URBAN_RURAL_TRIPLES:
            a, b, w = row.values[part_a], row.values[part_b], row.values[whole]
            if a is not None and b is not None and w is not None and a + b != w:
                settlement_mismatches.append(f"{row.source_row_ref} {row.raw_name} ({whole})")

    return {
        "units_count": len(units),
        "per_column": per_column,
        "all_columns_match": all(item["matches"] for item in per_column.values()),
        "gender_mismatches": gender_mismatches,
        "settlement_mismatches": settlement_mismatches,
    }


# --- Алиасы -----------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class AliasCandidate:
    """Кандидат в алиасы до свёртки и отсева дублей."""

    alias: str
    kind: AliasKind
    source_layer: str | None = None


def short_form(name_ru: str, level: TerritoryLevel) -> str | None:
    """Короткое написание названия — «Талгарский р-н», «Алматинская обл.».

    Такие формы встречаются в документах повсеместно, поэтому они заводятся
    наравне с полными. Практически все они сворачиваются функцией нормализации
    к той же форме, что и официальное название, и отдельной строкой в таблице
    не сохранятся — см. `dedupe_aliases`.
    """
    if name_ru.endswith(" район"):
        return name_ru.removesuffix(" район") + " р-н"
    if name_ru.endswith(" область"):
        return name_ru.removesuffix(" область") + " обл."
    if level is TerritoryLevel.CITY:
        return f"г. {name_ru}"
    return None


def alias_candidates(feature: TerritoryFeature) -> list[AliasCandidate]:
    """Все написания названия территории, известные из геонабора."""
    candidates: list[AliasCandidate] = [
        AliasCandidate(feature.name_ru, AliasKind.OFFICIAL),
    ]
    if feature.name_kk:
        candidates.append(AliasCandidate(feature.name_kk, AliasKind.OFFICIAL))
    if feature.name_osm:
        candidates.append(AliasCandidate(feature.name_osm, AliasKind.OFFICIAL))
    short = short_form(feature.name_ru, feature.level)
    if short:
        candidates.append(AliasCandidate(short, AliasKind.SHORT))
    if feature.name_en:
        candidates.append(AliasCandidate(feature.name_en, AliasKind.TRANSLITERATION))
    if feature.int_name:
        candidates.append(AliasCandidate(feature.int_name, AliasKind.TRANSLITERATION))
    candidates.extend(
        AliasCandidate(name, AliasKind.HISTORICAL) for name in feature.old_names
    )
    return candidates


@dataclass(frozen=True, slots=True)
class AliasRow:
    """Алиас, готовый к записи."""

    alias: str
    normalized: str
    kind: AliasKind
    source_layer: str | None
    notes: str | None


def dedupe_aliases(candidates: list[AliasCandidate]) -> list[AliasRow]:
    """Свернуть кандидатов к строкам таблицы.

    Ограничение `uq_alias_normalized_territory` допускает одну строку на пару
    «свёрнутая форма + территория». Это не досадное ограничение, а следствие
    смысла таблицы: свёртка на то и нужна, чтобы «Талгарский район» и
    «Талгарский р-н» считались одним написанием. Поэтому лишние варианты не
    выбрасываются молча, а перечисляются в `notes` оставшейся строки — след
    того, что источник писал название иначе, должен сохраниться.
    """
    primary_by_form: dict[str, AliasCandidate] = {}
    folded_by_form: dict[str, list[AliasCandidate]] = {}
    for candidate in candidates:
        normalized = normalize_territory_name(candidate.alias)
        if not normalized:
            continue
        primary = primary_by_form.get(normalized)
        if primary is None:
            primary_by_form[normalized] = candidate
            folded_by_form[normalized] = []
            continue
        if candidate.alias != primary.alias and all(
            candidate.alias != item.alias for item in folded_by_form[normalized]
        ):
            folded_by_form[normalized].append(candidate)

    rows: list[AliasRow] = []
    for normalized, primary in primary_by_form.items():
        folded = folded_by_form[normalized]
        note = None
        if folded:
            listed = ", ".join(
                f"«{item.alias}»"
                + (f" ({item.source_layer})" if item.source_layer else f" ({item.kind})")
                for item in folded
            )
            note = f"К той же свёрнутой форме приведены написания: {listed}"
        rows.append(
            AliasRow(
                alias=primary.alias,
                normalized=normalized,
                kind=primary.kind,
                source_layer=primary.source_layer,
                notes=note,
            )
        )
    return rows


def book_alias_candidates(aliases_payload: dict[str, Any]) -> dict[str, list[AliasCandidate]]:
    """Написания регионов из книги 8.3, разложенные по кодам территорий.

    Написание источника всегда заводится как `SOURCE_SPELLING`, включая четыре
    расхождения с нормой («Западно-Казахстанкая», «Мангыстауская»,
    «Северо-Казахстанкая», «Туркистанская»). Данные книги при этом не
    исправляются: алиас — это способ связать неверное написание с территорией,
    а не разрешение переписать источник.
    """
    result: dict[str, list[AliasCandidate]] = {}
    for record in aliases_payload["regions"]:
        iso = record["iso3166_2"]
        code = REGION_CODES[iso]
        bucket = result.setdefault(code, [])
        seen: set[str] = set()
        for name in (
            record["book_name_source_spelling"],
            record["book_name_normalized_column"],
            *record.get("aliases", []),
        ):
            text_value = str(name).strip()
            if not text_value or text_value in seen:
                continue
            seen.add(text_value)
            bucket.append(
                AliasCandidate(text_value, AliasKind.SOURCE_SPELLING, source_layer="8.3")
            )
    return result


# --- Отчёт ------------------------------------------------------------------


@dataclass
class LoadReport:
    """Что сделал (или сделал бы) импорт."""

    dry_run: bool = False
    territories_by_level: dict[str, int] = field(default_factory=dict)
    territories_created: int = 0
    territories_updated: int = 0
    geometries_written: int = 0
    geometries_repaired: int = 0
    aliases_by_kind: dict[str, int] = field(default_factory=dict)
    aliases_folded: int = 0
    population_written: int = 0
    population_skipped: int = 0
    ambiguous_aliases: tuple[str, ...] = ()
    issues: list[tuple[str, str, str]] = field(default_factory=list)
    reconciliation: dict[str, Any] = field(default_factory=dict)

    @property
    def territories_total(self) -> int:
        return sum(self.territories_by_level.values())

    def summary_ru(self) -> str:
        levels = ", ".join(
            f"{level}: {count}" for level, count in self.territories_by_level.items()
        )
        kinds = ", ".join(f"{kind}: {count}" for kind, count in self.aliases_by_kind.items())
        prefix = "СУХОЙ ПРОГОН — ничего не записано\n" if self.dry_run else ""
        return (
            f"{prefix}"
            f"территорий {self.territories_total} ({levels}); "
            f"создано {self.territories_created}, обновлено {self.territories_updated}\n"
            f"геометрий {self.geometries_written}, из них починено {self.geometries_repaired}\n"
            f"алиасов {sum(self.aliases_by_kind.values())} ({kinds}); "
            f"свёрнуто дублей {self.aliases_folded}\n"
            f"строк населения {self.population_written}, пропущено {self.population_skipped}\n"
            f"неоднозначных написаний {len(self.ambiguous_aliases)}; "
            f"замечаний {len(self.issues)}"
        )


# --- SQL для геометрий ------------------------------------------------------

# Геометрия обрабатывается целиком в PostGIS одним запросом. Причина простая:
# считать площадь на эллипсоиде, упрощать с сохранением топологии и искать
# точку внутри полигона должен тот же движок, который потом эти геометрии
# отдаёт карте, иначе расчёт и хранение разойдутся на краевых случаях.
#
# Важно про `is_valid`: колонка описывает геометрию ИСХОДНУЮ, а не сохранённую.
# Сохраняется всегда исправленная — иначе её нельзя показать, — и если бы
# признак относился к ней, он был бы всегда true и ничего не значил. False плюс
# заполненный `validity_note` читается как «пришло битым, починено вот так»,
# и найти все починенные объекты можно одним условием.
_GEOMETRY_SQL = text(
    """
WITH raw AS (
    SELECT ST_SetSRID(ST_GeomFromGeoJSON(CAST(:geojson AS text)), 4326) AS g
),
checked AS (
    SELECT g, ST_IsValid(g) AS is_ok, ST_IsValidReason(g) AS reason FROM raw
),
fixed AS (
    SELECT
        ST_Multi(
            CASE WHEN is_ok THEN g ELSE ST_CollectionExtract(ST_MakeValid(g), 3) END
        ) AS g,
        is_ok,
        reason
    FROM checked
)
INSERT INTO territory_geometries AS tg (
    id, territory_id, geom, geom_simplified_mid, geom_simplified_low,
    centroid, is_valid, validity_note
)
SELECT
    CAST(:row_id AS uuid),
    CAST(:territory_id AS uuid),
    g,
    ST_Multi(ST_SimplifyPreserveTopology(g, :tolerance_mid)),
    ST_Multi(ST_SimplifyPreserveTopology(g, :tolerance_low)),
    ST_PointOnSurface(g),
    is_ok,
    CASE
        WHEN is_ok THEN NULL
        ELSE 'ST_IsValid = false, причина: ' || reason
             || '. Исправлено ST_MakeValid + ST_CollectionExtract(…, 3) + ST_Multi; '
             || 'исходная геометрия в файле-источнике не менялась.'
    END
FROM fixed
ON CONFLICT (territory_id) DO UPDATE SET
    geom = EXCLUDED.geom,
    geom_simplified_mid = EXCLUDED.geom_simplified_mid,
    geom_simplified_low = EXCLUDED.geom_simplified_low,
    centroid = EXCLUDED.centroid,
    is_valid = EXCLUDED.is_valid,
    validity_note = EXCLUDED.validity_note,
    updated_at = now()
RETURNING
    tg.is_valid,
    tg.validity_note,
    ST_Area(tg.geom::geography) / 1e6 AS area_km2,
    ST_Within(tg.centroid, tg.geom) AS centroid_inside,
    ST_IsEmpty(tg.geom) AS is_empty
"""
)

_UNION_AREA_SQL = text(
    """
SELECT ST_Area(ST_Union(g.geom)::geography) / 1e6
FROM territory_geometries g
JOIN territories t ON t.id = g.territory_id
WHERE t.parent_id = CAST(:parent_id AS uuid)
"""
)


# --- Загрузчик --------------------------------------------------------------


class TerritoryLoader:
    """Загрузка справочника территорий в базу.

    Весь импорт идёт одной транзакцией. При `dry_run` в конце выполняется
    откат: так «показать, что будет сделано» опирается на настоящие вставки и
    настоящие проверки БД, а не на их предсказание, но в базе не остаётся ничего.
    """

    def __init__(
        self,
        session: Session,
        *,
        data_dir: Path,
        source_dir: Path,
        dry_run: bool = False,
    ) -> None:
        self.session = session
        self.boundaries_dir = data_dir / "boundaries"
        self.source_dir = source_dir
        self.dry_run = dry_run
        self.report = LoadReport(dry_run=dry_run)
        self.job: ImportJob | None = None
        self._territories: dict[str, Territory] = {}
        self._alias_candidates: dict[str, list[AliasCandidate]] = {}

    # --- вспомогательное ---------------------------------------------------

    def _issue(
        self,
        severity: IssueSeverity,
        code: str,
        message: str,
        *,
        source_row_ref: str | None = None,
        column_name: str | None = None,
        raw_value: str | None = None,
        context: dict[str, Any] | None = None,
    ) -> None:
        assert self.job is not None
        self.session.add(
            DataQualityIssue(
                import_job=self.job,
                severity=severity,
                code=code,
                message=message,
                source_row_ref=source_row_ref,
                column_name=column_name,
                raw_value=raw_value,
                context=context,
            )
        )
        self.report.issues.append((str(severity), code, message))

    def _source_file(self, path: Path, *, origin: str, expected_sha256: str | None) -> SourceFile:
        """Зафиксировать файл-источник по хешу.

        Хеш считается заново при каждом запуске и сверяется с зафиксированным
        в PROVENANCE.md. Расхождение — отказ: происхождение, записанное для
        другого содержимого, ничего не доказывает про это.
        """
        from scripts.source_manifest import normalize_name, sha256_of

        digest = sha256_of(path)
        if expected_sha256 is not None and digest != expected_sha256:
            raise ValueError(
                f"{path.name}: SHA-256 файла {digest} не совпадает с зафиксированным "
                f"в PROVENANCE.md {expected_sha256}. Импорт остановлен: происхождение "
                "набора больше не подтверждено."
            )
        existing = self.session.scalars(
            select(SourceFile).where(SourceFile.sha256 == digest)
        ).one_or_none()
        if existing is not None:
            return existing
        stat = path.stat()
        source_file = SourceFile(
            file_name=path.name,
            normalized_name=normalize_name(path.name),
            sha256=digest,
            size_bytes=stat.st_size,
            origin=origin,
        )
        self.session.add(source_file)
        self.session.flush()
        return source_file

    def _dataset(
        self,
        source_file: SourceFile,
        *,
        sheet_name: str,
        role: str,
        row_count: int | None,
        data_as_of: date | None,
        header_row: int | None = None,
    ) -> SourceDataset:
        existing = self.session.scalars(
            select(SourceDataset).where(
                SourceDataset.source_file_id == source_file.id,
                SourceDataset.sheet_name == sheet_name,
            )
        ).one_or_none()
        dataset = existing or SourceDataset(
            source_file_id=source_file.id, sheet_name=sheet_name, role=role
        )
        dataset.role = role
        dataset.row_count = row_count
        dataset.data_as_of = data_as_of
        dataset.header_row = header_row
        if existing is None:
            self.session.add(dataset)
        self.session.flush()
        return dataset

    # --- шаги --------------------------------------------------------------

    def _start_job(self) -> None:
        self.job = ImportJob(
            importer=IMPORTER_NAME,
            status=ImportStatus.DRY_RUN if self.dry_run else ImportStatus.RUNNING,
            is_dry_run=self.dry_run,
            started_at=utcnow(),
        )
        self.session.add(self.job)
        self.session.flush()

    def _boundary_version(self, spec: BoundarySetSpec) -> BoundaryVersion:
        existing = self.session.scalars(
            select(BoundaryVersion).where(BoundaryVersion.code == spec.code)
        ).one_or_none()
        version = existing or BoundaryVersion(code=spec.code)
        version.title = spec.title
        version.source_name = SOURCE_NAME
        version.source_url = SOURCE_URL
        version.downloaded_at = spec.downloaded_at
        version.license_name = LICENSE_NAME
        version.license_url = LICENSE_URL
        version.attribution_text = ATTRIBUTION_TEXT
        # ODbL прямо разрешает распространение производной базы при условии
        # атрибуции и share-alike, поэтому признак выставлен, а не занулён
        # «на всякий случай»: запрет распространения закрыл бы выгрузку слоя.
        version.redistribution_allowed = True
        version.administrative_division_as_of = spec.administrative_division_as_of
        version.is_current = True
        version.sha256 = spec.sha256
        version.notes = spec.notes
        if existing is None:
            self.session.add(version)
        self.session.flush()
        return version

    def _upsert_territory(
        self,
        *,
        code: str,
        boundary_version: BoundaryVersion,
        dataset: SourceDataset | None,
        level: TerritoryLevel,
        parent: Territory | None,
        name_ru: str,
        name_kk: str | None = None,
        name_en: str | None = None,
        iso3166_2: str | None = None,
        kato_code: str | None = None,
        osm_relation_id: int | None = None,
        natural_key: str,
        source_row_ref: str | None,
        data_as_of: date | None,
        area_km2: Decimal | None = None,
        admin_center_name: str | None = None,
        valid_from: date | None = None,
        notes: str | None = None,
    ) -> Territory:
        assert self.job is not None
        existing = self.session.scalars(
            select(Territory).where(
                Territory.code == code,
                Territory.boundary_version_id == boundary_version.id,
            )
        ).one_or_none()
        territory = existing or Territory(code=code, boundary_version_id=boundary_version.id)
        territory.kato_code = kato_code
        territory.iso3166_2 = iso3166_2
        territory.osm_relation_id = osm_relation_id
        territory.name_ru = name_ru
        territory.name_kk = name_kk
        territory.name_en = name_en
        territory.level = level
        territory.parent_id = parent.id if parent else None
        territory.admin_center_name = admin_center_name
        territory.area_km2 = area_km2
        territory.valid_from = valid_from
        territory.notes = notes
        territory.source_dataset_id = dataset.id if dataset else None
        territory.import_job_id = self.job.id
        territory.source_row_ref = source_row_ref
        territory.natural_key = natural_key
        territory.imported_at = utcnow()
        territory.data_as_of = data_as_of
        if existing is None:
            self.session.add(territory)
            self.report.territories_created += 1
        else:
            self.report.territories_updated += 1
        self.session.flush()
        self._territories[code] = territory
        level_name = str(level)
        self.report.territories_by_level[level_name] = (
            self.report.territories_by_level.get(level_name, 0) + 1
        )
        return territory

    def _write_geometry(self, territory: Territory, feature: TerritoryFeature) -> None:
        row = self.session.execute(
            _GEOMETRY_SQL,
            {
                "geojson": json.dumps(feature.geometry),
                "row_id": str(uuid.uuid4()),
                "territory_id": str(territory.id),
                "tolerance_mid": SIMPLIFY_TOLERANCE_MID,
                "tolerance_low": SIMPLIFY_TOLERANCE_LOW,
            },
        ).one()
        is_valid = bool(row[0])
        validity_note = row[1]
        area_km2 = float(row[2])
        centroid_inside = bool(row[3])
        is_empty = bool(row[4])

        self.report.geometries_written += 1
        if is_empty:
            # ST_MakeValid способна вернуть линию или точку, если полигон
            # выродился. После ST_CollectionExtract(…, 3) от такой геометрии не
            # остаётся ничего, и territory окажется без границы — молча этого
            # допускать нельзя.
            self._issue(
                IssueSeverity.ERROR,
                "geometry_empty_after_repair",
                f"{territory.code}: после исправления геометрия пуста — полигон выродился.",
                source_row_ref=feature.source_row_ref,
            )
        if not is_valid or validity_note:
            self.report.geometries_repaired += 1
            self._issue(
                IssueSeverity.WARNING,
                "geometry_repaired",
                f"{territory.code}: геометрия была невалидна и исправлена. {validity_note}",
                source_row_ref=feature.source_row_ref,
                context={"territory": territory.code},
            )
        if not centroid_inside:
            # ST_PointOnSurface обязана вернуть точку внутри полигона. Если это
            # не так, подпись и зум карты уедут в пустое место.
            self._issue(
                IssueSeverity.ERROR,
                "centroid_outside_polygon",
                f"{territory.code}: точка подписи оказалась вне полигона.",
                source_row_ref=feature.source_row_ref,
            )

        territory.area_km2_computed = Decimal(f"{area_km2:.2f}")
        self._check_area(territory, feature, area_km2)
        self.session.flush()

    def _check_area(
        self, territory: Territory, feature: TerritoryFeature, area_km2: float
    ) -> None:
        """Сверить площадь с заявленной в документе и с расчётом выгрузки."""
        if feature.area_km2_osm is not None:
            drift = abs(area_km2 - feature.area_km2_osm)
            if drift > max(1.0, feature.area_km2_osm * 0.005):
                self._issue(
                    IssueSeverity.WARNING,
                    "area_differs_from_extract",
                    f"{territory.code}: площадь по PostGIS {area_km2:.1f} км² расходится "
                    f"с расчётом выгрузки {feature.area_km2_osm:.1f} км².",
                    source_row_ref=feature.source_row_ref,
                )
        declared = territory.area_km2
        if declared is None:
            return
        declared_value = float(declared)
        delta = (area_km2 - declared_value) / declared_value
        if abs(delta) >= AREA_MISMATCH_THRESHOLD:
            self._issue(
                IssueSeverity.WARNING,
                "area_mismatch_with_document",
                f"{territory.code}: площадь по границам OSM {area_km2:.0f} км², "
                f"по ведомственному документу {declared_value:.0f} км² "
                f"({delta:+.1%}). Обе величины сохранены раздельно.",
                source_row_ref=feature.source_row_ref,
                context={"computed_km2": round(area_km2, 1), "declared_km2": declared_value},
            )

    def _load_boundary_set(
        self,
        spec: BoundarySetSpec,
        features: list[TerritoryFeature],
        dataset: SourceDataset,
        version: BoundaryVersion,
    ) -> None:
        for feature in features:
            parent = self._territories.get(feature.parent_code) if feature.parent_code else None
            declared = DOCUMENT_AREA_KM2.get(feature.code)
            notes: list[str] = []
            if feature.code in VALID_FROM_NOTES:
                notes.append(VALID_FROM_NOTES[feature.code])
            if feature.code in APPROXIMATE_AREA_CODES:
                notes.append("Площадь по документу дана приблизительно (со знаком «~»).")
            if feature.kato_code is None:
                notes.append("Код КАТО в источниках отсутствует; поле оставлено пустым.")
            territory = self._upsert_territory(
                code=feature.code,
                boundary_version=version,
                dataset=dataset,
                level=feature.level,
                parent=parent,
                name_ru=feature.name_ru,
                name_kk=feature.name_kk,
                name_en=feature.name_en,
                iso3166_2=feature.iso3166_2,
                kato_code=feature.kato_code,
                osm_relation_id=feature.osm_relation_id,
                natural_key=feature.osm_ref,
                source_row_ref=feature.source_row_ref,
                data_as_of=spec.osm_data_as_of,
                area_km2=Decimal(str(declared)) if declared is not None else None,
                valid_from=VALID_FROM.get(feature.code),
                notes=" ".join(notes) or None,
            )
            self._write_geometry(territory, feature)
            self._alias_candidates.setdefault(feature.code, []).extend(alias_candidates(feature))

    def _load_aliases(self) -> None:
        """Записать алиасы. Вызывается дважды: до и после разбора населения.

        Написания из книги населения становятся известны только в момент
        сопоставления её строк, а сопоставление невозможно, пока алиасов нет.
        Повторный проход по тем же кандидатам ничего не дублирует — upsert
        идёт по паре «свёрнутая форма + территория».
        """
        assert self.job is not None
        folded = 0
        for code, candidates in self._alias_candidates.items():
            territory = self._territories[code]
            rows = dedupe_aliases(candidates)
            folded += len(candidates) - len(rows)
            for row in rows:
                existing = self.session.scalars(
                    select(TerritoryAlias).where(
                        TerritoryAlias.territory_id == territory.id,
                        TerritoryAlias.normalized == row.normalized,
                    )
                ).one_or_none()
                alias = existing or TerritoryAlias(
                    territory_id=territory.id, normalized=row.normalized
                )
                alias.alias = row.alias
                alias.kind = row.kind
                alias.source_layer = row.source_layer
                alias.notes = row.notes
                alias.is_ambiguous = False
                if existing is None:
                    self.session.add(alias)
        self.session.flush()
        self.report.aliases_folded = folded
        self._recount_aliases()

    def build_resolver(self) -> TerritoryResolver:
        """Собрать сопоставитель названий по тому, что лежит в базе."""
        resolver = TerritoryResolver()
        rows = self.session.execute(
            select(TerritoryAlias.alias, Territory.code).join(
                Territory, Territory.id == TerritoryAlias.territory_id
            )
        ).all()
        resolver.add_many((str(alias), str(code)) for alias, code in rows)
        return resolver

    def _mark_ambiguous(self, resolver: TerritoryResolver) -> None:
        """Пометить написания, подходящие нескольким территориям.

        Такой алиас не годится для автоматического связывания: строка книги с
        этим названием не может быть привязана без разбора человеком. Признак
        ставится в таблице, чтобы импортёры других слоёв видели его до попытки
        сопоставления, а не после.
        """
        ambiguous = resolver.ambiguous_names
        self.report.ambiguous_aliases = ambiguous
        if not ambiguous:
            return
        for name in ambiguous:
            self.session.execute(
                text(
                    "UPDATE territory_aliases SET is_ambiguous = true, updated_at = now() "
                    "WHERE normalized = :normalized"
                ),
                {"normalized": name},
            )
            self._issue(
                IssueSeverity.WARNING,
                "ambiguous_alias",
                f"Написание «{name}» подходит нескольким территориям и помечено как "
                "неоднозначное: автоматическое связывание по нему запрещено.",
                raw_value=name,
            )

    def _load_population(self, rows: list[PopulationRow], dataset: SourceDataset) -> None:
        assert self.job is not None
        resolver = self.build_resolver()
        wanted = [row for row in rows if row.kind in {"unit", "region"}]

        for row in rows:
            if row.kind == "center":
                self.report.population_skipped += 1
            elif row.kind == "country":
                self.report.population_skipped += 1
                self._issue(
                    IssueSeverity.INFO,
                    "row_out_of_scope",
                    "Строка «Республика Казахстан» в книге есть, но не загружается: "
                    "задача ограничена Алматинской областью. Цифра остаётся доступной "
                    "в источнике и используется только для контроля доли области.",
                    source_row_ref=row.source_row_ref,
                )
        if self.report.population_skipped:
            self._issue(
                IssueSeverity.INFO,
                "level_not_in_reference",
                f"Пропущено строк райцентров: "
                f"{sum(1 for r in rows if r.kind == 'center')}. Территорий уровня "
                "«населённый пункт» в справочнике нет — их границы отсутствуют во всех "
                "наличных геонаборах, а заводить единицу без границы бессмысленно.",
            )

        for row in wanted:
            resolution = resolver.resolve(row.raw_name)
            if not resolution.ok or resolution.territory_code is None:
                self._issue(
                    IssueSeverity.ERROR,
                    "territory_not_resolved",
                    f"Строка «{row.raw_name}» не сопоставлена с территорией: "
                    f"{resolution.reason}.",
                    source_row_ref=row.source_row_ref,
                    raw_value=row.raw_name,
                    context={"candidates": list(resolution.candidates)},
                )
                continue

            territory = self._territories[resolution.territory_code]
            total = row.values["total"]
            if total is None:
                self._issue(
                    IssueSeverity.ERROR,
                    "population_total_missing",
                    f"Строка «{row.raw_name}»: нет численности всего населения.",
                    source_row_ref=row.source_row_ref,
                )
                continue

            existing = self.session.scalars(
                select(PopulationStat).where(
                    PopulationStat.territory_id == territory.id,
                    PopulationStat.as_of_date == POPULATION_AS_OF,
                )
            ).one_or_none()
            stat = existing or PopulationStat(
                territory_id=territory.id, as_of_date=POPULATION_AS_OF
            )
            stat.total = total
            for column in POPULATION_COLUMNS[1:]:
                setattr(stat, column, row.values[column])
            stat.source_dataset_id = dataset.id
            stat.import_job_id = self.job.id
            stat.source_row_ref = row.source_row_ref
            stat.natural_key = f"{territory.code}@{POPULATION_AS_OF.isoformat()}"
            stat.imported_at = utcnow()
            stat.data_as_of = POPULATION_AS_OF
            # Прочерк, разобранный в подтверждённый ноль, — не претензия к
            # строке: категория населения действительно отсутствует. Статус
            # «warning» остаётся только там, где прочерк пришлось оставить NULL.
            stat.validation_status = (
                "warning"
                if any(not d.startswith("0:") for d in row.dash_decisions.values())
                else "ok"
            )
            stat.validation_notes = (
                {"dash_decisions": row.dash_decisions} if row.dash_decisions else None
            )
            if existing is None:
                self.session.add(stat)
            self.report.population_written += 1

            # Написание из книги населения — тоже свидетельство источника.
            self._alias_candidates.setdefault(territory.code, []).append(
                AliasCandidate(
                    row.raw_name, AliasKind.SOURCE_SPELLING, source_layer="население"
                )
            )
            if row.center_name and territory.admin_center_name is None:
                territory.admin_center_name = row.center_name

            for column, decision in row.dash_decisions.items():
                severity = (
                    IssueSeverity.INFO if decision.startswith("0:") else IssueSeverity.WARNING
                )
                self._issue(
                    severity,
                    "dash_interpreted",
                    f"«{row.raw_name}», колонка {column}: прочерк в источнике → {decision}",
                    source_row_ref=row.source_row_ref,
                    column_name=column,
                    raw_value="-",
                )
        self.session.flush()

    def _check_geometry_coverage(self) -> dict[str, Any]:
        """Сверить сумму площадей единиц с площадью области.

        Проверяется не сумма чисел, а площадь объединения полигонов: сумма
        отдельных площадей скрывает и перекрытия, и дыры, а объединение — нет.
        """
        oblast = self._territories.get(REGION_CODES["KZ-19"])
        if oblast is None:
            return {}
        union_area = self.session.execute(
            _UNION_AREA_SQL, {"parent_id": str(oblast.id)}
        ).scalar_one_or_none()
        oblast_area = float(oblast.area_km2_computed or 0)
        if union_area is None or not oblast_area:
            return {}
        union_value = float(union_area)
        delta = (union_value - oblast_area) / oblast_area
        result = {
            "units_union_km2": round(union_value, 1),
            "oblast_polygon_km2": round(oblast_area, 1),
            "delta_pct": round(delta * 100, 4),
        }
        if abs(delta) > 0.01:
            self._issue(
                IssueSeverity.WARNING,
                "coverage_mismatch",
                f"Объединение 11 единиц {union_value:.1f} км² расходится с полигоном "
                f"области {oblast_area:.1f} км² на {delta:+.2%}.",
                context=result,
            )
        return result

    def _finish_job(self, reconciliation: dict[str, Any]) -> None:
        assert self.job is not None
        self.job.rows_read = (
            EXPECTED_REGIONS + EXPECTED_ALMATY_UNITS + self.report.population_written
        )
        self.job.rows_created = self.report.territories_created
        self.job.rows_updated = self.report.territories_updated
        self.job.rows_skipped = self.report.population_skipped
        self.job.rows_failed = sum(
            1 for severity, _, _ in self.report.issues if severity == str(IssueSeverity.ERROR)
        )
        self.job.reconciliation = reconciliation
        self.job.territory_match_report = {
            "territories_by_level": self.report.territories_by_level,
            "aliases_by_kind": self.report.aliases_by_kind,
            "aliases_folded": self.report.aliases_folded,
            "ambiguous_names": list(self.report.ambiguous_aliases),
        }
        self.job.finished_at = utcnow()
        self.job.status = ImportStatus.DRY_RUN if self.dry_run else ImportStatus.SUCCEEDED
        self.session.flush()

    # --- основной сценарий -------------------------------------------------

    def run(self) -> LoadReport:
        from scripts.source_manifest import resolve_source

        self._start_job()

        regions_path = self.boundaries_dir / REGIONS_SET.file_name
        almaty_path = self.boundaries_dir / ALMATY_SET.file_name
        aliases_path = self.boundaries_dir / ALIASES_8_3_FILE
        population_path = resolve_source(self.source_dir, POPULATION_FILE_NAME)

        regions_file = self._source_file(
            regions_path, origin="boundaries_dir", expected_sha256=REGIONS_SET.sha256
        )
        almaty_file = self._source_file(
            almaty_path, origin="boundaries_dir", expected_sha256=ALMATY_SET.sha256
        )
        aliases_file = self._source_file(
            aliases_path, origin="boundaries_dir", expected_sha256=ALIASES_8_3_SHA256
        )
        population_file = self._source_file(
            population_path, origin="source_data_dir", expected_sha256=None
        )

        regions = parse_regions(regions_path)
        almaty_units = parse_almaty_units(almaty_path)
        if len(regions) != EXPECTED_REGIONS:
            raise ValueError(f"Ожидалось {EXPECTED_REGIONS} регионов, в файле {len(regions)}")
        if len(almaty_units) != EXPECTED_ALMATY_UNITS:
            raise ValueError(
                f"Ожидалось {EXPECTED_ALMATY_UNITS} единиц области, в файле {len(almaty_units)}"
            )

        regions_dataset = self._dataset(
            regions_file,
            sheet_name=REGIONS_SET.file_name,
            role="raw",
            row_count=len(regions),
            data_as_of=REGIONS_SET.osm_data_as_of,
        )
        almaty_dataset = self._dataset(
            almaty_file,
            sheet_name=ALMATY_SET.file_name,
            role="raw",
            row_count=len(almaty_units),
            data_as_of=ALMATY_SET.osm_data_as_of,
        )
        aliases_dataset = self._dataset(
            aliases_file,
            sheet_name=ALIASES_8_3_FILE,
            role="reconciliation",
            row_count=EXPECTED_REGIONS,
            data_as_of=REGIONS_SET.osm_data_as_of,
        )
        population_dataset = self._dataset(
            population_file,
            sheet_name=POPULATION_SHEET,
            role="raw",
            row_count=POPULATION_LAST_ROW - POPULATION_FIRST_ROW + 1,
            data_as_of=POPULATION_AS_OF,
            header_row=4,
        )

        regions_version = self._boundary_version(REGIONS_SET)
        almaty_version = self._boundary_version(ALMATY_SET)

        # Страна — синтетический корень иерархии: отдельного полигона страны в
        # поставляемых файлах нет, а иерархия без корня не строится. Всё, что
        # про эту запись неизвестно, оставлено пустым и объяснено в notes.
        self._upsert_territory(
            code=COUNTRY_CODE,
            boundary_version=regions_version,
            dataset=regions_dataset,
            level=TerritoryLevel.COUNTRY,
            parent=None,
            name_ru="Республика Казахстан",
            name_kk="Қазақстан Республикасы",
            name_en="Republic of Kazakhstan",
            iso3166_2=None,
            osm_relation_id=214665,
            natural_key="relation/214665",
            source_row_ref=None,
            data_as_of=REGIONS_SET.osm_data_as_of,
            notes=(
                "Корень иерархии. Полигон страны в поставку не включён "
                "(выгружался только как эталон проверки топологии), поэтому "
                "геометрия и вычисленная площадь отсутствуют. КАТО и ISO 3166-2 "
                "не заполнены: у страны это коды другого классификатора."
            ),
        )
        self._issue(
            IssueSeverity.INFO,
            "synthetic_root",
            "Запись «Республика Казахстан» создана импортёром как корень иерархии; "
            "геометрия отсутствует, потому что контур страны в набор не входит.",
        )

        self._load_boundary_set(REGIONS_SET, regions, regions_dataset, regions_version)
        self._load_boundary_set(ALMATY_SET, almaty_units, almaty_dataset, almaty_version)

        self._issue(
            IssueSeverity.INFO,
            "duplicate_osm_object",
            "relation/215718 (Алматинская область) присутствует в обоих геонаборах. "
            "Загружен один раз, из набора регионов; второй экземпляр пропущен, "
            "чтобы не хранить два полигона одного объекта.",
            context={"osm_ref": "relation/215718"},
        )

        without_kato = sorted(
            code for code, territory in self._territories.items() if territory.kato_code is None
        )
        self._issue(
            IssueSeverity.INFO,
            "kato_missing",
            f"КАТО отсутствует у {len(without_kato)} из {len(self._territories)} территорий. "
            "Код есть только у Алматинской и Карагандинской областей; у районов его нет "
            "ни в одном источнике. Коды не достраивались.",
            context={"codes": without_kato},
        )

        aliases_payload: dict[str, Any] = json.loads(aliases_path.read_text(encoding="utf-8"))
        for code, candidates in book_alias_candidates(aliases_payload).items():
            self._alias_candidates.setdefault(code, []).extend(candidates)
        self._load_aliases()

        population_rows = parse_population(population_path)
        self._load_population(population_rows, population_dataset)
        # Написания из книги населения стали известны при сопоставлении —
        # второй проход доводит таблицу алиасов до полного состава.
        self._load_aliases()

        self._mark_ambiguous(self.build_resolver())

        population_check = check_population_totals(population_rows)
        if not population_check["all_columns_match"]:
            self._issue(
                IssueSeverity.ERROR,
                "population_control_sum_failed",
                "Сумма 11 единиц не совпала с итогом области.",
                context=population_check["per_column"],
            )
        for label, key in (
            ("мужчины + женщины ≠ всё население", "gender_mismatches"),
            ("город + село ≠ всё население", "settlement_mismatches"),
        ):
            if population_check[key]:
                self._issue(
                    IssueSeverity.ERROR,
                    "population_row_sum_failed",
                    f"{label}: {', '.join(population_check[key])}",
                )

        reconciliation = {
            "population": population_check,
            "geometry_coverage": self._check_geometry_coverage(),
            "aliases_used": aliases_dataset.sheet_name,
        }
        self._finish_job(reconciliation)
        self.report.reconciliation = reconciliation

        if self.dry_run:
            # Откат — единственный способ показать настоящий результат вставок
            # и при этом ничего не записать. Отчёт уже собран в памяти.
            self.session.rollback()
        return self.report

    def _recount_aliases(self) -> None:
        """Пересчитать алиасы по видам прямо в базе.

        Считается состояние таблицы, а не число обработанных кандидатов:
        отчёт обязан показывать то, что записано, а не то, что предполагалось.
        """
        rows = self.session.execute(
            select(TerritoryAlias.kind, func.count()).group_by(TerritoryAlias.kind)
        ).all()
        self.report.aliases_by_kind = {str(kind): int(count) for kind, count in rows}


def load_territories(
    session: Session,
    *,
    data_dir: Path,
    source_dir: Path,
    dry_run: bool = False,
) -> LoadReport:
    """Загрузить справочник территорий целиком."""
    return TerritoryLoader(
        session, data_dir=data_dir, source_dir=source_dir, dry_run=dry_run
    ).run()


__all__ = [
    "ALIASES_8_3_FILE",
    "ALMATY_SET",
    "ALMATY_UNIT_CODES",
    "ATTRIBUTION_TEXT",
    "COUNTRY_CODE",
    "DOCUMENT_AREA_KM2",
    "EXPECTED_ALMATY_UNITS",
    "EXPECTED_POPULATION_ROWS",
    "EXPECTED_REGIONS",
    "LICENSE_NAME",
    "POPULATION_AS_OF",
    "POPULATION_COLUMNS",
    "POPULATION_FILE_NAME",
    "REGIONS_SET",
    "REGION_CODES",
    "AliasCandidate",
    "AliasRow",
    "BoundarySetSpec",
    "LoadReport",
    "PopulationRow",
    "TerritoryFeature",
    "TerritoryLoader",
    "alias_candidates",
    "book_alias_candidates",
    "check_population_totals",
    "classify_population_row",
    "dedupe_aliases",
    "load_territories",
    "parse_almaty_units",
    "parse_population",
    "parse_regions",
    "short_form",
]
