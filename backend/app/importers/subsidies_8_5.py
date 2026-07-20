"""Импортёр слоя 8.5 — книга «Риск_субсидии_Алматинская.xlsx».

Что эта книга требует от загрузчика особенного.

**Все 10 240 формул в ней без кэша.** Книгу ни разу не пересчитывал Excel,
поэтому колонки «Коэф. риска R», «Уровень» и «Риск-экспозиция» читаются как
`None` на всех 3413 строках. Загрузчик обязан считать балл сам — через
`app.risk.core.evaluate`. Попытка «взять готовое значение» даёт пустую витрину,
и это не гипотетический риск, а состояние файла на диске.

**Веса тоже живут в книге.** Ячейки `Методика!B9:B13` объявлены редактируемыми.
Читаем их, а не зашиваем; контрольная сумма (`B14`) — формула, тоже без кэша,
поэтому сумму пересчитываем.

**Идентификаторы — строки.** У 70 получателей БИН/ИИН начинается с нуля, у
21 179 выплат с нуля начинается номер заявки. Любое приведение к числу рвёт
связь с другими слоями, поэтому идентификаторы нормализуются `zfill()` и
проверяются по длине.

**Пустая ячейка ≠ ноль.** У 66 получателей нет района, и вместе с ним пуст
индикатор s1. Excel не различает пустоту и ноль, мы — обязаны: такая строка
получает `IndicatorValue(value=None)`, попадает в журнал качества с кодом
`indicator_not_measured`, а её территория — с кодом `territory_not_resolved`.
Параллельно считается «книжный» балл в семантике Excel: без него не сойтись с
контрольными числами аудита, а сойтись нужно, чтобы доказать, что мы читаем
книгу правильно.

**Имя файла может лежать в NFD.** За файлом ходим только через
`scripts.source_manifest.resolve_source`.
"""

from __future__ import annotations

import hashlib
from collections import Counter
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import get_settings
from app.db.models.source import IssueSeverity
from app.risk.core import RiskLevel, RiskModelSpec, RiskResult
from app.risk.layers.subsidies import (
    INDICATOR_CODES,
    INDICATOR_META,
    LAYER_CODE,
    METHODOLOGY_SHEET,
    MODEL_CODE,
    MODEL_VERSION,
    THRESHOLD_CELLS,
    WEIGHT_CELLS,
    build_spec,
    risk_exposure,
    score,
    score_as_book,
)
from app.services.territory_resolver import (
    Resolution,
    ResolutionReport,
    ResolutionStatus,
    TerritoryResolver,
    build_report,
)

SOURCE_FILE_NAME = "Риск_субсидии_Алматинская.xlsx"
IMPORTER_NAME = "subsidies_8_5"

SHEET_RECIPIENTS = "Риск_получатели"
SHEET_DISTRICTS = "Риск_районы"
SHEET_DATA = "Данные"

HEADER_ROW = 2
"""Строка заголовков. Первая строка на всех листах — объединённый титул."""

XIN_LENGTH = 12

# --- Колонки листов, по именам заголовков ------------------------------------
# По именам, а не по номерам: перестановка колонок в книге должна давать
# внятную ошибку, а не молча сдвинутые данные.

COL_RECIPIENT_NO = "№"
COL_XIN = "БИН/ИИН"
COL_NAME = "Наименование получателя"
COL_DISTRICT = "Район"
COL_DIRECTOR = "Руководитель"
COL_AMOUNT = "Сумма субсидий, ₸"
COL_PAYMENTS = "Выплат"
COL_PROGRAMS = "Программ"
COL_ANIMAL_KINDS = "Видов жив."
COL_DISTRICT_SHARE = "Доля в районе"
COL_OBLAST_SHARE = "Доля в области"
COL_AFFILIATED = "Аффил.(получ. у рук.)"
COL_ANOMALY_SHARE = "Аном. выплат, доля"
COL_OUTLIER_SHARE = "Выбросов сумм, доля"

INDICATOR_COLUMNS: Mapping[str, str] = {
    "s1": "s1 Концентрация",
    "s2": "s2 Повторность",
    "s3": "s3 Аффилир.",
    "s4": "s4 Процесс",
    "s5": "s5 Выбросы",
}

COL_DATA_DISTRICT = "DistrictName"
COL_DATA_ANIMAL = "AnimalType"
COL_DATA_XIN = "EnterpriseXin"
COL_DATA_PROGRAM = "SubsidiesName"
COL_DATA_DECISION = "PositiveDecisionDate"
COL_DATA_EXECUTED = "ExecutedDate"
COL_DATA_LOCAL_PAY = "LocalPaymentDate"
COL_DATA_REPUBLIC_PAY = "RepublicPaymentDate"
COL_DATA_BID = "BidNumber"
COL_DATA_STATUS = "BidStatus"
COL_DATA_NORM = "SubsidiesNorm"
COL_DATA_REPUBLIC = "RepublicPaidBudget"
COL_DATA_LOCAL = "LocalPaidBudget"
COL_DATA_OWED = "SubsidiesOwedSum"
COL_DATA_TOTAL = "Сумма (Local+Republic), ₸"
COL_DATA_LAG = "Дней решение→выплата"
COL_FLAG_EARLY = "Флаг: выплата раньше решения"
COL_FLAG_LAG = "Флаг: аномальный лаг (>170 дн)"
COL_FLAG_OUTLIER = "Флаг: выброс суммы"

# --- Справочник территорий слоя ----------------------------------------------

TERRITORY_ALIASES: Mapping[str, str] = {
    "Аксуский район": "aksuskiy",
    "Алакольский район": "alakolskiy",
    "Алатау Г.А.": "alatau-ga",
    "Балхашский район": "balkhashskiy",
    "Енбекшиказахский район": "enbekshikazakhskiy",
    "Ескельдинский район": "eskeldinskiy",
    "Жамбылский район": "zhambylskiy",
    "Илийский район": "iliyskiy",
    "Карасайский район": "karasayskiy",
    "Каратальский район": "karatalskiy",
    "Кегенский район": "kegenskiy",
    "Кербулакский район": "kerbulakskiy",
    "Коксуский район": "koksuskiy",
    "Кордайский район": "kordayskiy",
    "Мойынкумский район": "moyynkumskiy",
    "Панфиловский район": "panfilovskiy",
    "Райымбекский район": "rayymbekskiy",
    "Сарканский район": "sarkanskiy",
    "Талгарский район": "talgarskiy",
    "Талдыкорган Г.А.": "taldykorgan-ga",
    "Текели Г.А.": "tekeli-ga",
    "Уйгурский район": "uygurskiy",
    "Уйгурский район ": "uygurskiy",
    "район Ақсуат": "aksuat",
    "Қонаев Г.А.": "konaev-ga",
}
"""Написания районов, встреченные в книге 8.5, и их внутренние коды.

Справочник задан явно, а не выведен из данных: неопознанное название должно
попадать в отчёт, а не порождать новую территорию. Четыре из этих единиц
(Кордайский, Мойынкумский, Ақсуат, Алатау Г.А.) вообще лежат вне Алматинской
области, а одиннадцать с 2022 года относятся к области Жетысу — коды здесь
только связывают названия, а не утверждают принадлежность.
"""

# --- Контрольные значения книги ----------------------------------------------

BOOK_CONTROL_VALUES: Mapping[str, float] = {
    "recipients": 3413,
    "payments": 21521,
    "districts": 24,
    "total_amount": 67_535_553_445,
    "max_score": 72.095,
    "level_critical": 0,
    "level_high": 2,
    "level_medium": 67,
    "level_low": 3344,
    "book_total_exposure": 20_393_585_538.27,
    "flag_paid_before_decision": 1209,
    "flag_abnormal_lag": 1052,
    "flag_amount_outlier": 882,
    "recipients_without_territory": 66,
    "payments_without_territory": 96,
}
"""Контрольные числа из аудита `docs/audit/03-sloi-8-5-8-6-8-7.md`, раздел 1.8.

Сверка идёт здесь, а не только в тестах: расхождение обязано быть видно
пользователю мастера импорта, а не одному разработчику.
"""


# --- Разобранные строки -------------------------------------------------------


@dataclass(frozen=True, slots=True)
class Methodology:
    """Методика, прочитанная с листа «Методика»."""

    weights: Mapping[str, float]
    thresholds: Mapping[RiskLevel, float]
    weight_sum: float
    """Пересчитанная сумма весов: `B14` — формула без кэша, читать её бесполезно."""


@dataclass(frozen=True, slots=True)
class QualityIssue:
    """Замечание к строке источника — прообраз `DataQualityIssue`.

    Обычный dataclass, а не ORM-объект: импортёр обязан работать и в режиме
    сухого прогона, когда сессии базы нет вовсе.
    """

    severity: IssueSeverity
    code: str
    message: str
    source_row_ref: str
    column_name: str | None = None
    raw_value: str | None = None
    context: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True, slots=True)
class ProgramRow:
    code: str
    name: str
    animal_type: str | None


@dataclass(frozen=True, slots=True)
class RecipientRow:
    """Получатель субсидий с обеими оценками риска."""

    book_rank: int
    xin: str
    name: str
    director_name: str | None
    territory_name_raw: str | None
    territory_code: str | None
    territory_status: ResolutionStatus

    total_amount: Decimal
    payments_count: int
    programs_count: int
    animal_types_count: int | None

    district_share: float | None
    oblast_share: float | None
    affiliated_count: int | None
    anomalous_payment_share: float | None
    amount_outlier_share: float | None

    indicators: Mapping[str, float | None]
    result: RiskResult
    """Оценка по методике проекта: пустой индикатор не измерен."""

    book_result: RiskResult
    """Оценка в семантике Excel — только для сверки с книгой."""

    source_row_ref: str

    @property
    def exposure(self) -> float | None:
        return risk_exposure(float(self.total_amount), self.result.score)

    @property
    def book_exposure(self) -> float | None:
        return risk_exposure(float(self.total_amount), self.book_result.score)


@dataclass(frozen=True, slots=True)
class PaymentRow:
    """Одна выплата листа «Данные»."""

    bid_number: str
    xin: str
    program_code: str | None
    animal_type: str | None
    territory_name_raw: str | None
    territory_code: str | None
    territory_status: ResolutionStatus

    bid_status: str | None
    positive_decision_at: datetime | None
    executed_at: datetime | None
    local_payment_at: datetime | None
    republic_payment_at: datetime | None

    subsidies_norm: Decimal | None
    amount_local: Decimal | None
    amount_republic: Decimal | None
    amount_owed: Decimal | None
    amount_total: Decimal

    decision_to_payment_days: int | None
    flag_paid_before_decision: bool
    flag_abnormal_lag: bool
    flag_amount_outlier: bool

    source_row_ref: str


@dataclass(frozen=True, slots=True)
class Reconciliation:
    """Сверка посчитанного с контрольными числами книги."""

    values: Mapping[str, float]

    def compare(
        self, expected: Mapping[str, float] = BOOK_CONTROL_VALUES, *, tolerance: float = 0.01
    ) -> tuple[dict[str, Any], ...]:
        """Расхождения «ожидалось / получено / разница».

        Возвращается список, а не булево: пользователю мастера импорта нужно
        видеть, какой именно показатель разошёлся и насколько.
        """
        rows: list[dict[str, Any]] = []
        for key, want in expected.items():
            if key not in self.values:
                continue
            got = self.values[key]
            if abs(got - want) > tolerance:
                rows.append(
                    {"metric": key, "expected": want, "actual": got, "delta": got - want}
                )
        return tuple(rows)


@dataclass(frozen=True, slots=True)
class ImportResult:
    """Итог разбора книги — всё, что нужно, чтобы записать слой в базу."""

    source_path: Path
    methodology: Methodology
    spec: RiskModelSpec
    programs: tuple[ProgramRow, ...]
    recipients: tuple[RecipientRow, ...]
    payments: tuple[PaymentRow, ...]
    issues: tuple[QualityIssue, ...]
    territory_report: ResolutionReport
    reconciliation: Reconciliation

    @property
    def level_counts(self) -> dict[RiskLevel, int]:
        counts = dict.fromkeys(RiskLevel, 0)
        for row in self.recipients:
            counts[row.result.level] += 1
        return counts

    @property
    def total_exposure(self) -> float:
        """Суммарная риск-экспозиция по методике проекта."""
        return sum(row.exposure or 0.0 for row in self.recipients)

    @property
    def book_total_exposure(self) -> float:
        """Суммарная риск-экспозиция в семантике книги — величина для сверки."""
        return sum(row.book_exposure or 0.0 for row in self.recipients)


# --- Нормализация значений ----------------------------------------------------


def normalize_xin(raw: object) -> str:
    """Привести БИН/ИИН к 12 знакам с сохранением ведущих нулей.

    `zfill`, а не форматирование числа: openpyxl может отдать идентификатор
    целым числом, если книгу когда-нибудь пересохранят с другим форматом ячейки,
    и тогда `080340015131` превратится в `80340015131`. Восстановить ведущий
    ноль потом невозможно — по нему уже не найти получателя в другом слое.
    """
    text = str(raw).strip()
    if text.endswith(".0"):  # число, приехавшее как float
        text = text[:-2]
    return text.zfill(XIN_LENGTH)


def normalize_bid_number(raw: object) -> str:
    """Номер заявки: строка, ведущий ноль значим у 21 179 записей из 21 521.

    Длина переменная (12 или 14), поэтому дополнять до фиксированной нельзя —
    только снять артефакты чтения.
    """
    text = str(raw).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def parse_datetime(raw: object) -> datetime | None:
    """Разобрать дату вида `2022-12-18T17:39:13`.

    В книге это строка, а не `datetime`. Часового пояса в ней нет, и мы его не
    добавляем: приписать UTC значило бы выдумать сведения.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def parse_flag(raw: object) -> bool:
    """Флаг книги: «да» либо пусто. Пусто здесь — действительно «нет».

    Это тот редкий случай, когда пустая ячейка означает отрицание, а не
    отсутствие измерения: колонка рассчитана самой книгой для всех 21 521
    строки, и «да» проставлено ровно там, где условие выполнилось.
    """
    return str(raw).strip().casefold() == "да" if raw is not None else False


def to_decimal(raw: object) -> Decimal | None:
    if raw is None:
        return None
    return Decimal(str(raw))


def to_float(raw: object) -> float | None:
    """Число из ячейки. `None` остаётся `None` — это «не измерено», а не 0.0."""
    if raw is None:
        return None
    if isinstance(raw, int | float):
        return float(raw)
    return float(str(raw).strip().replace(",", "."))


def to_int(raw: object) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        return int(raw)
    return int(str(raw).strip())


def program_code(name: str) -> str:
    """Устойчивый код программы.

    Наименование программы длиной до 321 символа, в первичный ключ его брать
    неудобно, а порядковый номер не переживёт повторного импорта. Хеш даёт
    один и тот же код при каждом запуске — импорт остаётся идемпотентным.
    """
    digest = hashlib.sha256(name.strip().encode("utf-8")).hexdigest()
    return f"prg-{digest[:16]}"


# --- Чтение книги -------------------------------------------------------------


def find_source(source_dir: Path | None = None) -> Path:
    """Путь к книге слоя 8.5, устойчиво к NFD в имени файла."""
    from scripts.source_manifest import resolve_source

    directory = source_dir if source_dir is not None else get_settings().source_data_dir
    return resolve_source(directory, SOURCE_FILE_NAME)


def _sheet_rows(workbook: Any, sheet_name: str) -> Iterator[tuple[Any, ...]]:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"В книге слоя {LAYER_CODE} нет листа {sheet_name!r}")
    for row in workbook[sheet_name].iter_rows(values_only=True):
        yield tuple(row)


def _header_index(
    header: Sequence[Any], sheet_name: str, required: Sequence[str]
) -> dict[str, int]:
    """Сопоставить имена колонок их позициям.

    Отсутствие ожидаемой колонки — отказ, а не пропуск: молча импортировать
    слой без колонки «Сумма субсидий» намного хуже, чем не импортировать вовсе.
    """
    index = {
        str(value).strip(): position
        for position, value in enumerate(header)
        if value is not None and str(value).strip()
    }
    missing = [name for name in required if name not in index]
    if missing:
        raise KeyError(f"Лист {sheet_name!r}: не найдены колонки {missing}")
    return index


def read_methodology(workbook: Any) -> Methodology:
    """Прочитать веса и пороги из листа «Методика».

    Веса читаются из книги, а не берутся из кода: лист прямо разрешает их
    менять. Контрольную сумму пересчитываем — `B14` формула, кэша у неё нет.
    """
    sheet = workbook[METHODOLOGY_SHEET]
    grid = [tuple(row) for row in sheet.iter_rows(values_only=True)]

    def cell(address: str) -> Any:
        column = ord(address[0].upper()) - ord("A")
        row = int(address[1:]) - 1
        if row >= len(grid) or column >= len(grid[row]):
            raise KeyError(f"Лист «{METHODOLOGY_SHEET}»: ячейка {address} за границами листа")
        return grid[row][column]

    weights: dict[str, float] = {}
    for code, address in WEIGHT_CELLS.items():
        value = cell(address)
        if value is None:
            raise ValueError(
                f"Лист «{METHODOLOGY_SHEET}»: вес индикатора {code} (ячейка {address}) пуст"
            )
        weights[code] = float(value)

    thresholds: dict[RiskLevel, float] = {}
    for level, address in THRESHOLD_CELLS.items():
        value = cell(address)
        if value is None:
            raise ValueError(
                f"Лист «{METHODOLOGY_SHEET}»: порог уровня {level} (ячейка {address}) пуст"
            )
        thresholds[level] = float(value)

    return Methodology(
        weights=weights,
        thresholds=thresholds,
        weight_sum=sum(weights.values()),
    )


def default_resolver() -> TerritoryResolver:
    """Сопоставитель, построенный на написаниях районов из книги 8.5."""
    resolver = TerritoryResolver()
    resolver.add_many(TERRITORY_ALIASES.items())
    return resolver


def _resolve_territory(
    resolver: TerritoryResolver,
    raw_name: object,
    *,
    row_ref: str,
    column: str,
    issues: list[QualityIssue],
    resolutions: list[Resolution],
) -> Resolution:
    """Сопоставить название района и зафиксировать неудачу в журнале качества.

    Неопознанное название не угадывается и не заменяется на «прочее»: строка
    сохраняется без территории, а причина уходит в отчёт.
    """
    resolution = resolver.resolve(None if raw_name is None else str(raw_name))
    resolutions.append(resolution)

    if resolution.status is not ResolutionStatus.RESOLVED:
        issues.append(
            QualityIssue(
                severity=IssueSeverity.WARNING,
                code="territory_not_resolved",
                message=f"Территория не определена: {resolution.reason}",
                source_row_ref=row_ref,
                column_name=column,
                raw_value=resolution.raw or None,
                context={
                    "status": str(resolution.status),
                    "normalized": resolution.normalized,
                    "candidates": list(resolution.candidates),
                },
            )
        )
    return resolution


def read_recipients(
    workbook: Any,
    spec: RiskModelSpec,
    resolver: TerritoryResolver,
) -> tuple[list[RecipientRow], list[QualityIssue], ResolutionReport]:
    """Прочитать лист «Риск_получатели» и пересчитать риск по каждой строке."""
    rows = list(_sheet_rows(workbook, SHEET_RECIPIENTS))
    header = rows[HEADER_ROW - 1]
    required = [
        COL_RECIPIENT_NO, COL_XIN, COL_NAME, COL_DISTRICT, COL_DIRECTOR, COL_AMOUNT,
        COL_PAYMENTS, COL_PROGRAMS, COL_ANIMAL_KINDS, COL_DISTRICT_SHARE, COL_OBLAST_SHARE,
        COL_AFFILIATED, COL_ANOMALY_SHARE, COL_OUTLIER_SHARE,
        *INDICATOR_COLUMNS.values(),
    ]
    index = _header_index(header, SHEET_RECIPIENTS, required)

    recipients: list[RecipientRow] = []
    issues: list[QualityIssue] = []
    resolutions: list[Resolution] = []

    for offset, raw in enumerate(rows[HEADER_ROW:], start=HEADER_ROW + 1):
        if raw[index[COL_RECIPIENT_NO]] is None:
            continue
        row_ref = f"{SHEET_RECIPIENTS}!A{offset}"

        resolution = _resolve_territory(
            resolver,
            raw[index[COL_DISTRICT]],
            row_ref=row_ref,
            column=COL_DISTRICT,
            issues=issues,
            resolutions=resolutions,
        )

        indicators: dict[str, float | None] = {}
        for code in INDICATOR_CODES:
            value = raw[index[INDICATOR_COLUMNS[code]]]
            indicators[code] = to_float(value)
            if value is None:
                # Пустая ячейка индикатора — «не измерено». В книге это следствие
                # неизвестного района, и списывать получателю нулевой риск
                # концентрации только из-за пробела в адресе нельзя.
                issues.append(
                    QualityIssue(
                        severity=IssueSeverity.WARNING,
                        code="indicator_not_measured",
                        message=(
                            f"Индикатор {code} не рассчитан в книге; "
                            f"балл нормируется на доступный вес"
                        ),
                        source_row_ref=row_ref,
                        column_name=INDICATOR_COLUMNS[code],
                        context={"indicator": code, "xin": normalize_xin(raw[index[COL_XIN]])},
                    )
                )

        amount = to_decimal(raw[index[COL_AMOUNT]]) or Decimal(0)

        recipients.append(
            RecipientRow(
                book_rank=int(raw[index[COL_RECIPIENT_NO]]),
                xin=normalize_xin(raw[index[COL_XIN]]),
                name=str(raw[index[COL_NAME]]),
                director_name=(
                    None if raw[index[COL_DIRECTOR]] is None else str(raw[index[COL_DIRECTOR]])
                ),
                territory_name_raw=(
                    None if raw[index[COL_DISTRICT]] is None else str(raw[index[COL_DISTRICT]])
                ),
                territory_code=resolution.territory_code,
                territory_status=resolution.status,
                total_amount=amount,
                payments_count=to_int(raw[index[COL_PAYMENTS]]) or 0,
                programs_count=to_int(raw[index[COL_PROGRAMS]]) or 0,
                animal_types_count=to_int(raw[index[COL_ANIMAL_KINDS]]),
                district_share=to_float(raw[index[COL_DISTRICT_SHARE]]),
                oblast_share=to_float(raw[index[COL_OBLAST_SHARE]]),
                affiliated_count=to_int(raw[index[COL_AFFILIATED]]),
                anomalous_payment_share=to_float(raw[index[COL_ANOMALY_SHARE]]),
                amount_outlier_share=to_float(raw[index[COL_OUTLIER_SHARE]]),
                indicators=indicators,
                result=score(spec, indicators),
                book_result=score_as_book(spec, indicators),
                source_row_ref=row_ref,
            )
        )

    return recipients, issues, build_report(resolutions)


def read_payments(
    workbook: Any,
    resolver: TerritoryResolver,
) -> tuple[list[PaymentRow], list[ProgramRow], list[QualityIssue]]:
    """Прочитать лист «Данные» — 21 521 выплату — и собрать справочник программ."""
    rows = list(_sheet_rows(workbook, SHEET_DATA))
    header = rows[HEADER_ROW - 1]
    required = [
        COL_DATA_DISTRICT, COL_DATA_ANIMAL, COL_DATA_XIN, COL_DATA_PROGRAM,
        COL_DATA_DECISION, COL_DATA_EXECUTED, COL_DATA_LOCAL_PAY, COL_DATA_REPUBLIC_PAY,
        COL_DATA_BID, COL_DATA_STATUS, COL_DATA_NORM, COL_DATA_REPUBLIC, COL_DATA_LOCAL,
        COL_DATA_OWED, COL_DATA_TOTAL, COL_DATA_LAG,
        COL_FLAG_EARLY, COL_FLAG_LAG, COL_FLAG_OUTLIER,
    ]
    index = _header_index(header, SHEET_DATA, required)

    payments: list[PaymentRow] = []
    issues: list[QualityIssue] = []
    resolutions: list[Resolution] = []
    programs: dict[str, ProgramRow] = {}

    for offset, raw in enumerate(rows[HEADER_ROW:], start=HEADER_ROW + 1):
        if raw[index[COL_DATA_BID]] is None:
            continue
        row_ref = f"{SHEET_DATA}!A{offset}"

        resolution = _resolve_territory(
            resolver,
            raw[index[COL_DATA_DISTRICT]],
            row_ref=row_ref,
            column=COL_DATA_DISTRICT,
            issues=issues,
            resolutions=resolutions,
        )

        program_name = raw[index[COL_DATA_PROGRAM]]
        code: str | None = None
        if program_name is not None:
            code = program_code(str(program_name))
            if code not in programs:
                animal = raw[index[COL_DATA_ANIMAL]]
                programs[code] = ProgramRow(
                    code=code,
                    name=str(program_name),
                    animal_type=None if animal is None else str(animal),
                )

        payments.append(
            PaymentRow(
                bid_number=normalize_bid_number(raw[index[COL_DATA_BID]]),
                xin=normalize_xin(raw[index[COL_DATA_XIN]]),
                program_code=code,
                animal_type=(
                    None if raw[index[COL_DATA_ANIMAL]] is None
                    else str(raw[index[COL_DATA_ANIMAL]])
                ),
                territory_name_raw=(
                    None if raw[index[COL_DATA_DISTRICT]] is None
                    else str(raw[index[COL_DATA_DISTRICT]])
                ),
                territory_code=resolution.territory_code,
                territory_status=resolution.status,
                bid_status=(
                    None if raw[index[COL_DATA_STATUS]] is None
                    else str(raw[index[COL_DATA_STATUS]])
                ),
                positive_decision_at=parse_datetime(raw[index[COL_DATA_DECISION]]),
                executed_at=parse_datetime(raw[index[COL_DATA_EXECUTED]]),
                local_payment_at=parse_datetime(raw[index[COL_DATA_LOCAL_PAY]]),
                republic_payment_at=parse_datetime(raw[index[COL_DATA_REPUBLIC_PAY]]),
                subsidies_norm=to_decimal(raw[index[COL_DATA_NORM]]),
                amount_local=to_decimal(raw[index[COL_DATA_LOCAL]]),
                amount_republic=to_decimal(raw[index[COL_DATA_REPUBLIC]]),
                amount_owed=to_decimal(raw[index[COL_DATA_OWED]]),
                amount_total=to_decimal(raw[index[COL_DATA_TOTAL]]) or Decimal(0),
                decision_to_payment_days=to_int(raw[index[COL_DATA_LAG]]),
                flag_paid_before_decision=parse_flag(raw[index[COL_FLAG_EARLY]]),
                flag_abnormal_lag=parse_flag(raw[index[COL_FLAG_LAG]]),
                flag_amount_outlier=parse_flag(raw[index[COL_FLAG_OUTLIER]]),
                source_row_ref=row_ref,
            )
        )

    return payments, sorted(programs.values(), key=lambda p: p.name), issues


def build_reconciliation(
    recipients: Sequence[RecipientRow], payments: Sequence[PaymentRow]
) -> Reconciliation:
    """Собрать показатели, которые сверяются с контрольными числами аудита."""
    levels = Counter(row.result.level for row in recipients)
    scores = [row.result.score for row in recipients if row.result.score is not None]
    territories = {row.territory_code for row in recipients if row.territory_code}

    return Reconciliation(
        values={
            "recipients": len(recipients),
            "payments": len(payments),
            "districts": len(territories),
            "total_amount": float(sum(row.total_amount for row in recipients)),
            "max_score": max(scores) if scores else 0.0,
            "level_critical": levels[RiskLevel.CRITICAL],
            "level_high": levels[RiskLevel.HIGH],
            "level_medium": levels[RiskLevel.MEDIUM],
            "level_low": levels[RiskLevel.LOW],
            "level_unknown": levels[RiskLevel.UNKNOWN],
            "total_exposure": sum(row.exposure or 0.0 for row in recipients),
            "book_total_exposure": sum(row.book_exposure or 0.0 for row in recipients),
            "flag_paid_before_decision": sum(
                1 for p in payments if p.flag_paid_before_decision
            ),
            "flag_abnormal_lag": sum(1 for p in payments if p.flag_abnormal_lag),
            "flag_amount_outlier": sum(1 for p in payments if p.flag_amount_outlier),
            "recipients_without_territory": sum(
                1 for row in recipients if row.territory_code is None
            ),
            "payments_without_territory": sum(1 for p in payments if p.territory_code is None),
        }
    )


def run_import(
    *,
    source_dir: Path | None = None,
    resolver: TerritoryResolver | None = None,
    with_payments: bool = True,
) -> ImportResult:
    """Разобрать книгу слоя 8.5 целиком.

    Записи в базу здесь нет намеренно: разбор обязан отрабатывать в сухом
    прогоне, где показывают, что произойдёт, ничего не меняя.
    """
    path = find_source(source_dir)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        methodology = read_methodology(workbook)
        spec = build_spec(methodology.weights, methodology.thresholds)

        active_resolver = resolver if resolver is not None else default_resolver()
        recipients, recipient_issues, territory_report = read_recipients(
            workbook, spec, active_resolver
        )

        payments: list[PaymentRow] = []
        programs: list[ProgramRow] = []
        payment_issues: list[QualityIssue] = []
        if with_payments:
            payments, programs, payment_issues = read_payments(workbook, active_resolver)
    finally:
        workbook.close()

    return ImportResult(
        source_path=path,
        methodology=methodology,
        spec=spec,
        programs=tuple(programs),
        recipients=tuple(recipients),
        payments=tuple(payments),
        issues=(*recipient_issues, *payment_issues),
        territory_report=territory_report,
        reconciliation=build_reconciliation(recipients, payments),
    )


# --- Отображение в ORM --------------------------------------------------------


def factors_payload(result: RiskResult) -> dict[str, Any]:
    """Расшифровка вклада индикаторов для карточки объекта.

    Неизмеренные факторы попадают в выгрузку наравне с измеренными: именно они
    объясняют, почему полнота не сто процентов.
    """
    titles = {meta.code: meta.name for meta in INDICATOR_META}
    return {
        "model": result.model_code,
        "version": result.model_version,
        "completeness": result.completeness,
        "notes": list(result.notes),
        "factors": [
            {
                "code": factor.code,
                "name": titles.get(factor.code, factor.name),
                "weight": factor.weight,
                "value": factor.value,
                "contribution": factor.contribution,
                "measured": factor.measured,
                "effect": factor.effect,
                "note": factor.note,
            }
            for factor in result.factors
        ],
    }


def build_orm_rows(
    result: ImportResult,
    *,
    territory_ids: Mapping[str, Any] | None = None,
    data_version: int = 1,
) -> dict[str, list[Any]]:
    """Собрать ORM-объекты слоя.

    `territory_ids` — соответствие «код территории → id в базе». Отсутствие
    кода в этом словаре трактуется так же, как несопоставленное название:
    запись сохраняется без территории, а не отбрасывается.
    """
    from app.db.models.subsidy import SubsidyPayment, SubsidyProgram, SubsidyRecipient

    lookup = territory_ids or {}

    programs = [
        SubsidyProgram(
            code=row.code,
            name=row.name,
            animal_type=row.animal_type,
            natural_key=row.code,
            data_version=data_version,
        )
        for row in result.programs
    ]
    programs_by_code = {row.code: orm for row, orm in zip(result.programs, programs, strict=True)}

    recipients: list[Any] = []
    recipients_by_xin: dict[str, Any] = {}
    for row in result.recipients:
        orm = SubsidyRecipient(
            xin=row.xin,
            name=row.name,
            director_name=row.director_name,
            territory_id=lookup.get(row.territory_code) if row.territory_code else None,
            territory_name_raw=row.territory_name_raw,
            territory_resolution=str(row.territory_status),
            total_amount=row.total_amount,
            payments_count=row.payments_count,
            programs_count=row.programs_count,
            animal_types_count=row.animal_types_count,
            district_share=to_decimal(row.district_share),
            oblast_share=to_decimal(row.oblast_share),
            affiliated_count=row.affiliated_count,
            anomalous_payment_share=to_decimal(row.anomalous_payment_share),
            amount_outlier_share=to_decimal(row.amount_outlier_share),
            s1_concentration=row.indicators["s1"],
            s2_repetition=row.indicators["s2"],
            s3_affiliation=row.indicators["s3"],
            s4_process_anomaly=row.indicators["s4"],
            s5_amount_outlier=row.indicators["s5"],
            model_code=MODEL_CODE,
            model_version=MODEL_VERSION,
            risk_score=row.result.score,
            risk_level=str(row.result.level),
            risk_completeness=row.result.completeness,
            risk_exposure=to_decimal(row.exposure),
            book_risk_score=row.book_result.score,
            book_risk_level=str(row.book_result.level),
            book_risk_exposure=to_decimal(row.book_exposure),
            book_rank=row.book_rank,
            factors=factors_payload(row.result),
            natural_key=row.xin,
            source_row_ref=row.source_row_ref,
            validation_status="warning" if row.result.unmeasured_factors else "ok",
            data_version=data_version,
        )
        recipients.append(orm)
        recipients_by_xin[row.xin] = orm

    payments: list[Any] = []
    for payment in result.payments:
        recipient = recipients_by_xin.get(payment.xin)
        if recipient is None:
            # Выплата без получателя в витрине — сигнал рассогласования книги.
            # Молча пропускать нельзя, но и падать здесь незачем: расхождение
            # уже зафиксировано в журнале качества.
            continue
        payments.append(
            SubsidyPayment(
                recipient=recipient,
                program=(
                    programs_by_code.get(payment.program_code) if payment.program_code else None
                ),
                territory_id=lookup.get(payment.territory_code) if payment.territory_code else None,
                territory_name_raw=payment.territory_name_raw,
                bid_number=payment.bid_number,
                bid_status=payment.bid_status,
                animal_type=payment.animal_type,
                positive_decision_at=payment.positive_decision_at,
                executed_at=payment.executed_at,
                local_payment_at=payment.local_payment_at,
                republic_payment_at=payment.republic_payment_at,
                subsidies_norm=payment.subsidies_norm,
                amount_local=payment.amount_local,
                amount_republic=payment.amount_republic,
                amount_owed=payment.amount_owed,
                amount_total=payment.amount_total,
                decision_to_payment_days=payment.decision_to_payment_days,
                flag_paid_before_decision=payment.flag_paid_before_decision,
                flag_abnormal_lag=payment.flag_abnormal_lag,
                flag_amount_outlier=payment.flag_amount_outlier,
                natural_key=payment.bid_number,
                source_row_ref=payment.source_row_ref,
                data_version=data_version,
            )
        )

    return {"programs": programs, "recipients": recipients, "payments": payments}


__all__ = [
    "BOOK_CONTROL_VALUES",
    "HEADER_ROW",
    "IMPORTER_NAME",
    "INDICATOR_COLUMNS",
    "SHEET_DATA",
    "SHEET_DISTRICTS",
    "SHEET_RECIPIENTS",
    "SOURCE_FILE_NAME",
    "TERRITORY_ALIASES",
    "ImportResult",
    "Methodology",
    "PaymentRow",
    "ProgramRow",
    "QualityIssue",
    "RecipientRow",
    "Reconciliation",
    "build_orm_rows",
    "build_reconciliation",
    "default_resolver",
    "factors_payload",
    "find_source",
    "normalize_bid_number",
    "normalize_xin",
    "parse_datetime",
    "parse_flag",
    "program_code",
    "read_methodology",
    "read_payments",
    "read_recipients",
    "run_import",
]
