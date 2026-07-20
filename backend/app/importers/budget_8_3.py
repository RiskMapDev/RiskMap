"""Импортёр слоя 8.3 «Бюджетные риски».

Книга `Слой_8_3_бюджетных_рисков_КЗ_2025 (1).xlsx`, 11 листов, 10,8 МБ.
Читаются три из них:

* `Параметры` — веса, пороги и направления 15 индикаторов. Лист объявлен
  редактируемой моделью, поэтому значения именно вычитываются, а не считаются
  зашитыми: расхождение с константами кода — повод остановиться, а не молча
  пересчитать по своим числам.
* `Расчет_месяц` — 240 расчётных строк «область × месяц».
* `RAW_DATA_Бюджет_все_регионы_КЗ_` — 74 831 строка бюджетной иерархии.

Про расчётный лист важно понимать вот что: **колонки A–AL в нём константы, а
не формулы**. Агрегация 74 831 сырой строки в 240 расчётных выполнена вне
Excel и внутри книги невоспроизводима. Поэтому импортёр берёт из книги сырые
показатели индикаторов (колонки AT…BG) как входные данные, а всё, что считается
после них — баллы, взвешенную сумму, пол, уровень, — пересчитывает сам через
`app.risk.layers.budget`. Именно это делает сверку с колонкой `Risk Score`
осмысленной: сходится не копия, а независимый расчёт.

Четыре области записаны в источнике с опечатками и вариантами написания.
Они заводятся алиасами вида `SOURCE_SPELLING`, а не «исправляются» в данных:
данные — свидетельство, и переписывать их импортом нельзя. Что именно было
написано в книге, обязано остаться видимым.
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook

from app.risk.core import IndicatorDirection
from app.risk.layers.budget import (
    BANDS,
    INDICATORS,
    BudgetRawIndicators,
    BudgetRowInputs,
    BudgetRowResult,
    NormalizationBand,
    evaluate_row,
)

LAYER_CODE: Final[str] = "8.3"
SOURCE_FILE_NAME: Final[str] = "Слой_8_3_бюджетных_рисков_КЗ_2025 (1).xlsx"

SHEET_PARAMS: Final[str] = "Параметры"
SHEET_MONTHLY: Final[str] = "Расчет_месяц"
SHEET_RAW: Final[str] = "RAW_DATA_Бюджет_все_регионы_КЗ_"
SHEET_FUNCTIONS: Final[str] = "Функции_расходов"

EXPECTED_SHEETS: Final[int] = 11
EXPECTED_MONTHLY_ROWS: Final[int] = 240
EXPECTED_RAW_ROWS: Final[int] = 74_831
EXPECTED_TERRITORIES: Final[int] = 20
EXPECTED_PERIODS: Final[int] = 12

# Написания областей, встреченные в книге и расходящиеся с нормой. Слева —
# как в источнике, справа — каноническое название. Заводятся алиасами с
# указанием слоя: «Западно-Казахстанкая» без «ск» — опечатка, «Мангыстауская»
# и «Туркистанская» — допустимые варианты транслитерации, но join по названию
# падает одинаково на всех четырёх.
SOURCE_SPELLINGS: Final[dict[str, str]] = {
    "Западно-Казахстанкая область": "Западно-Казахстанская область",
    "Мангыстауская область": "Мангистауская область",
    "Северо-Казахстанкая область": "Северо-Казахстанская область",
    "Туркистанская область": "Туркестанская область",
}

# Колонки листа «Расчет_месяц», 0-based. Заданы поимённо: обращение по числу
# в коде расчёта нечитаемо и переживает перестановку колонок молча.
_COL_GEO_LEVEL: Final[int] = 0
_COL_TERRITORY_ID: Final[int] = 1
_COL_PARENT_TERRITORY: Final[int] = 2
_COL_SOURCE_REGION: Final[int] = 3
_COL_TERRITORY_NAME: Final[int] = 4
_COL_MONTH: Final[int] = 5
_COL_PERIOD: Final[int] = 6
_COL_CLOSING_BALANCE: Final[int] = 29
_COL_MISSING_ROOTS: Final[int] = 35
_COL_RAW_FIRST: Final[int] = 45
_COL_BOOK_SCORE: Final[int] = 74
_COL_BOOK_LEVEL: Final[int] = 75
_COL_BOOK_RANK: Final[int] = 76
_COL_BOOK_COMPLETENESS: Final[int] = 80
_COL_BOOK_KEY: Final[int] = 81
_COL_BOOK_OVERRIDE: Final[int] = 82

# Колонки листа RAW_DATA, 0-based.
_RAW_COLUMNS: Final[tuple[str, ...]] = (
    "id", "code", "name", "utv", "utch", "plg", "plgp", "plgo", "sumrg",
    "obz", "obzsumrg", "plgpsumrg", "plgsumrg", "level", "isLeaf", "loaded",
    "expanded", "region", "period", "parentCode", "parent_id",
)


def _as_float(value: object) -> float:
    """Число из ячейки, где денежные поля смешаны int и float.

    В книге `utv` — 74 371 целых и 460 дробных, `plgp` — наоборот. Приводим
    явно: сравнение int и float в накопительных суммах даёт разные результаты
    округления.
    """
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, int | float):
        return float(value)
    text = str(value).replace(" ", "").replace(" ", "").replace(",", ".").strip()
    if not text:
        return 0.0
    return float(text)


def _as_int(value: object) -> int:
    """Целое из ячейки.

    Нужен отдельно от `_as_float`, потому что openpyxl отдаёт объединённый тип
    (число, строка, дата, формула), и прямой `int()` по нему не типизируется.
    Молча подставлять ноль здесь нельзя: эти поля — идентификаторы и уровни
    иерархии, и ноль в них означал бы существующую запись, а не пропуск.
    """
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int | float):
        return int(value)
    text = str(value).strip()
    if not text:
        raise ValueError("ожидалось целое, ячейка пуста")
    return int(float(text.replace(",", ".")))


def _as_int_or_none(value: object) -> int | None:
    """То же, но пустая ячейка — законное «значения нет», а не ошибка."""
    if value is None or not str(value).strip():
        return None
    return _as_int(value)


def parse_period(period: str) -> tuple[int, int]:
    """Разобрать период вида `01.2025` в пару «месяц, год».

    В книге все 74 831 значение — строки, а не даты. Лексикографическая
    сортировка по ним даёт правильный порядок только внутри одного года, и
    полагаться на неё нельзя.
    """
    month_text, _, year_text = period.strip().partition(".")
    return int(month_text), int(year_text)


def normalize_region_name(raw: str) -> str:
    """Каноническое название области по написанию из книги.

    Возвращает норму для сопоставления. Исходное написание при этом обязано
    сохраняться в записи отдельным полем — иначе исчезнет след того, что
    источник писал название иначе.
    """
    return SOURCE_SPELLINGS.get(raw.strip(), raw.strip())


@dataclass(frozen=True, slots=True)
class BookIndicatorParameter:
    """Строка листа «Параметры» — вес и пороги одного индикатора."""

    code: str
    name: str
    weight: float
    no_risk: float
    critical: float
    direction: IndicatorDirection
    unit: str

    @property
    def band(self) -> NormalizationBand:
        return NormalizationBand(self.no_risk, self.critical, self.direction)


@dataclass(frozen=True, slots=True)
class BudgetMonthlyRow:
    """Расчётная строка «область × месяц» вместе с эталоном книги.

    Значения книги (`book_*`) хранятся рядом с входными данными не для
    использования в расчёте, а для сверки. Как только импортёр начнёт
    подставлять их вместо собственного результата, проверка перестанет что-либо
    доказывать.
    """

    inputs: BudgetRowInputs
    source_region_name: str
    """Написание области ровно как в книге, включая опечатки."""

    geo_level: str
    parent_territory_id: str
    source_row_ref: str

    book_score: float
    book_level: str
    book_rank: int
    book_completeness: float
    book_override: bool
    book_key: str
    missing_roots_flag: int
    """Флаг «нет корневых строк» (колонка AJ) — источник 32 неполных строк."""

    @property
    def spelling_differs(self) -> bool:
        return self.source_region_name != self.inputs.territory_name


@dataclass(frozen=True, slots=True)
class BudgetRawFact:
    """Строка сырой бюджетной иерархии: статья × область × месяц."""

    source_id: int
    code: int | None
    name: str
    level: int
    is_leaf: bool
    parent_id: int | None
    parent_code: int | None

    region_source: str
    region_normalized: str
    period: str
    month: int
    year: int

    utv: float
    utch: float
    plg: float
    plgp: float
    plgo: float
    sumrg: float
    obz: float
    obzsumrg: float
    plgpsumrg: float
    plgsumrg: float


def resolve_workbook(source_dir: Path) -> Path:
    """Найти книгу 8.3 в каталоге источников устойчиво к NFD.

    Часть исходников пришла с macOS с именами в Unicode NFD, где «й» записана
    двумя кодовыми точками. Прямая конкатенация пути такие файлы не находит.
    """
    from scripts.source_manifest import resolve_source

    return resolve_source(source_dir, SOURCE_FILE_NAME)


def load_parameters(path: Path) -> dict[str, BookIndicatorParameter]:
    """Прочитать веса и пороги с листа «Параметры»."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[SHEET_PARAMS]
        parameters: dict[str, BookIndicatorParameter] = {}
        for row in sheet.iter_rows(min_row=5, max_row=19, values_only=True):
            code = row[0]
            if code is None:
                continue
            direction = (
                IndicatorDirection.LOWER_IS_RISKIER
                if str(row[5]).strip().upper() == "LOW"
                else IndicatorDirection.HIGHER_IS_RISKIER
            )
            parameters[str(code)] = BookIndicatorParameter(
                code=str(code),
                name=str(row[1]),
                weight=_as_float(row[2]),
                no_risk=_as_float(row[3]),
                critical=_as_float(row[4]),
                direction=direction,
                unit=str(row[6] or ""),
            )
        return parameters
    finally:
        workbook.close()


def check_parameters_match_model(path: Path) -> list[str]:
    """Сверить лист «Параметры» с константами модели.

    Возвращает список расхождений. Пустой список — книга и код описывают одну
    и ту же методику. Лист объявлен редактируемым, и если администратор
    поправит вес в книге, расчёт обязан об этом сообщить, а не разойтись тихо.
    """
    parameters = load_parameters(path)
    problems: list[str] = []

    for spec in INDICATORS:
        parameter = parameters.get(spec.code)
        if parameter is None:
            problems.append(f"{spec.code}: индикатор отсутствует в листе «Параметры»")
            continue
        if parameter.weight != spec.weight:
            problems.append(
                f"{spec.code}: вес в книге {parameter.weight:g}, в модели {spec.weight:g}"
            )
        band = BANDS[spec.code]
        if (parameter.no_risk, parameter.critical) != (band.no_risk, band.critical):
            problems.append(
                f"{spec.code}: пороги в книге {parameter.no_risk:g}/{parameter.critical:g}, "
                f"в модели {band.no_risk:g}/{band.critical:g}"
            )
        if parameter.direction is not band.direction:
            problems.append(
                f"{spec.code}: направление в книге {parameter.direction}, "
                f"в модели {band.direction}"
            )

    total = sum(p.weight for p in parameters.values())
    if total != 100.0:
        problems.append(f"сумма весов в книге {total:g}, ожидается 100")

    return problems


def load_monthly_rows(path: Path) -> list[BudgetMonthlyRow]:
    """Прочитать 240 расчётных строк «область × месяц»."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[SHEET_MONTHLY]
        rows: list[BudgetMonthlyRow] = []
        for offset, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if row[_COL_TERRITORY_ID] is None:
                continue
            rows.append(_build_monthly_row(row, excel_row=offset + 2))
        return rows
    finally:
        workbook.close()


def _build_monthly_row(row: tuple[Any, ...], *, excel_row: int) -> BudgetMonthlyRow:
    raw_values = [_as_float(row[_COL_RAW_FIRST + i]) for i in range(14)]
    raw = BudgetRawIndicators(
        r01_dohody_ispolnenie=raw_values[0],
        r02_zatraty_ispolnenie=raw_values[1],
        r04_intensivnost_utochneniy=raw_values[2],
        r05_oshibka_profilya=raw_values[3],
        r06_otklonenie_saldo=raw_values[4],
        r07_kassovyy_bufer=raw_values[5],
        r08_izbytochnye_ostatki=raw_values[6],
        r09_davlenie_ostatka=raw_values[7],
        r10_otstavanie_obyazatelstv=raw_values[8],
        r11_neoplachennye_obyazatelstva=raw_values[9],
        r12_shirina_nedoispolneniya=raw_values[10],
        r13_hhi=raw_values[11],
        r14_finansovye_operatsii=raw_values[12],
        r15_flagi_kachestva=raw_values[13],
    )
    source_region = str(row[_COL_SOURCE_REGION])
    inputs = BudgetRowInputs(
        territory_id=str(row[_COL_TERRITORY_ID]),
        territory_name=normalize_region_name(str(row[_COL_TERRITORY_NAME])),
        month=int(_as_float(row[_COL_MONTH])),
        period=str(row[_COL_PERIOD]),
        raw=raw,
        closing_balance=_as_float(row[_COL_CLOSING_BALANCE]),
    )
    return BudgetMonthlyRow(
        inputs=inputs,
        source_region_name=source_region,
        geo_level=str(row[_COL_GEO_LEVEL]),
        parent_territory_id=str(row[_COL_PARENT_TERRITORY]),
        source_row_ref=f"{SHEET_MONTHLY}!A{excel_row}",
        book_score=_as_float(row[_COL_BOOK_SCORE]),
        book_level=str(row[_COL_BOOK_LEVEL]),
        book_rank=int(_as_float(row[_COL_BOOK_RANK])),
        book_completeness=_as_float(row[_COL_BOOK_COMPLETENESS]),
        book_override=bool(_as_float(row[_COL_BOOK_OVERRIDE])),
        book_key=str(row[_COL_BOOK_KEY]),
        missing_roots_flag=int(_as_float(row[_COL_MISSING_ROOTS])),
    )


def evaluate_rows(rows: list[BudgetMonthlyRow]) -> list[BudgetRowResult]:
    """Посчитать риск по всем строкам собственным расчётом."""
    return [evaluate_row(row.inputs) for row in rows]


def iter_raw_facts(path: Path) -> Iterator[BudgetRawFact]:
    """Построчно читать сырую бюджетную иерархию (74 831 строка).

    Генератор, а не список: лист занимает большую часть десятимегабайтной
    книги, и материализовать его целиком незачем — импорт идёт партиями.
    """
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[SHEET_RAW]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            period = str(row[18])
            month, year = parse_period(period)
            region_source = str(row[17])
            yield BudgetRawFact(
                source_id=_as_int(row[0]),
                code=_as_int_or_none(row[1]),
                name=str(row[2]),
                level=_as_int(row[13]),
                is_leaf=bool(row[14]),
                parent_id=_as_int_or_none(row[20]),
                parent_code=_as_int_or_none(row[19]),
                region_source=region_source,
                region_normalized=normalize_region_name(region_source),
                period=period,
                month=month,
                year=year,
                utv=_as_float(row[3]),
                utch=_as_float(row[4]),
                plg=_as_float(row[5]),
                plgp=_as_float(row[6]),
                plgo=_as_float(row[7]),
                sumrg=_as_float(row[8]),
                obz=_as_float(row[9]),
                obzsumrg=_as_float(row[10]),
                plgpsumrg=_as_float(row[11]),
                plgsumrg=_as_float(row[12]),
            )
    finally:
        workbook.close()


def territory_aliases(rows: list[BudgetMonthlyRow]) -> list[tuple[str, str, str]]:
    """Алиасы «написание из книги → territory_id → вид алиаса».

    Возвращаются все встреченные написания, а не только расходящиеся: алиас
    совпадающего названия тоже нужен, иначе резолвер не найдёт территорию по
    каноническому имени.
    """
    seen: dict[tuple[str, str], str] = {}
    for row in rows:
        territory_id = row.inputs.territory_id
        seen[(row.source_region_name, territory_id)] = (
            "source_spelling" if row.spelling_differs else "official"
        )
        seen[(row.inputs.territory_name, territory_id)] = "official"
    return [(alias, code, kind) for (alias, code), kind in sorted(seen.items())]


__all__ = [
    "EXPECTED_MONTHLY_ROWS",
    "EXPECTED_PERIODS",
    "EXPECTED_RAW_ROWS",
    "EXPECTED_SHEETS",
    "EXPECTED_TERRITORIES",
    "LAYER_CODE",
    "SHEET_MONTHLY",
    "SHEET_PARAMS",
    "SHEET_RAW",
    "SOURCE_FILE_NAME",
    "SOURCE_SPELLINGS",
    "BookIndicatorParameter",
    "BudgetMonthlyRow",
    "BudgetRawFact",
    "check_parameters_match_model",
    "evaluate_rows",
    "iter_raw_facts",
    "load_monthly_rows",
    "load_parameters",
    "normalize_region_name",
    "parse_period",
    "resolve_workbook",
    "territory_aliases",
]
